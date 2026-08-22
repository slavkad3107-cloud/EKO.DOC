"""Форма № 2-ТП (отходы) — сведения об образовании, обработке, утилизации,
обезвреживании, размещении отходов производства и потребления (слова
«транспортировании» в названии действующей формы нет). Годовая: ТО
Росприроднадзора — 1 февраля (за 2025 — до 02.02.2026), Росприроднадзору —
15 марта; подача через ЛКПП РПН. Код формы по ОКУД 0609013.

ДЕЙСТВУЮЩАЯ редакция — Приказ Росстата от 06.11.2025 № 614 (начиная с отчёта
за 2025); приказы № 627 от 09.10.2020 и № 698 от 13.11.2020 утратили силу.

Структура: стр.1 (титул по бланку № 614: три стандартные надписи, блок
«Предоставляют / Сроки предоставления / Форма / Приказ / Годовая», кодовая
часть); стр.2 — Раздел I (движение отходов, 29 граф, многоуровневая шапка
дословно по бланку); стр.3 — Раздел II (регоператоры ТКО: собственные 29 граф
тремя подтаблицами гр.1–9 / 10–17 / 18–29 с шапкой А|Б|В|Г, данные из
extra.tko_operators; если организация не регоператор — шапка печатается
целиком, как в принятых отчётах, с пояснением под ней по п. 14 Указаний) +
Раздел III (фиксированные строки 11–31, значения из
extra.disposal_objects_s3) + подпись по бланку.

Раздел I: только виды отходов (без итоговых строк — их в бланке нет),
последовательно с I по V класс опасности (п. 7 Указаний), внутри класса — по
коду ФККО, нумерация строк с 1 (п. 13). Код ФККО в графе В печатается в
каноническом виде с пробелами («7 33 100 01 72 4»), как в печати ЛКПП; в XML
(WST_CODE) — без пробелов. Точность массы — по классу опасности: I–III =
3 знака, IV–V = 1 знак (п. 8). Баланс гр.29 (п. 13) проверяется.
Графа 14 (ТКО региональному оператору) — только по получателю-регоператору;
гр.27/28 — размещение на хранение/захоронение раздельно (placed_storage /
placed_burial). XML — конверт «Модуля природопользователя» (DATA_PACKET_NI,
DocType=3); XSD Модуля за 2025 не сверен — схем в открытом доступе нет.

Данные — из ctx.wastes (та же модель, что и учёт движения №1028).
Эталон формулировок — принятый отчёт ИП Миних за 2025 (печать ЛКПП).
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from ecodoc.core import fkko as _fkko
from ecodoc.core.models import Issue
from ecodoc.core.money import D
from ecodoc.core.registry import register
from ecodoc.render import xlsx
from ecodoc.render.xmlutil import _is_nvos_code as _is_nvos, data_packet_ni, el, write_tree
from ecodoc.reports.base import Report


# Раздел III формы (приказ Росстата № 614): фиксированные строки 11–31.
# Значения подставляются из ctx.extra['disposal_objects_s3'] = {"11": 2, …};
# незаполненные печатаются прочерком, как в принятых отчётах. Формулировки —
# дословно по печатной форме ЛКПП (в бланке «из них ТКО», не расшифровка).
_S3_ROWS = [
    (11, "Количество эксплуатируемых респондентом объектов захоронения отходов, ед"),
    (12, "из них ТКО, ед"),
    (13, "Количество эксплуатируемых респондентом объектов хранения отходов, ед"),
    (14, "Количество эксплуатируемых респондентом объектов захоронения отходов, "
         "отвечающих установленным требованиям, ед"),
    (15, "из них ТКО, ед"),
    (16, "Количество эксплуатируемых респондентом объектов хранения отходов, "
         "отвечающих установленным требованиям, ед"),
    (17, "Вместимость эксплуатируемых респондентом объектов захоронения отходов "
         "согласно проектной документации, т"),
    (18, "из них ТКО, т"),
    (19, "Остаточная вместимость эксплуатируемых респондентом объектов "
         "захоронения отходов, т"),
    (20, "из них ТКО, т"),
    (21, "Вместимость эксплуатируемых респондентом объектов захоронения отходов "
         "согласно проектной документации, м3"),
    (22, "из них ТКО, м3"),
    (23, "Остаточная вместимость эксплуатируемых респондентом объектов "
         "захоронения отходов, м3"),
    (24, "из них ТКО, м3"),
    (25, "Вместимость эксплуатируемых респондентом объектов хранения отходов "
         "согласно проектной документации, т"),
    (26, "Остаточная вместимость эксплуатируемых респондентом объектов "
         "хранения отходов, т"),
    (27, "Вместимость эксплуатируемых респондентом объектов хранения отходов "
         "согласно проектной документации, м3"),
    (28, "Остаточная вместимость эксплуатируемых респондентом объектов "
         "хранения отходов, м3"),
    (29, "Площадь, занимаемая эксплуатируемыми респондентом объектами "
         "захоронения отходов, га"),
    (30, "из них ТКО, га"),
    (31, "Площадь, занимаемая эксплуатируемыми респондентом объектами "
         "хранения отходов, га"),
]

# Заголовки разделов — дословно по печатной форме ЛКПП (приказ № 614).
_S1_TITLE = ("Раздел I. Сведения об образовании, обработке, утилизации, "
             "обезвреживании, размещении отходов производства и потребления; "
             "сведения об образовании и передаче твердых коммунальных отходов "
             "региональному оператору, тонна")
_S2_TITLE = ("Раздел II. Сведения об образовании, обработке, утилизации, "
             "обезвреживании, размещении отходов производства и потребления, "
             "представляемые региональными операторами, осуществляющими "
             "деятельность с твердыми коммунальными отходами, тонна")

# Служебные графы А|Б|В|Г — одинаковы в Разделах I и II (бланк № 614).
_SERVICE = [("А", "N строки"), ("Б", "Наименование видов отходов"),
            ("В", "Код отхода по федеральному классификационному каталогу отходов"),
            ("Г", "Класс опасности отхода")]

# --- Многоуровневая шапка Раздела I: для каждой графы — «путь» подписей от
# верхнего уровня к нижнему, дословно по бланку (эталон МИНИХ, таблицы 8 и 10).
# Почему путь, а не плоская подпись: госорган сверяет заголовки граф с бланком,
# а в бланке графы сгруппированы («Поступление … из других хозяйствующих
# субъектов» → «всего» / «из графы 3» → «из других субъектов РФ» / «по импорту
# …»). Соседние графы с одинаковым началом пути объединяются (merge), а
# последний элемент пути растягивается вниз до нижней строки шапки.
_P_IN_OTHER = "Поступление отходов из других хозяйствующих субъектов"
_P_IN_OWN = "Поступление отходов с собственных объектов"
_P_UTIL = "Утилизировано отходов"
_P_TR = "Передача отходов (за исключением ТКО) другим хозяйствующим субъектам"
_P_TR_OWN = "Передача отходов (за исключением ТКО) на собственные объекты"
_P_PLACE = "Размещение отходов на эксплуатируемых объектах за отчетный год"
_OTHER_RF = "из них в другие субъекты РФ"
_G29 = [
    (1, ["Наличие отходов на начало отчетного года"]),
    (2, ["Образование отходов за отчетный год"]),
    (3, [_P_IN_OTHER, "всего"]),
    (4, [_P_IN_OTHER, "из графы 3", "из других субъектов РФ"]),
    (5, [_P_IN_OTHER, "из графы 3", "по импорту из других государств"]),
    (6, [_P_IN_OWN, "всего"]),
    (7, [_P_IN_OWN, "из них из других субъектов РФ"]),
    (8, ["Образование других видов отходов после обработки за отчетный год"]),
    (9, ["Обработано отходов"]),
    (10, [_P_UTIL, "всего"]),
    (11, [_P_UTIL, "из графы 10", "для повторного применения (рециклинг)"]),
    (12, [_P_UTIL, "из графы 10", "предварительно прошедших обработку"]),
    (13, ["Обезврежено отходов"]),
    (14, ["Передача ТКО региональному оператору"]),
    (15, [_P_TR, "для обработки", "всего передано для обработки"]),
    (16, [_P_TR, "для обработки", _OTHER_RF]),
    (17, [_P_TR, "для утилизации", "всего передано для утилизации"]),
    (18, [_P_TR, "для утилизации", _OTHER_RF]),
    (19, [_P_TR, "для обезвреживания", "всего передано для обезвреживания"]),
    (20, [_P_TR, "для обезвреживания", _OTHER_RF]),
    (21, [_P_TR, "для хранения", "всего передано для хранения"]),
    (22, [_P_TR, "для хранения", _OTHER_RF]),
    (23, [_P_TR, "для захоронения", "всего передано для захоронения"]),
    (24, [_P_TR, "для захоронения", _OTHER_RF]),
    (25, [_P_TR_OWN, "всего"]),
    (26, [_P_TR_OWN, _OTHER_RF]),
    (27, [_P_PLACE, "хранение"]),
    (28, [_P_PLACE, "захоронение"]),
    (29, ["Наличие отходов на конец отчетного года"]),
]

# --- 29 граф Раздела II — пути подписей дословно по бланку (эталон МИНИХ,
# таблицы 12–14). У Раздела II графы СВОИ — ТКО-показатели регоператора.
_P2_IN = ("Поступление ТКО к региональному оператору от других хозяйствующих "
          "субъектов, населения и субъектов РФ")
_P2_AFTER = ("Образование ТКО после обработки за отчетный год "
             "(отходы после обработки ТКО)")
_P2_TR = "Передача ТКО региональным оператором другим операторам"
_P2_TR_LONG = (_P2_TR + " (передача отходов после обработки ТКО другим "
               "операторам)")
_P2_BURIAL = "Захоронение ТКО на эксплуатируемых объектах за отчетный год"
_OTHER_RF_TKO = ("из них ТКО, переданных хозяйствующим субъектам (операторам), "
                 "осуществляющим деятельность в других субъектах Российской "
                 "Федерации")
_G29_II = [
    (1, ["Наличие ТКО на начало отчетного года"]),
    (2, ["Образование ТКО за отчетный год"]),
    (3, [_P2_IN, "всего ТКО"]),
    (4, [_P2_IN, "из графы 3", "ТКО, образованных в жилых помещениях в субъекте РФ"]),
    (5, [_P2_IN, "из графы 3", "ТКО, образованных в других субъектах РФ (по соглашению)"]),
    (6, [_P2_AFTER, "всего"]),
    (7, [_P2_AFTER, "из графы 6", "на объектах обработки регионального оператора"]),
    (8, [_P2_AFTER, "из графы 6", "на объектах оператора, осуществляющего обработку "
         "ТКО, передающего их после обработки региональному оператору"]),
    (9, [_P2_AFTER, "из графы 6", "на объектах оператора, осуществляющего обработку "
         "ТКО, не передающего их после обработки региональному оператору"]),
    (10, ["Обработано ТКО", "всего ТКО"]),
    (11, ["Обработано ТКО", "из них ТКО, образованных в жилых помещениях"]),
    (12, ["Утилизировано ТКО", "всего ТКО"]),
    (13, ["Утилизировано ТКО", "из графы 12", "для повторного применения (рециклинг)"]),
    (14, ["Утилизировано ТКО", "из графы 12", "энергетическая утилизация"]),
    (15, ["Обезврежено ТКО"]),
    (16, [_P2_TR, "для обработки", "всего ТКО"]),
    (17, [_P2_TR, "для обработки", _OTHER_RF_TKO]),
    (18, [_P2_TR_LONG, "для утилизации", "всего ТКО"]),
    (19, [_P2_TR_LONG, "для утилизации", "из графы 18",
          "ТКО, переданных хозяйствующим субъектам (операторам), осуществляющим "
          "деятельность в других субъектах Российской Федерации"]),
    (20, [_P2_TR_LONG, "для утилизации", "из графы 18", "на энергетическую утилизацию",
          "всего ТКО"]),
    (21, [_P2_TR_LONG, "для утилизации", "из графы 18", "на энергетическую утилизацию",
          "из них ТКО переданных в другие субъекты Российской Федерации на "
          "энергетическую утилизацию"]),
    (22, [_P2_TR_LONG, "для обезвреживания", "всего ТКО"]),
    (23, [_P2_TR_LONG, "для обезвреживания", _OTHER_RF_TKO]),
    (24, [_P2_TR_LONG, "для захоронения", "всего ТКО"]),
    (25, [_P2_TR_LONG, "для захоронения", _OTHER_RF_TKO]),
    (26, ["Хранение отходов после обработки ТКО"]),
    (27, [_P2_BURIAL, "всего"]),
    (28, [_P2_BURIAL, "из них ТКО, образованных в жилых помещениях"]),
    (29, ["Наличие ТКО на конец отчетного года"]),
]

# разбиение граф Раздела II на три подтаблицы — как в печатной форме
# (иначе 29 граф не помещаются на лист)
_S2_SPLITS = [(1, 9), (10, 17), (18, 29)]

# обратная совместимость: старые ключи extra['tko_operators'] → номер графы
_S2_LEGACY = {3: "received", 10: "processed", 27: "placed"}

# Перечень респондентов и сроки — дословно по бланку № 614 (таблица 5 эталона).
_RESPONDENTS = (
    "юридические лица, физические лица, занимающиеся предпринимательской "
    "деятельностью без образования юридического лица (индивидуальные "
    "предприниматели), осуществляющие деятельность в области обращения с "
    "отходами производства и потребления, региональные операторы по обращению "
    "с твердыми коммунальными отходами, операторы по обращению с твердыми "
    "коммунальными отходами:")
_NOTE_CONFID = "КОНФИДЕНЦИАЛЬНОСТЬ ГАРАНТИРУЕТСЯ ПОЛУЧАТЕЛЕМ ИНФОРМАЦИИ"
_NOTE_KOAP = ("Нарушение порядка предоставления первичных статистических данных "
              "или несвоевременное предоставление этих данных, либо предоставление "
              "недостоверных первичных статистических данных влечет "
              "ответственность, установленную Кодексом Российской Федерации об "
              "административных правонарушениях")
_NOTE_152FZ = ("В соответствии с пунктом 9 части 1 статьи 6 Федерального закона от "
               "27 июля 2006 г. N 152-ФЗ \"О персональных данных\" обработка "
               "персональных данных осуществляется для статистических целей при "
               "условии обязательного обезличивания персональных данных")
_SIGN_TITLE = ("Должностное лицо, ответственное за предоставление первичных "
               "статистических данных (лицо, уполномоченное предоставлять "
               "первичные статистические данные от имени юридического лица или "
               "от имени физического лица, занимающегося предпринимательской "
               "деятельностью без образования юридического лица)")
# пояснение под Разделом II — круг заполняющих по п. 14 Указаний к № 614
_S2_NOTE = ("— раздел не заполняется: по п. 14 Указаний его заполняют "
            "региональные операторы по обращению с ТКО; операторы по обращению "
            "с ТКО в субъекте РФ; юридические лица (ИП), эксплуатирующие "
            "собственные объекты размещения ТКО; операторы, не передающие "
            "отходы после обработки ТКО региональному оператору — респондент "
            "к ним не относится")


def _header_tree(ws, top_row: int, first_col: int, cols: list, depth: int,
                 size: int = 7):
    """Напечатать многоуровневую шапку граф по «путям» подписей.

    cols — [(номер_графы, [подпись_уровня_1, …])]; строки top_row..top_row+depth-1
    — уровни, строка top_row+depth — номера граф. Соседние графы с одинаковым
    началом пути объединяются по горизонтали; последний элемент пути тянется
    вниз до последнего уровня (так бланк рисует «всего» рядом с «из графы 3»)."""
    from openpyxl.utils import get_column_letter
    n = len(cols)
    done = [[False] * depth for _ in range(n)]
    for i, (num, path) in enumerate(cols):
        col = get_column_letter(first_col + i)
        xlsx.cell(ws, f"{col}{top_row + depth}", num, italic=True, size=8)
        for lvl, label in enumerate(path):
            if done[i][lvl]:
                continue
            last = lvl == len(path) - 1
            # ширина: соседи справа с тем же префиксом пути (и тем же уровнем)
            j = i
            while (j + 1 < n and len(cols[j + 1][1]) > lvl
                   and cols[j + 1][1][:lvl + 1] == path[:lvl + 1]):
                j += 1
            bottom = top_row + (depth - 1 if last else lvl)
            for k in range(i, j + 1):
                for lv in range(lvl, bottom - top_row + 1):
                    done[k][lv] = True
            rng = (f"{col}{top_row + lvl}:"
                   f"{get_column_letter(first_col + j)}{bottom}")
            if rng.split(":")[0] == rng.split(":")[1]:
                xlsx.cell(ws, rng.split(":")[0], label, bold=True, fill=True, size=size)
            else:
                xlsx.merge(ws, rng, label, bold=True, fill=True, size=size)


def _service_header(ws, top_row: int, depth: int):
    """Графы А|Б|В|Г: подпись на всю высоту шапки + буква в строке номеров."""
    from openpyxl.utils import get_column_letter
    for i, (g, lbl) in enumerate(_SERVICE):
        col = get_column_letter(i + 1)
        if depth > 1:
            xlsx.merge(ws, f"{col}{top_row}:{col}{top_row + depth - 1}", lbl,
                       bold=True, fill=True, size=8)
        else:
            xlsx.cell(ws, f"{col}{top_row}", lbl, bold=True, fill=True, size=8)
        xlsx.cell(ws, f"{col}{top_row + depth}", g, italic=True, size=8)


def _s2_val(op: dict, num: int):
    """Значение графы Раздела II: op['gN'], иначе старый ключ, иначе «-».

    Ничего не выдумываем: для графы без источника печатается прочерк,
    как в принятых отчётах."""
    v = op.get(f"g{num}")
    if v in (None, ""):
        v = op.get(_S2_LEGACY.get(num, ""), "")
    return v if v not in (None, "") else "-"


def _n(v) -> float:
    return float(D(v))


def _f6(v) -> str:
    return f"{_n(v):.6f}"


def _prec(hazard_class) -> int:
    """Точность массы по Указаниям: I–III класс — 3 знака, IV–V — 1 знак."""
    try:
        return 3 if int(hazard_class) <= 3 else 1
    except (TypeError, ValueError):
        return 1


def _fmt_class(v, hazard_class) -> str:
    return f"{_n(v):.{_prec(hazard_class)}f}"


def _round_class(v, hazard_class) -> float:
    return round(_n(v), _prec(hazard_class))


def _is_tko(fkko) -> bool:
    """ТКО — блок ФККО «7 3…» (передаётся региональному оператору, графа 14).
    Единое определение для всех форм — core/waste_agg.is_tko."""
    from ecodoc.core.waste_agg import is_tko
    return is_tko(fkko)


def _by_kind(w) -> D:
    """Сумма передачи по видам (гр.15+17+19+21+23)."""
    return (D(w.transferred_processing) + D(w.transferred_util)
            + D(w.transferred_neutral) + D(w.transferred_storage)
            + D(w.transferred_burial))


def _placed(w) -> tuple[D, D, bool]:
    """(на хранение — гр.27, на захоронение — гр.28, fallback).

    Обратная совместимость: в старых базах заполнены только placed_norm/over
    (в пределах/сверх лимита — это для платы, не вид размещения); тогда всё
    размещённое считаем захоронением, как раньше, и помечаем fallback=True,
    чтобы validate() предупредил."""
    st, bu = D(w.placed_storage), D(w.placed_burial)
    if st == 0 and bu == 0:
        legacy = D(w.placed_norm) + D(w.placed_over)
        if legacy > 0:
            return D(0), legacy, True
    return st, bu, False


def _sort_key(w):
    """Порядок строк Раздела I по п. 7 Указаний: с I по V класс, внутри
    класса — по коду ФККО (неизвестный класс — в конец)."""
    try:
        hc = int(w.hazard_class)
    except (TypeError, ValueError):
        hc = 9
    return (hc, _fkko.norm(w.fkko_code) or "9" * 11, (w.name or "").lower())


@register
class TP2Waste(Report):
    code = "2tp-waste"
    title = "2-ТП (отходы)"

    # --- получатели по ФККО (для графы 14: ТКО региональному оператору) ---
    def _receivers(self, fkko) -> list[dict]:
        e = self.ctx.extra if isinstance(self.ctx.extra, dict) else {}
        code = _fkko.norm(fkko)
        return [r for r in (e.get("waste_receivers") or [])
                if isinstance(r, dict) and _fkko.norm(r.get("fkko")) == code]

    def _regop_mass(self, w):
        """Масса ТКО, переданная региональному оператору (по получателям из
        актов/справок с признаком regional_operator). None — по этому отходу
        получатели вообще не известны (ручной ввод без актов)."""
        recs = self._receivers(w.fkko_code)
        if not recs:
            return None
        return sum((D(r.get("mass") or 0) for r in recs
                    if r.get("regional_operator")), D(0))

    def _g14(self, w) -> D:
        """Графа 14 — ТКО региональному оператору (п. 13 Указаний).

        По актам масса регоператору не раскладывается по видам, поэтому
        гр.14 = transferred − сумма по видам. Ручной ввод без получателей
        (как в принятых отчётах: всё ТКО — регоператору) даёт то же самое.
        Если получатели известны и среди них регоператора нет — гр.14 = 0,
        а нераспознанный остаток ловит validate()."""
        if not _is_tko(w.fkko_code):
            return D(0)
        rest = D(w.transferred) - _by_kind(w)
        regop = self._regop_mass(w)
        if regop is not None and regop == 0:
            return D(0)
        return max(rest, D(0))

    def _sorted_wastes(self):
        return sorted(self.ctx.wastes, key=_sort_key)

    def validate(self) -> list[Issue]:
        from ecodoc.core.validators import inn_valid, ogrn_valid
        issues: list[Issue] = []
        o = self.ctx.organization
        e = self.ctx.extra if isinstance(self.ctx.extra, dict) else {}
        if not o.inn:
            issues.append(Issue("error", "ИНН", "не указан ИНН"))
        elif not inn_valid(o.inn):
            issues.append(Issue("error", "ИНН", f"ИНН {o.inn} не проходит проверку — опечатка"))
        if not o.okpo:
            issues.append(Issue("warning", "ОКПО", "для 2-ТП обязателен код ОКПО"))
        # --- preflight под загрузку в ЛКПП РПН ---
        if not o.ogrn:
            kind = "ОГРНИП" if o.is_individual else "ОГРН"
            issues.append(Issue("error", kind,
                                f"не указан {kind} — обязателен для сверки с ЕГРЮЛ/ЕГРИП "
                                "при загрузке в ЛКПП"))
        elif not ogrn_valid(o.ogrn):
            issues.append(Issue("warning", "ОГРН", f"ОГРН {o.ogrn} не проходит проверку — сверьте"))
        okt = (o.oktmo or "").strip()
        if not okt:
            issues.append(Issue("error", "ОКТМО", "не указан ОКТМО территории (Код по ОКТМО)"))
        elif len(okt) not in (8, 11) or not okt.isdigit():
            issues.append(Issue("warning", "ОКТМО",
                                f"ОКТМО «{okt}» — ожидается 8 или 11 цифр; проверьте, что это "
                                "именно ОКТМО, а не ОКАТО"))
        if not any(_is_nvos(ob.code) for ob in self.ctx.objects):
            issues.append(Issue("warning", "объект",
                                "нет кода объекта НВОС (напр. 41-0247-000123-П) — если объект "
                                "на учёте, укажите код, иначе РПН отклонит «объект не стоит на учёте»"))
        # предупреждения агрегации актов (нераспознанное назначение передачи)
        for msg in e.get("waste_agg_warnings") or []:
            issues.append(Issue("warning", "акты", str(msg)))
        eps = D("0.0005")
        for w in self.ctx.wastes:
            code = w.fkko_code
            # передано всего = гр.14 + сумма по видам (гр.15/17/19/21/23) —
            # иначе масса «передана в никуда» и гр.29 не сойдётся (п. 13)
            rest = D(w.transferred) - _by_kind(w)
            if rest < -eps:
                issues.append(Issue(
                    "error", f"передача/{code}",
                    f"сумма передачи по видам {_by_kind(w)} больше «передано всего» "
                    f"{D(w.transferred)}"))
            elif rest > eps and self._g14(w) < rest - eps:
                issues.append(Issue(
                    "error", f"передача/{code}",
                    f"transferred ≠ сумма по видам передачи: не указано назначение "
                    f"передачи по акту — {rest} т не разнесено по гр.15–23 "
                    "(утилизация/обезвреживание/захоронение/хранение/обработка)"
                    + (" или гр.14 (региональному оператору)" if _is_tko(code) else "")))
            st, bu, fallback = _placed(w)
            if fallback:
                issues.append(Issue(
                    "warning", f"размещение/{code}",
                    f"вид размещения не указан (placed_storage/placed_burial = 0): "
                    f"{bu} т (placed_norm+placed_over) показано как захоронение "
                    "(гр.28); если это хранение на собственном объекте — "
                    "заполните placed_storage"))
            # баланс гр.29 по п. 13 Указаний: начало + образовано + поступило −
            # обработано − утилизировано − обезврежено − размещено − передано
            # (поступление с собственных объектов / образование после обработки
            # в модели нет — гр.6/гр.8 печатаются нулями)
            bal = (D(w.accumulated_start) + D(w.accumulated_start_nakopl)
                   + D(w.generated) + D(w.received)
                   - D(w.processed) - D(w.used) - D(w.neutralized)
                   - D(w.transferred) - st - bu)
            if abs(bal - D(w.accumulated_end)) > D("0.001"):
                issues.append(Issue(
                    "warning", f"баланс/{code}",
                    f"наличие на конец {D(w.accumulated_end)} ≠ расчётный баланс "
                    f"{bal} (гр.29 = приход − расход)"))
        if not self.ctx.period.year:
            issues.append(Issue("error", "период", "не указан отчётный год"))
        if not self.ctx.wastes:
            issues.append(Issue("error", "отходы", "нет позиций отходов"))
        return issues

    # ---------------- XML (Модуль природопользователя) ----------------
    def render_xml(self, out_path: Path) -> Path:
        out_path = self._ensure_dir(out_path)
        exp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        root = data_packet_ni(self.ctx, doc_type=3, body_fn=self._body, exp_date=exp)
        write_tree(root, out_path)
        return out_path

    def _body(self, org, obj, exp_date):
        o = self.ctx.organization
        e = self.ctx.extra if isinstance(self.ctx.extra, dict) else {}
        ob, eo_id = obj if obj else (None, "")
        okato = (ob.oktmo if ob else "") or o.oktmo
        rpt = el(org, "RPT_2TP_WASTE")
        el(rpt, "DOC_YEAR", self.ctx.period.year)
        el(rpt, "RPN_CODE", str(e.get("rpn_to", "") or "1"))
        el(rpt, "ROSPRIR", e.get("rospr", ""))
        el(rpt, "FNAME", o.name)
        el(rpt, "SNAME", o.short_name or o.name)
        el(rpt, "ADDR_STR", o.address)
        el(rpt, "OGRN", o.ogrn)
        el(rpt, "INN", o.inn)
        el(rpt, "KPP", o.kpp)
        el(rpt, "OKPO", o.okpo)
        el(rpt, "OFFICIAL", o.official_title)
        el(rpt, "FIO_OFFICIAL", o.director_name)
        el(rpt, "CREATE_DATE", exp_date[:10])
        el(rpt, "OKATO", o.oktmo)
        el(rpt, "RPT_OKATO", okato)
        el(rpt, "ID_EO", eo_id)
        codes = [c.strip() for c in (o.okved or "").replace(";", ",").split(",") if c.strip()]
        if codes:
            okv = el(rpt, "OKVED")
            el(okv, "OKVED_CODE", codes[0])
        # тот же порядок, что и в печати: с I по V класс, затем по коду ФККО
        for w in self._sorted_wastes():
            hc = w.hazard_class
            st, bu, _ = _placed(w)
            fact = el(rpt, "RPT_2TP_WASTE_FACT")
            el(fact, "NONE_FKKO_NAME", w.name)
            el(fact, "WST_CODE", _fkko.norm(w.fkko_code) or str(w.fkko_code))
            el(fact, "WSTYPE", hc)
            # наличие на начало = хранение + накопление
            el(fact, "TP2_BP_ACCUM_WASTE",
               _fmt_class(D(w.accumulated_start) + D(w.accumulated_start_nakopl), hc))
            el(fact, "TP2_FORMING", _fmt_class(w.generated, hc))
            el(fact, "TP2_ARRIVAL", _fmt_class(w.received, hc))
            el(fact, "TP2_TRANSF", _fmt_class(w.transferred, hc))
            # TR_* = ПЕРЕДАНО другим (для утилизации/обезвреживания/захоронения),
            # а не собственные утилизация/обезвреживание
            el(fact, "TP2_TR_ISPOTX", _fmt_class(w.transferred_util, hc))
            el(fact, "TP2_TR_SOTX", _fmt_class(w.transferred_neutral, hc))
            el(fact, "TP2_TR_DISP", _fmt_class(w.transferred_burial, hc))
            # размещение на собственных объектах: захоронение (гр.28) и
            # хранение (гр.27) — раздельно, из данных, а не «0.0» константой
            el(fact, "TP2_RAZM", _fmt_class(bu, hc))
            el(fact, "TP2_RAZM_STOR", _fmt_class(st, hc))
            el(fact, "TP2_ACCUM_WASTE", _fmt_class(w.accumulated_end, hc))

    # ---------------- Печатная форма (3 страницы) ---------------------
    def render_print(self, out_path: Path) -> Path:
        out_path = self._ensure_dir(out_path)
        wb = xlsx.new_workbook()
        self._page1(wb)
        self._page2(wb)
        self._page3(wb)
        return xlsx.save(wb, out_path)

    def _page1(self, wb):
        """Титул — по бланку № 614 (эталон МИНИХ, таблицы 0–7)."""
        ws = wb.create_sheet("стр.1")
        o = self.ctx.organization
        year = self.ctx.period.year or ""
        xlsx.widths(ws, {"A": 26, "B": 26, "C": 18, "D": 22, "E": 18, "F": 24})
        # три стандартные надписи бланка — без них титул не по форме
        xlsx.merge(ws, "A1:F1", "ФЕДЕРАЛЬНОЕ СТАТИСТИЧЕСКОЕ НАБЛЮДЕНИЕ", bold=True)
        xlsx.merge(ws, "A2:F2", _NOTE_CONFID, bold=True)
        xlsx.merge(ws, "A3:F3", _NOTE_KOAP, size=9)
        xlsx.merge(ws, "A4:F4", _NOTE_152FZ, size=9)
        # название формы — по действующему приказу Росстата № 614: слова
        # «транспортировании» в нём нет (оно осталось от прежней редакции)
        xlsx.merge(ws, "A6:F7", "СВЕДЕНИЯ ОБ ОБРАЗОВАНИИ, ОБРАБОТКЕ, УТИЛИЗАЦИИ, "
                   "ОБЕЗВРЕЖИВАНИИ, РАЗМЕЩЕНИИ ОТХОДОВ ПРОИЗВОДСТВА И ПОТРЕБЛЕНИЯ "
                   f"за {year} г.", bold=True, border=False)
        # блок «Предоставляют / Сроки предоставления / Форма / Приказ / Годовая»
        # — отдельными ячейками, как таблица 5 бланка
        xlsx.merge(ws, "A9:C9", "Предоставляют:", bold=True, align="left")
        xlsx.cell(ws, "D9", "Сроки предоставления", bold=True)
        xlsx.merge(ws, "E9:F9", "Форма N 2-ТП (отходы)", bold=True)
        xlsx.merge(ws, "A10:C10", _RESPONDENTS, align="left", size=9)
        xlsx.cell(ws, "D10", "")
        xlsx.merge(ws, "E10:F11", "Приказ Росстата:\nОб утверждении формы\n"
                   "от 06.11.2025 N 614\nО внесении изменений (при наличии)\n"
                   "от ______ N ___\nот ______ N ___", size=9, align="left")
        xlsx.merge(ws, "A11:C11", "- территориальному органу Росприроднадзора в "
                   "субъекте Российской Федерации", align="left", size=9)
        xlsx.cell(ws, "D11", "1 февраля после отчетного периода", size=9)
        xlsx.merge(ws, "A12:C12", "- Росприроднадзору", align="left", size=9)
        xlsx.cell(ws, "D12", "15 марта после отчетного периода", size=9)
        xlsx.merge(ws, "E12:F12", "Годовая", bold=True)
        xlsx.heights(ws, {3: 40, 4: 40, 6: 24, 7: 24, 10: 96, 11: 30})
        # реквизиты респондента
        xlsx.cell(ws, "A14", "Наименование отчитывающейся организации:",
                  border=False, align="left")
        xlsx.merge(ws, "B14:F14", o.name, border=False, align="left")
        xlsx.cell(ws, "A15", "Почтовый адрес:", border=False, align="left")
        xlsx.merge(ws, "B15:F15", o.address, border=False, align="left")
        # кодовая часть — подписи граф по таблице 7 бланка
        head = ["Код Формы по ОКУД",
                "Код отчитывающейся организации (индивидуального предпринимателя) "
                "по ОКПО (для обособленного подразделения и головного "
                "подразделения юридического лица - идентификационный номер)",
                "Код вида деятельности по ОКВЭД", "Код территории по ОКТМО",
                "ИНН", "ОГРН (ОГРНИП)"]
        okved1 = next((c.strip() for c in (o.okved or "").replace(";", ",").split(",")
                       if c.strip()), "")
        vals = ["0609013", o.okpo, okved1, o.oktmo, o.inn, o.ogrn]
        for i, (h, v) in enumerate(zip(head, vals)):
            col = chr(65 + i)
            xlsx.cell(ws, f"{col}17", h, bold=True, fill=True, size=8)
            xlsx.cell(ws, f"{col}18", i + 1, italic=True, size=9)
            xlsx.cell(ws, f"{col}19", v or "")
        xlsx.heights(ws, {17: 96})

    def _row_values(self, w) -> list[float]:
        """29 граф Раздела I для одного отхода (индекс 0 = графа 1)."""
        hc = w.hazard_class
        rc = lambda v: _round_class(v, hc)  # noqa: E731
        st, bu, _ = _placed(w)
        g = [0.0] * 29
        # 1 — наличие на начало (хранение + накопление)
        g[0] = rc(D(w.accumulated_start) + D(w.accumulated_start_nakopl))
        g[1] = rc(w.generated)             # 2
        g[2] = rc(w.received)              # 3
        g[8] = rc(w.processed)             # 9
        g[9] = rc(w.used)                  # 10
        g[12] = rc(w.neutralized)          # 13
        # 14 — ТКО региональному оператору (только по получателю-регоператору,
        # п. 13 Указаний); остальное ТКО — по гр.15–23 как прочие отходы
        g[13] = rc(self._g14(w))
        g[14] = rc(w.transferred_processing)   # 15 — для обработки
        g[16] = rc(w.transferred_util)         # 17 — для утилизации
        g[18] = rc(w.transferred_neutral)      # 19 — для обезвреживания
        g[20] = rc(w.transferred_storage)      # 21 — для хранения
        g[22] = rc(w.transferred_burial)       # 23 — для захоронения
        g[26] = rc(st)                         # 27 — размещение: хранение
        g[27] = rc(bu)                         # 28 — размещение: захоронение
        g[28] = rc(w.accumulated_end)          # 29
        return g

    # строки листа стр.2: 1 — заголовок, 2–4 — уровни шапки, 5 — номера граф,
    # с 6 — данные
    _S1_DEPTH = 3
    _S1_FIRST_ROW = 6

    def _page2(self, wb):
        from openpyxl.utils import get_column_letter
        ws = wb.create_sheet("стр.2")
        # заголовок Раздела I — полная официальная формулировка; merge на всю
        # ширину таблицы (4 служебных + 29 граф), иначе длинный текст обрезается
        xlsx.merge(ws, "A1:AG1", _S1_TITLE, bold=True, border=False)
        _service_header(ws, 2, self._S1_DEPTH)
        _header_tree(ws, 2, 5, _G29, self._S1_DEPTH)
        cols = [get_column_letter(5 + i) for i in range(29)]
        # только виды отходов — итоговых строк «ВСЕГО»/«по классу» в бланке
        # нет (эталон МИНИХ: строка 1 = первый отход); порядок — п. 7 Указаний
        r = self._S1_FIRST_ROW
        for row_no, w in enumerate(self._sorted_wastes(), start=1):
            g = self._row_values(w)
            xlsx.cell(ws, f"A{r}", row_no)
            xlsx.cell(ws, f"B{r}", w.name, align="left")
            # графа В — канонический вид ФККО с пробелами, как печатает ЛКПП
            # («7 33 100 01 72 4»); в XML код остаётся без пробелов
            xlsx.cell(ws, f"C{r}", _fkko.fmt(w.fkko_code))
            xlsx.cell(ws, f"D{r}", w.hazard_class)
            for c, v in zip(cols, g):
                xlsx.cell(ws, f"{c}{r}", v)
            r += 1
        xlsx.widths(ws, {"A": 5, "B": 28, "C": 13, "D": 7,
                         **{c: 10 for c in cols}})
        xlsx.heights(ws, {2: 60, 3: 48, 4: 60})

    def _page3(self, wb):
        """стр.3 — Раздел II (регоператоры ТКО) + Раздел III (объекты размещения)
        + подпись. Разделы II/III — из ctx.extra."""
        from openpyxl.utils import get_column_letter
        ws = wb.create_sheet("стр.3")
        o = self.ctx.organization
        e = self.ctx.extra if isinstance(self.ctx.extra, dict) else {}
        # A–D — служебные А|Б|В|Г, E..P — до 12 граф самой широкой подтаблицы;
        # столбец B широкий — в нём же длинные показатели Раздела III
        xlsx.widths(ws, {"A": 7, "B": 32, "C": 15, "D": 9,
                         **{get_column_letter(5 + i): 11 for i in range(12)}})
        # Раздел II — по бланку № 614: собственные 29 граф тремя подтаблицами
        # (гр.1–9 / 10–17 / 18–29) с шапкой А|Б|В|Г. В принятых отчётах шапка
        # печатается целиком даже у нерегоператоров, поэтому структуру не
        # подменяем текстом, а печатаем всегда.
        xlsx.merge(ws, "A1:P2", _S2_TITLE, bold=True, border=False)
        tko_ops = e.get("tko_operators", [])
        r = 3
        for start, end in _S2_SPLITS:
            if start > 1:
                xlsx.merge(ws, f"A{r}:D{r}", "продолжение раздела II",
                           border=False, italic=True, align="left")
                r += 1
            sub = [(num, path) for num, path in _G29_II if start <= num <= end]
            depth = max(len(p) for _, p in sub)
            _service_header(ws, r, depth)
            _header_tree(ws, r, 5, sub, depth)
            xlsx.heights(ws, {rr: 48 for rr in range(r, r + depth)})
            r += depth + 1
            for idx, op in enumerate(tko_ops, start=1):
                xlsx.cell(ws, f"A{r}", idx)
                xlsx.cell(ws, f"B{r}", op.get("name", ""), align="left")
                xlsx.cell(ws, f"C{r}", _fkko.fmt(op.get("fkko", "")))
                xlsx.cell(ws, f"D{r}", op.get("hazard_class", ""))
                for i, num in enumerate(range(start, end + 1)):
                    xlsx.cell(ws, f"{get_column_letter(5 + i)}{r}",
                              _s2_val(op, num))
                r += 1
        if not tko_ops:
            xlsx.merge(ws, f"A{r}:P{r}", _S2_NOTE, border=False, italic=True,
                       align="left")
            xlsx.heights(ws, {r: 44})
            r += 1
        # Раздел III — по бланку формы (приказ Росстата № 614): три графы и
        # 21 строка с фиксированными номерами 11–31.
        r += 1
        self._s3_row = r
        xlsx.merge(ws, f"A{r}:C{r}", "Раздел III. Сведения об эксплуатируемых "
                   "объектах размещения отходов", bold=True, border=False)
        r += 1
        xlsx.header_row(ws, r, ["N строки", "Наименование показателя", "Фактически"])
        r += 1
        s3 = e.get("disposal_objects_s3", {}) or {}
        for num, label in _S3_ROWS:
            value = s3.get(str(num), s3.get(num, ""))
            xlsx.data_row(ws, r, [num, label, value if value not in (None, "") else "-"])
            r += 1
        # подпись — дословно по бланку № 614 (таблица 16 эталона)
        r += 2
        self._sign_row = r
        xlsx.merge(ws, f"A{r}:C{r}", _SIGN_TITLE, border=False, align="left", size=9)
        xlsx.heights(ws, {r: 48})
        xlsx.cell(ws, f"A{r+2}", o.director_position or "", border=False)
        xlsx.cell(ws, f"B{r+2}", o.director_name or "", border=False)
        xlsx.cell(ws, f"C{r+2}", "", border=False)
        xlsx.cell(ws, f"A{r+3}", "(должность)", border=False, italic=True, size=8)
        xlsx.cell(ws, f"B{r+3}", "(Ф.И.О.)", border=False, italic=True, size=8)
        xlsx.cell(ws, f"C{r+3}", "(подпись)", border=False, italic=True, size=8)
        xlsx.cell(ws, f"A{r+4}", o.phone or "", border=False)
        xlsx.cell(ws, f"B{r+4}", o.email or "", border=False)
        xlsx.cell(ws, f"C{r+4}", "«____» __________ 20___ г.", border=False)
        xlsx.cell(ws, f"A{r+5}", "(номер контактного телефона)", border=False,
                  italic=True, size=8)
        xlsx.cell(ws, f"B{r+5}", "(e-mail)", border=False, italic=True, size=8)
        xlsx.cell(ws, f"C{r+5}", "(дата составления документа)", border=False,
                  italic=True, size=8)
