"""Сведения в региональный кадастр отходов Санкт-Петербурга.

Печатная форма повторяет официальный бланк Комитета по природопользованию
СПб: Форма 1 (сведения о ЮЛ/ИП и объекте НВОС) + Форма 2 (места накопления) +
Форма 3 (движение отходов, 22 графы) + Форма 4 (объекты обработки/утилизации/
обезвреживания) + Форма 5 (уведомление о представлении сведений).

Основание — Распоряжение Комитета по природопользованию, охране окружающей
среды и обеспечению экологической безопасности СПб от 26.04.2023 № 87-р
(в ред. Распоряжения от 14.02.2025 № 28-р). Срок — не позднее 31 марта года,
следующего за отчётным, по каждому обособленному подразделению отдельно.
Каналы подачи (офиц. страница Комитета gov.spb.ru/gov/otrasl/ecology/
otchotnost/regionalnyj-kadastr-othodov/, обновлена 10.06.2026, и «Руководство
по внесению сведений в РКО» от 16.10.2025): с 2025 г. сведения готовятся в
подсистеме «Ведение регионального кадастра отходов» ГИС СПб «Обращение с
отходами производства и потребления в Санкт-Петербурге»
(oopp.kpoos.gov.spb.ru) либо по типовым формам 1-5; подача — через Подсистему
с УКЭП, либо файлом Excel (.xlsx/.xls) на kadastr@kpoos.gov.spb.ru, либо
почтой на адрес Комитета: 191144, Санкт-Петербург, Дегтярный пер., д. 9
стр. 1 (адрес ул. Чайковского, 20, лит. В из Руководства 10.2025 уже не
актуален). XML-выгрузки нет: программа делает Excel по типовым формам.
Сведения, поданные с нарушением Порядка (в т.ч. без района СПб в п. 2.2
Формы 1), по п. 27 Порядка считаются не представленными — отсюда error в
validate().
Места накопления — из ctx.extra['accumulation_sites'] (плоские записи либо
запись места со списком wastes); объекты обработки — из
ctx.extra['treatment_objects'] (обычно отсутствуют, тогда строка «-»).
Шапки и подписи граф Форм 2-4 сверены с принятым Комитетом отчётом за 2024 г.
(ООО «ВСМ-Сервис», подписан УКЭП): Форма 3 делит передачу на регоператора ТКО
(гр.17) и прочих (гр.18) ПО ФАКТИЧЕСКОМУ ПРИЁМЩИКУ из
ctx.extra['waste_receivers'], а не по коду ФККО; гр.19-21 (в другие субъекты
РФ / субъект / цель) берутся оттуда же.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl.utils import column_index_from_string, get_column_letter

from ecodoc.core import fkko
from ecodoc.core.models import Issue
from ecodoc.core.refdata import oktmo_ref
from ecodoc.core.money import D
from ecodoc.core.registry import register
from ecodoc.core.waste_agg import norm_fkko
from ecodoc.render import xlsx
from ecodoc.reports.base import Report

from ecodoc.calendar.obligations import CADASTRE_REGIONS as _REGIONS  # один список


def _num(v) -> float:
    return float(D(v))


@register
class CadastreSPB(Report):
    code = "cadastre-spb"
    title = "Кадастр отходов (СПб)"
    has_xml = False  # сдаётся подписанным Excel, XML-выгрузки нет

    def validate(self) -> list[Issue]:
        issues: list[Issue] = []
        o = self.ctx.organization
        if not o.inn:
            issues.append(Issue("error", "ИНН", "не указан ИНН"))
        if not self.ctx.period.year:
            issues.append(Issue("error", "период", "не указан отчётный год"))
        if not self.ctx.wastes:
            issues.append(Issue("error", "отходы", "нет позиций отходов"))
        if not self._extra_list("accumulation_sites"):
            issues.append(Issue("warning", "места накопления",
                                "нет данных о местах накопления отходов "
                                "(extra.accumulation_sites) — Форма 2 уйдёт в Комитет "
                                "пустой; в принятых отчётах она заполнена (описание "
                                "места, вместимость т/м³, перечень отходов, способ "
                                "накопления)"))
        from ecodoc.core.nvos import subject_code
        regions = {subject_code(ob.region_code) for ob in self.ctx.objects if ob.region_code}
        if regions and not (regions & _REGIONS):
            issues.append(Issue("warning", "регион",
                                "объекты вне СПб/ЛО — проверьте, требуется ли региональный "
                                f"кадастр (коды: {', '.join(sorted(regions))})"))
        # Район СПб в п. 2.2 Формы 1 обязателен (подсказка бланка в D22 и офиц.
        # страница Комитета); без него сведения по п. 27 Порядка «считаются не
        # представленными» — поэтому error, а не warning. Для объектов в ЛО
        # (region_code 47) форма Комитета СПб не сдаётся — там не проверяем.
        ob = self.ctx.objects[0] if self.ctx.objects else None
        if ob is not None and subject_code(ob.region_code or "") in ("", "78") \
                and not self._district(ob):
            issues.append(Issue("error", "район",
                                "в п. 2.2 Формы 1 не определён район Санкт-Петербурга "
                                "(бланк: адрес объекта «с указанием района "
                                "Санкт-Петербурга»; по п. 27 Порядка сведения без "
                                "района считаются не представленными) — добавьте "
                                "«<Название> район» в адрес объекта или укажите "
                                "extra.district"))
        year_next = (self.ctx.period.year + 1) if self.ctx.period.year else "след."
        issues.append(Issue("warning", "срок",
                            f"срок подачи — не позднее 31 марта {year_next} по каждому "
                            "обособленному подразделению (Распоряжение Комитета по "
                            "природопользованию СПб № 87-р в ред. № 28-р); готовить в "
                            "Подсистеме «Ведение регионального кадастра отходов» ГИС "
                            "СПб (oopp.kpoos.gov.spb.ru, подача с УКЭП) либо по типовым "
                            "формам — Excel на kadastr@kpoos.gov.spb.ru; для ЛО/иных "
                            "регионов — сверьте свой НПА"))
        return issues

    def render_xml(self, out_path: Path) -> Path:  # pragma: no cover
        raise NotImplementedError("Кадастр СПб сдаётся подписанным Excel, XML нет")

    def render_print(self, out_path: Path) -> Path:
        out_path = self._ensure_dir(out_path)
        wb = xlsx.new_workbook()
        self._form1(wb)
        self._form2(wb)
        self._form3(wb)
        self._form4(wb)
        self._form5(wb)
        return xlsx.save(wb, out_path)

    # Форма 1 — сведения о ЮЛ/ИП и объекте НВОС ------------------------
    def _form1(self, wb):
        ws = wb.create_sheet("Форма 1")
        o = self.ctx.organization
        ob = self.ctx.objects[0] if self.ctx.objects else None
        xlsx.widths(ws, {"A": 2, "B": 6, "C": 48, "D": 55})
        xlsx.cell(ws, "D1", "Форма 1", border=False, bold=True, align="right")
        xlsx.cell(ws, "C3", "Сведения об индивидуальном предпринимателе/юридическом лице",
                  border=False, bold=True, align="left")
        xlsx.cell(ws, "B5", "1", bold=True)
        xlsx.merge(ws, "C5:D5",
                   "Общие сведения об индивидуальном предпринимателе, юридическом лице",
                   bold=True, fill=True, align="left")
        director = (f"{o.director_position}\n{o.director_name}".strip()
                    if o.director_name else o.director_position)
        rows1 = [
            ("1.1", "ИНН", o.inn),
            ("1.2", "ОГРН/ОГРНИП", o.ogrn),
            ("1.3", "Полное наименование", o.name),
            ("1.4", "Краткое (сокращенное) наименование", o.short_name or o.name),
            ("1.5", "Код ОКПО", o.okpo),
            ("1.6", "Адрес регистрации индивидуального предпринимателя, юридического лица",
             o.address),
            ("1.7", "Почтовый адрес", o.address),
            ("1.8", "Телефон", o.phone),
            ("1.9", "Факс", "-"),
            ("1.10", "Электронная почта", o.email),
            ("1.11", "Адрес интернет-сайта", "-"),
            ("1.12", "ФИО руководителя с указанием должности", director),
            ("1.13", "Коды ОКВЭД (основной и вспомогательные)", o.okved),
            ("1.14", "Должность, ФИО, номер телефона, адрес электронной почты "
                     "ответственного за представление сведений в региональный кадастр",
             self._responsible()),
        ]
        r = 6
        for num, label, value in rows1:
            xlsx.cell(ws, f"B{r}", num)
            xlsx.cell(ws, f"C{r}", label, align="left")
            xlsx.cell(ws, f"D{r}", value or "", align="left")
            r += 1
        xlsx.cell(ws, f"B{r}", "2", bold=True)
        xlsx.merge(ws, f"C{r}:D{r}",
                   "Сведения об объекте, оказывающем негативное воздействие на окружающую среду",
                   bold=True, fill=True, align="left")
        r += 1
        rows2 = [
            ("2.1", "Код объекта, оказывающего негативное воздействие на окружающую среду",
             ob.code if ob else ""),
            # Бланк (D22): «адрес объекта … (с указанием района Санкт-Петербурга)»;
            # в принятом отчёте район дописан в скобках после адреса.
            ("2.2", "Адрес местонахождения объекта, оказывающего негативное воздействие "
                    "на окружающую среду", self._address_with_district(ob) if ob else ""),
            ("2.3", "Код ОКТМО объекта, оказывающего негативное воздействие на окружающую среду",
             (ob.oktmo if ob else "") or o.oktmo),
        ]
        for num, label, value in rows2:
            xlsx.cell(ws, f"B{r}", num)
            xlsx.cell(ws, f"C{r}", label, align="left")
            xlsx.cell(ws, f"D{r}", value or "", align="left")
            r += 1

    # Форма 2 — места накопления --------------------------------------
    def _form2(self, wb):
        ws = wb.create_sheet("Форма 2")
        xlsx.widths(ws, {"A": 2, "B": 6, "C": 34, "D": 9, "E": 9, "F": 34, "G": 15,
                         "H": 9, "I": 20})
        xlsx.cell(ws, "I1", "Форма 2", border=False, bold=True, align="right")
        xlsx.cell(ws, "B3", "Сведения о местах накопления отходов",
                  border=False, bold=True, align="left")
        xlsx.merge(ws, "B5:B6", "№ п/п", bold=True, fill=True)
        xlsx.merge(ws, "C5:C6", "Краткое описание места накопления отходов", bold=True, fill=True)
        xlsx.merge(ws, "D5:E5", "Вместимость места накопления отходов", bold=True, fill=True)
        xlsx.cell(ws, "D6", "тонн", bold=True, fill=True, size=9)
        xlsx.cell(ws, "E6", "м3", bold=True, fill=True, size=9)
        xlsx.merge(ws, "F5:H5", "Перечень накапливаемых отходов", bold=True, fill=True)
        xlsx.cell(ws, "F6", "Наименование отхода по федеральному классификационному "
                            "каталогу отходов", bold=True, fill=True, size=9)
        xlsx.cell(ws, "G6", "Код отхода по федеральному классификационному каталогу отходов",
                  bold=True, fill=True, size=9)
        xlsx.cell(ws, "H6", "Класс опасности отхода", bold=True, fill=True, size=9)
        xlsx.merge(ws, "I5:I6", "Способ накопления (отдельно от других отходов/"
                                "в смеси с другими отходами)", bold=True, fill=True, size=9)
        for i, col in enumerate("BCDEFGHI"):
            xlsx.cell(ws, f"{col}7", i + 1, italic=True, size=9)
        sites = self._extra_list("accumulation_sites")
        r = 8
        if sites:
            # В принятом отчёте отходы сгруппированы под пронумерованными
            # местами: описание и вместимость — один раз на место, дальше по
            # строке на каждый накапливаемый отход. Поддерживаем оба формата
            # записи: место со списком wastes и плоскую запись «место+отход».
            for n, s in enumerate(sites, 1):
                wastes = s.get("wastes")
                wastes = ([w for w in wastes if isinstance(w, dict)]
                          if isinstance(wastes, list) else [])
                if not wastes:
                    wastes = [s]  # плоская запись: отход в тех же полях
                for i, w in enumerate(wastes):
                    xlsx.cell(ws, f"B{r}", n if i == 0 else "")
                    xlsx.cell(ws, f"C{r}", s.get("description", "") if i == 0 else "",
                              align="left")
                    xlsx.cell(ws, f"D{r}", s.get("capacity_t", "") if i == 0 else "")
                    xlsx.cell(ws, f"E{r}", s.get("capacity_m3", "") if i == 0 else "")
                    xlsx.cell(ws, f"F{r}", w.get("waste_name") or w.get("name", ""),
                              align="left")
                    # Код — в формате ФККО «7 33 100 01 72 4» (так в бланке,
                    # Подсистеме и принятом отчёте), класс — римской цифрой.
                    xlsx.cell(ws, f"G{r}", fkko.fmt(w.get("fkko", "")))
                    xlsx.cell(ws, f"H{r}", _hazard(w.get("hazard_class", "")))
                    xlsx.cell(ws, f"I{r}", w.get("method", "отдельно от других отходов"),
                              align="left")
                    r += 1
        else:
            # Данных о местах накопления нет — печатаем прочерки, как в принятых
            # отчётах, а не многоточие-заглушку (validate() при этом даёт
            # предупреждение: суррогат из ctx.wastes без привязки к местам и
            # вместимости Комитет не примет, выдумывать его нельзя).
            for col in "BCDEFGHI":
                xlsx.cell(ws, f"{col}8", "-")
        xlsx.heights(ws, {5: 30, 6: 40})

    # Форма 3 — движение отходов (22 графы) ---------------------------
    def _form3(self, wb):
        ws = wb.create_sheet("Форма 3")
        xlsx.cell(ws, "V5", "Форма 3", border=False, bold=True, align="right")
        xlsx.cell(ws, "A7", "Сведения о движении отходов", border=False, bold=True, align="left")
        # верхний ярус (row 9), спаны групп; single-колонки объединены до row 11
        xlsx.merge(ws, "A9:A11", "№ п/п", bold=True, fill=True)
        xlsx.merge(ws, "B9:B11", "Наименование отхода по федеральному классификационному "
                                 "каталогу отходов", bold=True, fill=True, size=9)
        xlsx.merge(ws, "C9:C11", "Код отхода по федеральному классификационному "
                                 "каталогу отходов", bold=True, fill=True, size=9)
        xlsx.merge(ws, "D9:D11", "Класс опасности отхода", bold=True, fill=True, size=9)
        xlsx.merge(ws, "E9:E11", "Наличие отходов на начало отчетного года, тонн",
                   bold=True, fill=True, size=9)
        xlsx.merge(ws, "F9:F11", "Образовано отходов за отчетный год, тонн",
                   bold=True, fill=True, size=9)
        xlsx.merge(ws, "G9:J9", "Поступило отходов от других хозяйствующих субъектов, тонн",
                   bold=True, fill=True, size=9)
        xlsx.merge(ws, "K9:L9", "Обработано отходов, тонн", bold=True, fill=True, size=9)
        xlsx.merge(ws, "M9:N9", "Утилизировано отходов, тонн", bold=True, fill=True, size=9)
        xlsx.merge(ws, "O9:P9", "Обезврежено отходов, тонн", bold=True, fill=True, size=9)
        xlsx.merge(ws, "Q9:Q11", "Передано региональному оператору по обращению с ТКО, тонн",
                   bold=True, fill=True, size=9)
        xlsx.merge(ws, "R9:U9", "Передано отходов другим хозяйствующим субъектам "
                                "(за исключением регионального оператора по обращению с ТКО), тонн",
                   bold=True, fill=True, size=9)
        xlsx.merge(ws, "V9:V11", "Наличие отходов на конец отчетного года, тонн",
                   bold=True, fill=True, size=9)
        # второй ярус (row 10, объединено до row 11 для подписей)
        sub = {
            "G": "всего", "H": "из них из других субъектов РФ",
            "I": "наименование субъекта РФ", "J": "цель поступления отходов",
            "K": "всего", "L": "из них поступивших из других субъектов РФ",
            "M": "всего", "N": "из них поступивших из других субъектов РФ",
            "O": "всего", "P": "из них поступивших из других субъектов РФ",
            "R": "всего", "S": "из них в другие субъекты РФ",
            "T": "наименование субъекта РФ", "U": "цель передачи отходов",
        }
        for col, label in sub.items():
            xlsx.merge(ws, f"{col}10:{col}11", label, bold=True, fill=True, size=8)
        for i in range(22):
            xlsx.cell(ws, f"{get_column_letter(i+1)}12", i + 1, italic=True, size=9)
        # Приёмщики по ФККО: первая запись на код (aggregate_acts даёт уникальные
        # пары код+приёмщик, для формы берём первого — как в журнале №1028).
        receivers: dict[str, dict] = {}
        for rec in self._extra_list("waste_receivers"):
            receivers.setdefault(norm_fkko(rec.get("fkko")), rec)
        r = 13
        for n, w in enumerate(self.ctx.wastes, 1):
            tko = _tko_mass(w, receivers)
            recv = _num(D(w.transferred) - D(tko))
            rec = receivers.get(norm_fkko(getattr(w, "fkko_code", "")))
            # Гр.19-21 — по принятому отчёту заполняются в каждой строке с
            # передачей: субъект РФ приёмщика (вне СПб → гр.19 масса + гр.20
            # название, внутри СПб → «-»), цель = вид обращения приёмщика;
            # неизвестное — «-», как в эталоне, а не пустая ячейка.
            if _num(w.transferred) > 0:
                region = _receiver_region(rec)
                out_of_spb = bool(region) and region != "Санкт-Петербург"
                to_other = recv if out_of_spb else 0.0
                region_name = region if out_of_spb else "-"
                purpose = ((rec.get("operation") or "").strip() if rec else "") or "-"
            else:
                to_other, region_name, purpose = 0.0, "", ""
            vals = {
                # после агрегации актов код хранится 11 цифрами без пробелов —
                # печатаем в формате каталога, класс римской (как в принятом отчёте)
                "A": n, "B": w.name, "C": fkko.fmt(w.fkko_code),
                "D": _hazard(w.hazard_class),
                "E": _num(w.accumulated_start), "F": _num(w.generated),
                "G": _num(w.received), "H": 0.0, "I": "", "J": "",
                "K": _num(w.processed), "L": 0.0, "M": _num(w.used), "N": 0.0,
                "O": _num(w.neutralized), "P": 0.0, "Q": tko,
                "R": recv, "S": to_other, "T": region_name, "U": purpose,
                "V": _num(w.accumulated_end),
            }
            for col, v in vals.items():
                xlsx.cell(ws, f"{col}{r}", v, align="left" if col in "BIJTU" else "center")
            r += 1
        xlsx.widths(ws, {"A": 5, "B": 30, "C": 15, "D": 8,
                         **{get_column_letter(i): 10 for i in range(5, 23)}})
        xlsx.heights(ws, {9: 40, 10: 40})

    # Форма 4 — объекты обработки/утилизации/обезвреживания ------------
    def _form4(self, wb):
        ws = wb.create_sheet("Форма 4")
        last = get_column_letter(27)
        xlsx.cell(ws, f"{last}1", "Форма 4", border=False, bold=True, align="right")
        xlsx.cell(ws, "A3", "Сведения об объекте обработки, утилизации, обезвреживания отходов",
                  border=False, bold=True, align="left")
        # Сетка шапки повторяет типовую форму Комитета (Типовые_формы.xlsx,
        # лист «Форма 4»): одиночные графы объединены по вертикали 5:8,
        # групповые заголовки — по горизонтали в строке 5, подписи граф внутри
        # групп объединены 6:8, номера граф — строка 9, данные — с 10-й.
        # Тексты подписей — полные формулировки эталона (без склейки
        # «Группа: подпись» и без многоточий-заглушек).
        singles = {
            1: "№, п/п",
            2: "Код объекта в государственном реестре объектов, оказывающих негативное "
               "воздействие на окружающую среду",
            3: "Адрес места нахождения объекта обработки, утилизации, обезвреживания отходов "
               "(с указанием района Санкт-Петербурга)",
            4: "Кадастровый номер земельного участка",
            5: "Категория земель (вид разрешенного использования земельного участка)",
            6: "ИНН, наименование (ФИО), адрес, телефон, факс, интернет-сайт юридического "
               "лица/индивидуального предпринимателя, эксплуатирующего объект обработки, "
               "утилизации, обезвреживания отходов",
            7: "Дата выдачи, номер лицензии на деятельность по сбору, использованию, "
               "обезвреживанию, транспортировке, размещению отходов I-IV классов опасности "
               "и наименование органа, выдавшего ее",
            8: "Применение технологии (промышленное, опытно-промышленное, опытное, иное)",
            9: "Наименование технологии",
            10: "Наименование и реквизиты технической документации на технику/технологию",
            11: "Основной вывод заключения государственной экологической экспертизы "
                "по проекту технической документации на технику/технологию и реквизиты "
                "(в случае проведения государственной экологической экспертизы)",
            12: "Назначение технологии (обработка, утилизация, обезвреживание отходов)",
            13: "Краткая характеристика технологического процесса обработки, утилизации "
                "и (или) обезвреживания отходов",
            19: "ИНН, наименование (ФИО), адрес, телефон, факс, интернет-сайт "
                "разработчика (собственника) технологии",
            20: "Наименование обрабатываемых, утилизируемых и (или) обезвреживаемых отходов",
            21: "Код по ФККО обрабатываемых, утилизируемых и (или) обезвреживаемых отходов",
            22: "Класс опасности",
            23: "Количество отходов, обработанных, утилизированных и (или) обезвреженных "
                "за отчетный год, тонн",
        }
        groups = {  # (первая графа, последняя графа) -> групповой заголовок
            (14, 16): "Используемые установки (необходимое оборудование)",
            (17, 18): "Получение вторичной продукции (энергии)",
            (24, 27): "Образование вторичных отходов",
        }
        subs = {  # подписи граф внутри групп (второй ярус)
            14: "Наименование",
            15: "Реквизиты технической документации (технические условия)",
            16: "Наличие положительного заключения государственной экологической экспертизы",
            17: "Наименование и код по Общероссийскому классификатору продукции по видам "
                "экономической деятельности (ОКПД) получаемой вторичной продукции "
                "(энергии), иное (в случае отсутствия кода по ОКПД)",
            18: "Производительность при получении вторичной продукции (энергии) "
                "(количество в год с указанием единицы измерения)",
            24: "Наименование",
            25: "Код по ФККО",
            26: "Класс опасности",
            27: "Образовано вторичных отходов за отчетный год, тонн",
        }
        for num, label in singles.items():
            col = get_column_letter(num)
            xlsx.merge(ws, f"{col}5:{col}8", label, bold=True, fill=True, size=8)
        for (c1, c2), label in groups.items():
            xlsx.merge(ws, f"{get_column_letter(c1)}5:{get_column_letter(c2)}5", label,
                       bold=True, fill=True, size=8)
        for num, label in subs.items():
            col = get_column_letter(num)
            xlsx.merge(ws, f"{col}6:{col}8", label, bold=True, fill=True, size=8)
        for i in range(27):
            col = get_column_letter(i + 1)
            xlsx.cell(ws, f"{col}9", i + 1, italic=True, size=9)
            ws.column_dimensions[col].width = 6 if i == 0 else 16
        objs = self._extra_list("treatment_objects")
        r = 10
        if objs:
            for n, obj in enumerate(objs, 1):
                xlsx.cell(ws, f"A{r}", n)
                for i in range(1, 27):
                    col = get_column_letter(i + 1)
                    xlsx.cell(ws, f"{col}{r}", obj.get(f"c{i+1}", ""), align="left", size=9)
                r += 1
        else:
            for i in range(27):
                xlsx.cell(ws, f"{get_column_letter(i+1)}10", "-")
        xlsx.heights(ws, {5: 45, 6: 75})

    # Форма 5 — уведомление -------------------------------------------
    def _form5(self, wb):
        ws = wb.create_sheet("Форма 5")
        o = self.ctx.organization
        year = self.ctx.period.year or ""
        xlsx.widths(ws, {"A": 10, "B": 12, "C": 12, "D": 8, "E": 10, "F": 12, "G": 10,
                         "H": 12, "I": 10})
        xlsx.cell(ws, "I1", "Форма 5", border=False, bold=True, align="right")
        for i, line in enumerate(["Председателю Комитета", "по природопользованию, охране",
                                  "окружающей среды и обеспечению",
                                  "экологической безопасности"]):
            xlsx.cell(ws, f"F{2+i}", line, border=False, align="left")
        xlsx.merge(ws, "A13:I13",
                   "Уведомление о представлении сведений в региональный кадастр отходов "
                   "производства и потребления в Санкт-Петербурге", border=False, bold=True)
        xlsx.merge(ws, "A15:I15",
                   f"     Направляю в Ваш адрес сведения в региональный кадастр отходов "
                   f"в Санкт-Петербурге за отчетный период {year} год.", border=False, align="left")
        xlsx.cell(ws, "A16", "     Уникальный идентификационный номер: ",
                  border=False, align="left")
        # «2» — номер сноски из бланка (A18), расшифровка в A34-A35
        xlsx.merge(ws, "A18:I18", "УИН, либо штрих код2", border=False, italic=True,
                   size=9, align="left")
        xlsx.cell(ws, "A21", "     Достоверность представленных сведений подтверждаю.",
                  border=False, align="left")
        xlsx.merge(ws, "A23:I23",
                   "     Приложение: Сведения в региональный кадастр отходов "
                   "в Санкт-Петербурге на ______ л. в 1 экз.", border=False, align="left")
        xlsx.merge(ws, "A27:B27", "Руководитель", border=False, align="left")
        xlsx.merge(ws, "C27:E27", o.name or o.short_name, border=False, align="center")
        xlsx.cell(ws, "H27", o.director_name or "", border=False, align="center")
        xlsx.cell(ws, "C28", "(наименование организации)", border=False, italic=True, size=8)
        xlsx.cell(ws, "F28", "(подпись)", border=False, italic=True, size=8)
        xlsx.cell(ws, "H28", "(Ф.И.О.)", border=False, italic=True, size=8)
        xlsx.cell(ws, "B30", "М.П.", border=False)
        xlsx.cell(ws, "F30", "(дата)", border=False, italic=True, size=8)
        xlsx.cell(ws, "B31", "(при наличии).", border=False, italic=True, size=8)
        # сноска бланка: A34 — линия, A35 — «2 Не заполняется» (УИН присваивает
        # Комитет, заявитель поле не трогает)
        xlsx.cell(ws, "A34", "___________________", border=False, align="left")
        xlsx.cell(ws, "A35", "2 Не заполняется", border=False, size=9, align="left")

    # --- вспомогательное ---
    def _district(self, ob) -> str:
        """Район Санкт-Петербурга объекта для п. 2.2 Формы 1 (пусто — не определён).

        Источники по порядку: extra['district'] (ручное указание эколога),
        «<Название> район»/«р-н» в адресе объекта, запись справочника
        data/oktmo_ref.json с тем же ОКТМО, где в value назван район (напр.
        «Санкт-Петербург, Приморский р-н (Богатырский пр.)»). ОКТМО СПб кодирует
        муниципальные округа (40 3xx 000), а не районы, поэтому таблицы
        «ОКТМО → район» здесь нет — её не из чего проверить."""
        e = self.ctx.extra if isinstance(self.ctx.extra, dict) else {}
        manual = str(e.get("district") or "").strip()
        if manual:
            return _district_name(manual) or manual
        found = _district_name(ob.address or "")
        if found:
            return found
        code = str(ob.oktmo or "").strip()
        if code:
            for rec in oktmo_ref().values():
                if isinstance(rec, dict) and str(rec.get("oktmo") or "") == code:
                    found = _district_name(str(rec.get("value") or ""))
                    if found:
                        return found
        return ""

    def _address_with_district(self, ob) -> str:
        """Адрес объекта с районом в скобках, как в принятом отчёте
        («…стр.1 (Всеволожский р-н)»); если район уже назван в адресе —
        адрес как есть."""
        addr = ob.address or ""
        district = self._district(ob)
        if not district or _district_name(addr):
            return addr
        return f"{addr} ({district} район)"

    def _responsible(self) -> str:
        e = self.ctx.extra if isinstance(self.ctx.extra, dict) else {}
        if e.get("responsible"):
            return str(e["responsible"])
        o = self.ctx.organization
        parts = [p for p in ["Эколог", o.director_name, o.phone, o.email] if p]
        return ", ".join(parts)

    def _extra_list(self, key: str) -> list[dict]:
        e = self.ctx.extra if isinstance(self.ctx.extra, dict) else {}
        v = e.get(key, [])
        return [x for x in v if isinstance(x, dict)] if isinstance(v, list) else []


# «Приморский район», «Приморский р-н», «р-н Центральный» → название района.
# Прилагательное с большой буквы (-ий/-ый/-ой: Приморский, Центральный,
# Курортный, Петродворцовый, Всеволожский) строго рядом со словом «район»/
# «р-н» — иначе слово из названия улицы за район не примется.
_RE_DISTRICT = re.compile(
    r"(?:(?<![А-Яа-яЁё])([А-ЯЁ][а-яё]+(?:ий|ый|ой))\s+(?:район|р-н)(?![А-Яа-яЁё]))"
    r"|(?:(?<![А-Яа-яЁё])(?:район|р-н)\s+([А-ЯЁ][а-яё]+(?:ий|ый|ой))(?![А-Яа-яЁё]))")


def _district_name(text: str) -> str:
    """Название района из текста адреса (без слова «район»); пусто — нет."""
    m = _RE_DISTRICT.search(text or "")
    if not m:
        return ""
    return m.group(1) or m.group(2) or ""


def _hazard(value) -> str:
    """Класс опасности для печати — римской цифрой (I…V), как в принятом
    Комитетом отчёте; уже римский/неизвестный текст — как есть."""
    return fkko.roman(value) or str(value or "")


# Резервные ТКО-подтипы для случая, когда приёмщик неизвестен: 7 31 … (отходы
# коммунальные из жилищ) и 7 33 100 01 72 х (мусор от офисных и бытовых
# помещений организаций). Весь блок «7 3…» ТКО НЕ считается: в принятом
# Комитетом отчёте смет с территории (7 33 39…) и мусор производственных/
# складских помещений (7 33 2…) стоят в гр.18 «прочим», а не у регоператора.
# core/waste_agg.is_tko не трогаем — он общий для декларации и 2-ТП.
_TKO_FALLBACK_PREFIXES = ("731", "7331000172")

# Код субъекта РФ в номере лицензии («Л020-00113-47/…» → 47). Это регион
# ВЫДАВШЕГО лицензию органа, поэтому только запасной вариант после адреса.
_RE_LICENSE_REGION = re.compile(r"-(\d{2})/")
_REGION_BY_CODE = {"78": "Санкт-Петербург", "47": "Ленинградская область"}


def _is_regoperator(name: str) -> bool:
    """Приёмщик — региональный оператор по обращению с ТКО.

    Для СПб это АО «Невский экологический оператор»; общий признак — слова
    «региональный оператор» в наименовании. «УК по обращению с отходами в
    Ленинградской области» регоператором СПб не является — её передачи идут
    в гр.18 (так в принятом отчёте)."""
    low = (name or "").lower()
    return "невский эколог" in low or ("региональн" in low and "оператор" in low)


def _tko_mass(w, receivers: dict[str, dict]) -> float:
    """Масса, переданная региональному оператору ТКО (гр.17 Формы 3).

    Признак — ФАКТИЧЕСКИЙ приёмщик из extra['waste_receivers'] (так делит
    принятый Комитетом отчёт), а не код ФККО; лишь когда приёмщик неизвестен —
    резервно узкий список ТКО-подтипов, а не весь блок «7 3…»."""
    code = norm_fkko(getattr(w, "fkko_code", ""))
    rec = receivers.get(code)
    if rec is not None:
        return float(D(w.transferred)) if _is_regoperator(rec.get("receiver", "")) else 0.0
    if any(code.startswith(p) for p in _TKO_FALLBACK_PREFIXES):
        return float(D(w.transferred))
    return 0.0


def _receiver_region(rec: dict | None) -> str:
    """Субъект РФ приёмщика для гр.19-20 Формы 3.

    Первичный источник — адрес (и наименование: «…в Ленинградской области»),
    запасной — код региона в номере лицензии. Пусто — субъект неизвестен."""
    if not rec:
        return ""
    text = " ".join(str(rec.get(k) or "")
                    for k in ("address", "receiver_address", "receiver")).lower()
    if "ленинградск" in text:
        return "Ленинградская область"
    if "санкт-петербург" in text or "с.-петербург" in text or "спб" in text:
        return "Санкт-Петербург"
    m = _RE_LICENSE_REGION.search(str(rec.get("license") or ""))
    return _REGION_BY_CODE.get(m.group(1), "") if m else ""
