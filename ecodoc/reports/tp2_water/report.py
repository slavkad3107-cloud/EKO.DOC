"""Форма 2-ТП (водхоз) — сведения об использовании воды. Годовая, с первого
рабочего дня по 22 января года, следующего за отчётным. Код ОКУД 0609060.

ДЕЙСТВУЮЩАЯ форма — Приказ Росстата от 02.10.2024 № 445 (заменил № 815 от
27.12.2019) В РЕДАКЦИИ приказа Росстата от 06.08.2026 № 473 (изменены
указания: п. 1.8 — первичный учёт по формам приказа МПР от 27.05.2026 № 311;
п. 1.10 — пояснительная записка при изменении объёмов/масс более чем на 10 %;
п. 2.1/3.1 — отдельные строки по каждому водозабору/выпуску И по каждому
разрешительному документу, при сбросе разных категорий качества через один
выпуск — отдельные строки; п. 2.2 — при водопользовании без забора из
поверхностного объекта гр. 1-3 не заполняются). Сам бланк № 473 не менял.
Сверено 22.08.2026 по normativ.kontur.ru (documentId=507797, 507773).
Адресат — территориальный орган Росводресурсов в субъекте РФ (БВУ); подача
электронно через Модуль респондента ИАС «2-ТП (водхоз)» (НЕ в Росстат).

Состав печатной формы — ДОСЛОВНО по бланку приложения к приказу № 445
(дополнительно сверен с принятым отчётом ПРОТЕЛЮКС за 2023):
титульный лист (кодовая часть гр.1-6 = ОКУД/ОКПО/ИНН/ОКВЭД2/ОКАТО/ГУИВ —
НЕ ОКТМО/ОГРН, как в 2-ТП отходы; блок подписи должностного лица),
Раздел 1 — 49 граф, строки нумеруются 11, 12, … (графа «А» = N строки):
  гр.1-3 документ, 4-6 источник, 7-10 коды, 11 лимит, 12 всего за год,
  13-24 по месяцам, 25 учтено приборами, 26 потери, 27-28 коды территорий
  использования, 29-30 расходы в системах оборотного/повторного
  водоснабжения, 31 использовано всего, 32-41 — ПЯТЬ пар «код вида
  использования (Прил. 3) / объём», 42-47 — ТРИ пары «код категории воды
  (гр. 5 Прил. 2) / объём» передано БЕЗ использования, 48-49 — ОДНА пара
  «код категории (гр. 6 Прил. 2) / объём» передано ПОСЛЕ использования.
Раздел 2 — строки 21, 22, …; 30 граф + пары «код ЗВ / масса» (гр. 31-78,
  до 24 веществ). Точность: объёмы — тыс м³ до 0,01; расстояние — км до 0,1;
  массы ЗВ — до 0,001 (примечание к бланку), единицы по Прил. 5 (т/кг).
  Итоговых строк в бланке НЕТ — итоги только на служебном листе «Сводка».

Данные — из ctx.extra['water'] (объёмы — тыс. м³/год, расстояния — км).
Графа без источника в данных печатается ПУСТОЙ — ничего не выдумываем.
  {
    "okato": "41221505",     # титульный лист, гр.5 — код территории ОКАТО
    "guiv": "412285",        # титульный лист, гр.6 — код водопользователя ГУИВ
    "intake": [{             # Раздел 1, словарь на источник/поставщика:
        "name": "...", "type": "...",       # справочно (в бланке граф нет)
        "doc_type": "Л",                    # гр.1 — Д(оговор)/Л(ицензия)/Р(ешение)
        "doc_no": "480082",                 # гр.2
        "doc_date": "25.02.2020",           # гр.3
        "source_code": "60",                # гр.4 — код типа источника (Прил. 1)
        "water_body_code": "БАЛ/НАРВА",     # гр.5 — код водного объекта
        "distance_km": 16.9,                # гр.6 — расстояние от устья, км
        "supplier_guiv": "",                # гр.7 — код поставщика по ГУИВ
        "quality": "ТН",                    # гр.8 — категория качества (Прил. 2)
        "okato": "41221505", "vhu": "01.03.00.004",   # гр.9-10
        "limit": 145.71,                    # гр.11 — допустимый объём забора
        "volume": 1.29,                     # гр.12 — забрано всего за год
        "months": [12 значений, янв..дек],  # гр.13-24 (Σ должна = гр.12)
        "measured": 1.29,                   # гр.25 — учтено средствами измерений
        "losses": 0.0,                      # гр.26 — потери при транспортировке
        "use_okato": "", "use_vhu": "",     # гр.27-28 — территория использования
        "recycled": 0.0, "reused": 0.0,     # гр.29-30 — оборотное/повторное
        "used_total": 1.29,                 # гр.31 — использовано всего за год
        "uses": [{"code": "102", "volume": 1.29}],      # гр.32-41 (до 5 пар)
        "transfers": [{"code": "ПК", "volume": 0}],     # гр.42-47 — передано
                                            # БЕЗ использования (до 3 пар,
                                            # код категории — гр.5 Прил. 2)
        "transfer_after_use": {"code": "СК", "volume": 0}   # гр.48-49 —
                                            # передано ПОСЛЕ использования
                                            # (код категории — гр.6 Прил. 2)
      }],
    "discharge": [{          # Раздел 2, словарь на выпуск:
        "receiver": "р. Нарва",             # справочно
        "doc_type": "Р", "doc_no": "...", "doc_date": "...",   # гр.1-3
        "receiver_code": "20",              # гр.4 — код типа приёмника (Прил. 1)
        "water_body_code": "...",           # гр.5
        "distance_km": 16.9,                # гр.6
        "quality": "СД",                    # гр.7 — код категории (Прил. 2)
        "okato": "...", "vhu": "...",       # гр.8-9
        "limit": 145.71,                    # гр.10 — допустимый объём водоотведения
        "volume": 1.29,                     # гр.11 — отведено всего за год
        "measured": 1.29,                   # гр.12 — учтено средствами измерений
        "polluted_no_treat": 0,             # гр.13 — загрязнённых без очистки
        "insufficiently_treated": 0,        # гр.14 — недостаточно очищенных
        "normatively_clean": 0,             # гр.15 — нормативно чистых без очистки
        "treatment_code": "5",              # гр.16 — код очистного сооружения (Прил. 4: 5/6/7)
        "normatively_treated": 1.29,        # гр.17 — нормативно очищенных
        "treatment_capacity": 2.5,          # гр.18 — мощность очистных сооружений
        "months": [12 значений],            # гр.19-30
        "pollutants": [{"code": "132", "name": "БПКполн", "mass": 0.05}]
                                            # гр.31-78 — пары «код ЗВ / масса»
      }],
    "recycled": 40.0, "reused": 0.0, "used_own": 10.0   # сводные показатели
  }
"""
from __future__ import annotations

from pathlib import Path

from lxml import etree
from openpyxl.utils import get_column_letter

from ecodoc.core.models import Issue
from ecodoc.core.money import D
from ecodoc.core.registry import register
from ecodoc.render import xlsx
from ecodoc.render.xmlutil import el, write_tree
from ecodoc.reports.base import Report

NPA_SHORT = ("Приказ Росстата от 02.10.2024 № 445 (в ред. приказа от "
             "06.08.2026 № 473)")

_MONTH_NAMES = ("январь", "февраль", "март", "апрель", "май", "июнь",
                "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь")


S1_USE_PAIRS = 5        # гр.32-41 (п. 2.15 указаний)
S1_TRANSFER_PAIRS = 3   # гр.42-47 (п. 2.16 указаний); гр.48-49 — п. 2.17
S2_ZV_PAIRS_MAX = 24    # гр.31-78 (п. 3.13 указаний)
# Прил. 4 к форме — коды сооружений очистки воды
TREATMENT_CODES = {"5": "Биологическая очистка", "6": "Физико-химическая "
                   "очистка", "7": "Механическая очистка"}
# Прил. 2 к форме — коды категорий качества воды (все графы 3-7)
QUALITY_CODES = {"ПО", "ПК", "ПД", "ТН", "ТД", "ТР", "ТП", "МР", "СК", "СД",
                 "СТ", "МН", "ТМ", "РВ", "КД", "РС", "КР", "ШР", "БЛ", "ЛВ"}
# Прил. 5 к форме — ЗВ, массы которых приводятся в ТОННАХ (прочие — в кг)
ZV_IN_TONNES = {"132": "БПК полн.", "113": "взвешенные вещества",
                "80": "нефтепродукты (нефть, углеводороды нефти)",
                "40": "сульфат-ион", "83": "сухой остаток (минерализация)",
                "52": "хлорид-ион", "90": "фосфат-ион", "2": "азот общий",
                "3": "аммоний-ион"}


def _build_s1_cols() -> list[tuple[int, str]]:
    """49 граф Раздела 1 дословно по бланку приложения к приказу № 445."""
    cols = [
        (1, "Тип разрешительного документа (Д, Л, Р)"),
        (2, "Номер документа"),
        (3, "Дата документа"),
        (4, "Код типа источника (Прил. 1)"),
        (5, "Код водного объекта"),
        (6, "Расстояние от устья, км"),
        (7, "Код поставщика по ГУИВ"),
        (8, "Код категории качества воды (Прил. 2)"),
        (9, "Код территории по ОКАТО"),
        (10, "Код ВХУ"),
        (11, "Допустимый объём забора воды, тыс м³"),
        (12, "Забрано или получено — всего за год, тыс м³"),
    ]
    cols += [(13 + i, _MONTH_NAMES[i]) for i in range(12)]
    cols += [
        (25, "Учтено средствами измерений, тыс м³"),
        (26, "Потери при транспортировке, тыс м³"),
        (27, "Использовано: коды территорий — по ОКАТО"),
        (28, "Использовано: коды территорий — ВХУ"),
        (29, "Расходы в системах водоснабжения, тыс м³ — оборотного"),
        (30, "Расходы в системах водоснабжения, тыс м³ — повторного"),
        (31, "Использовано — всего за год, тыс м³"),
    ]
    # По бланку № 445 (п. 2.15-2.17 указаний): гр.32-41 — ПЯТЬ пар «код вида
    # использования (Прил. 3) / объём»; гр.42-47 — ТРИ пары «код категории
    # воды (гр.5 Прил. 2) / объём» передано без использования; гр.48-49 —
    # ОДНА пара «код (гр.6 Прил. 2) / объём» передано после использования.
    # Раньше было 6 пар использования (32-43) и 3 пары передачи (44-49) —
    # графы 42-49 печатались со сдвигом на одну пару.
    for i in range(S1_USE_PAIRS):
        cols += [(32 + 2 * i, "Использовано за год по кодам видов "
                              "использования, тыс м³ — код"),
                 (33 + 2 * i, "Использовано за год по кодам видов "
                              "использования, тыс м³ — объём")]
    for i in range(S1_TRANSFER_PAIRS):
        cols += [(42 + 2 * i, "Передано без использования, по кодам "
                              "категорий воды, тыс м³ — код"),
                 (43 + 2 * i, "Передано без использования, по кодам "
                              "категорий воды, тыс м³ — объём")]
    cols += [(48, "Передано после использования, тыс м³ — код"),
             (49, "Передано после использования, тыс м³ — объём")]
    return cols


_S1_COLS = _build_s1_cols()

# 30 граф Раздела 2 дословно по бланку № 445; пары «код ЗВ / масса»
# (гр.31-78) добавляются динамически по числу веществ (не более 24)
_S2_COLS = [
    (1, "Решение (Р)/Лицензия (Л) — тип (Р, Л)"),
    (2, "Решение (Р)/Лицензия (Л) — номер"),
    (3, "Решение (Р)/Лицензия (Л) — дата"),
    (4, "Приёмник отведённых вод — код типа приёмника (Прил. 1)"),
    (5, "Приёмник отведённых вод — код водного объекта"),
    (6, "Приёмник отведённых вод — расстояние от устья, км"),
    (7, "Коды — категории качества воды (Прил. 2)"),
    (8, "Коды — по ОКАТО"),
    (9, "Коды — ВХУ"),
    (10, "Допустимый объём водоотведения, тыс м³"),
    (11, "Отведено воды, всего за год, тыс м³"),
    (12, "Учтено средствами измерений, тыс м³"),
    (13, "Отведено в водные объекты, тыс м³ — загрязнённых — без очистки"),
    (14, "Отведено в водные объекты, тыс м³ — загрязнённых — недостаточно "
         "очищенных"),
    (15, "Отведено в водные объекты, тыс м³ — нормативно чистых "
         "(без очистки)"),
    (16, "Отведено в водные объекты, тыс м³ — нормативно-очищенных — код "
         "очистного сооружения (Прил. 4)"),
    (17, "Отведено в водные объекты, тыс м³ — нормативно-очищенных — объём"),
    (18, "Мощность очистных сооружений, тыс м³"),
] + [(19 + i, f"Отведено за месяц, тыс м³ — {_MONTH_NAMES[i]}")
     for i in range(12)]


def _numv(v, places: int = 2):
    """Число для печати или пустая графа: данных нет — ничего не выдумываем.

    Точность по указаниям к форме: объёмы — тыс м³ до второго знака
    (п. 2.1, 3.1), расстояние — до 0,1 км (п. 2.3.3, 3.3), массы ЗВ — до
    трёх знаков (примечание к бланку). Округление — в момент печати, данные
    в базе не трогаем.
    """
    if v in (None, ""):
        return ""
    return float(round(D(v), places))


def _months_row(rec) -> list:
    """12 помесячных значений; недостающие — пустые графы."""
    ms = list(rec.get("months") or [])
    return [_numv(ms[i]) if i < len(ms) else "" for i in range(12)]


def _pairs_row(items, n, places: int = 2) -> list:
    """Пары «код + объём/масса» фиксированной ширины (пустые добиваются)."""
    out: list = []
    for i in range(n):
        if i < len(items):
            it = items[i]
            out += [str(it.get("code", "")),
                    _numv(it.get("volume", it.get("mass")), places)]
        else:
            out += ["", ""]
    return out


def _opt(parent, tag, val):
    """Необязательный текстовый элемент XML — только если значение задано."""
    if val not in (None, ""):
        el(parent, tag, val)


def _optn(parent, tag, val):
    """Необязательный числовой элемент XML — только если значение задано."""
    if val not in (None, ""):
        el(parent, tag, float(D(val)))


def _xml_months(parent, rec):
    """Помесячная разбивка (гр.13-24 / 19-30 бланка): <Месяцы><М н="1">…»."""
    ms = rec.get("months") or []
    if not ms:
        return
    box = el(parent, "Месяцы")
    for i in range(12):
        v = ms[i] if i < len(ms) else 0
        el(box, "М", float(D(v or 0)), н=str(i + 1))


def _xml_doc(parent, rec):
    """Разрешительный документ (гр.1-3): договор/лицензия/решение."""
    if any(rec.get(k) for k in ("doc_type", "doc_no", "doc_date")):
        el(parent, "Документ", тип=rec.get("doc_type", ""),
           номер=rec.get("doc_no", ""), дата=rec.get("doc_date", ""))


@register
class TP2Water(Report):
    code = "2tp-water"
    title = "2-ТП (водхоз)"

    def _water(self) -> dict:
        return self.ctx.extra.get("water", {}) or {}

    def validate(self) -> list[Issue]:
        issues: list[Issue] = []
        o = self.ctx.organization
        if not o.inn:
            issues.append(Issue("error", "ИНН", "не указан ИНН"))
        if not o.okpo:
            issues.append(Issue("warning", "ОКПО", "для 2-ТП обязателен код ОКПО"))
        if not self.ctx.period.year:
            issues.append(Issue("error", "период", "не указан отчётный год"))
        w = self._water()
        if not w.get("intake") and not w.get("discharge"):
            issues.append(Issue("error", "водоучёт",
                                "нет данных водозабора/водоотведения "
                                "(заполните extra.water: intake/discharge)"))
        for part in ("intake", "discharge"):
            for rec in w.get(part, []):
                q = str(rec.get("quality", "") or "")
                # по бланку № 445 категория качества задаётся только КОДОМ из
                # Прил. 2 (гр.8 Раздела 1 / гр.7 Раздела 2)
                if q and q.upper() not in QUALITY_CODES:
                    issues.append(Issue(
                        "warning", "качество",
                        f"{part}: категория качества «{q}» — не код Прил. 2 "
                        f"к форме № 445 ({', '.join(sorted(QUALITY_CODES))})"))
        for d in w.get("discharge", []):
            tc = str(d.get("treatment_code", "") or "")
            # Прил. 4 к форме № 445: только 5/6/7 (старые коды вроде «40»
            # Модуль не примет)
            if tc and tc not in TREATMENT_CODES:
                issues.append(Issue(
                    "warning", "очистка",
                    f"код очистного сооружения «{tc}» — не из Прил. 4 к "
                    f"форме № 445 (5 — биологическая, 6 — физико-химическая, "
                    f"7 — механическая)"))
            # п. 3.10: сумма гр.14 + гр.17 должна быть равна гр.11
            if d.get("insufficiently_treated") not in (None, "") or \
                    d.get("normatively_treated") not in (None, ""):
                s = D(d.get("insufficiently_treated") or 0) + \
                    D(d.get("normatively_treated") or 0)
                if (s - D(d.get("volume", 0))).copy_abs() > D("0.01"):
                    issues.append(Issue(
                        "warning", "очистка",
                        f"выпуск «{d.get('receiver', '')}»: сумма гр.14 и "
                        f"гр.17 ({s}) не равна гр.11 (п. 3.10 указаний)"))
            if len(d.get("pollutants", []) or []) > S2_ZV_PAIRS_MAX:
                issues.append(Issue(
                    "warning", "ЗВ", f"выпуск «{d.get('receiver', '')}»: "
                    f"в бланке место только для {S2_ZV_PAIRS_MAX} веществ "
                    "(гр.31-78), лишние не печатаются"))
        for s_ in w.get("intake", []):
            uses = s_.get("uses", []) or []
            if len(uses) > S1_USE_PAIRS:
                issues.append(Issue(
                    "warning", "использование",
                    f"источник «{s_.get('name', '')}»: видов использования "
                    f"{len(uses)}, а в бланке {S1_USE_PAIRS} пар (гр.32-41)"))
            # п. 2.15: сумма гр.33, 35, 37, 39, 41 должна быть равна гр.31
            if uses and s_.get("used_total") not in (None, ""):
                su = sum((D(u.get("volume") or 0) for u in uses), D(0))
                if (su - D(s_.get("used_total"))).copy_abs() > D("0.01"):
                    issues.append(Issue(
                        "warning", "использование",
                        f"источник «{s_.get('name', '')}»: сумма объёмов по "
                        f"видам использования ({su}) не равна гр.31 "
                        f"({s_.get('used_total')}) — п. 2.15 указаний"))
            if len(s_.get("transfers", []) or []) > S1_TRANSFER_PAIRS:
                issues.append(Issue(
                    "warning", "передача",
                    f"источник «{s_.get('name', '')}»: передач без "
                    f"использования больше {S1_TRANSFER_PAIRS} (гр.42-47)"))
        if w.get("discharge"):
            # п. 3.1 указаний (ред. № 473): к разделу 2 прилагаются результаты
            # расчёта НДС (II кат. — по данным ДВОС; III — для веществ I-II кл.)
            issues.append(Issue(
                "warning", "приложения",
                "к Разделу 2 прилагаются результаты расчёта нормативов "
                "допустимого сброса (п. 3.1 указаний к форме № 445)"))
        # помесячная разбивка: 12 значений, сумма = годовому объёму (гр.12/11);
        # без этого Модуль респондента отчёт не примет
        for part in ("intake", "discharge"):
            for rec in w.get(part, []):
                ms = rec.get("months")
                if not ms:
                    continue
                if len(ms) != 12:
                    issues.append(Issue(
                        "warning", "месяцы",
                        f"{part}: помесячная разбивка должна содержать 12 "
                        f"значений (январь-декабрь), получено {len(ms)}"))
                    continue
                diff = (sum((D(m or 0) for m in ms), D(0))
                        - D(rec.get("volume", 0))).copy_abs()
                if diff > D("0.01"):
                    issues.append(Issue(
                        "warning", "месяцы",
                        f"{part}: сумма помесячных объёмов не сходится с "
                        f"годовым (расхождение {diff} тыс. м³)"))
        return issues

    # ---------------- XML -------------------------------------------------
    def render_xml(self, out_path: Path) -> Path:
        out_path = self._ensure_dir(out_path)
        o = self.ctx.organization
        w = self._water()
        # version 0.3 — добавлены реквизиты документов, коды-классификаторы,
        # лимиты и помесячная разбивка (<Месяцы><М н="1">…</М></Месяцы>)
        root = etree.Element("Форма2ТПВодхоз", version="0.4", ОКУД="0609060",
                             НПА=NPA_SHORT)
        org = el(root, "Респондент")
        el(org, "Наименование", o.name)
        el(org, "ИНН", o.inn)
        el(org, "ОКПО", o.okpo)
        el(org, "ОКТМО", o.oktmo)
        _opt(org, "ОКАТО", w.get("okato"))
        _opt(org, "ГУИВ", w.get("guiv"))
        el(org, "Адрес", o.address)
        el(org, "Адресат", "территориальное Бассейновое водное управление "
                           "(Модуль респондента ИАС «2-ТП (водхоз)» / ГИС ЦП «Вода»)")
        el(root, "ОтчётныйГод", self.ctx.period.year)

        intk = el(root, "ЗаборВоды")
        for s in w.get("intake", []):
            x = el(intk, "Источник")
            el(x, "Наименование", s.get("name", ""))
            el(x, "Тип", s.get("type", ""))
            _xml_doc(x, s)                                       # гр.1-3
            _opt(x, "КодТипаИсточника", s.get("source_code"))    # гр.4
            _opt(x, "КодВодногоОбъекта", s.get("water_body_code"))  # гр.5
            _optn(x, "РасстояниеОтУстья", s.get("distance_km"))  # гр.6
            _opt(x, "ПоставщикГУИВ", s.get("supplier_guiv"))     # гр.7
            _opt(x, "КатегорияКачества", s.get("quality"))       # гр.8
            _opt(x, "ОКАТО", s.get("okato"))                     # гр.9
            _opt(x, "ВХУ", s.get("vhu"))                         # гр.10
            _optn(x, "ДопустимыйОбъём", s.get("limit"))          # гр.11
            el(x, "Объём", float(D(s.get("volume", 0))))         # гр.12
            _xml_months(x, s)                                    # гр.13-24
            _optn(x, "УчтеноПриборами", s.get("measured"))       # гр.25
            _optn(x, "Потери", s.get("losses"))                  # гр.26
            _opt(x, "ОКАТОИспользования", s.get("use_okato"))    # гр.27
            _opt(x, "ВХУИспользования", s.get("use_vhu"))        # гр.28
            _optn(x, "Оборотная", s.get("recycled"))             # гр.29
            _optn(x, "Повторная", s.get("reused"))               # гр.30
            _optn(x, "ИспользованоВсего", s.get("used_total"))   # гр.31
            for u in s.get("uses", []):                          # гр.32-43
                el(x, "ВидИспользования", float(D(u.get("volume", 0))),
                   код=str(u.get("code", "")))
            for t in s.get("transfers", []):                     # гр.42-47
                el(x, "ПереданоБезИспользования",
                   float(D(t.get("volume", 0))), код=str(t.get("code", "")))
            tau = s.get("transfer_after_use") or {}                # гр.48-49
            if tau.get("code") or tau.get("volume") not in (None, ""):
                el(x, "ПереданоПослеИспользования",
                   float(D(tau.get("volume", 0))),
                   код=str(tau.get("code", "")))
        disc = el(root, "Водоотведение")
        for d in w.get("discharge", []):
            x = el(disc, "Выпуск")
            el(x, "Приёмник", d.get("receiver", ""))
            el(x, "Качество", d.get("quality", ""))              # гр.7 (код)
            el(x, "Объём", float(D(d.get("volume", 0))))         # гр.11
            _xml_doc(x, d)                                       # гр.1-3
            _opt(x, "КодТипаПриёмника", d.get("receiver_code"))  # гр.4
            _opt(x, "КодВодногоОбъекта", d.get("water_body_code"))  # гр.5
            _optn(x, "РасстояниеОтУстья", d.get("distance_km"))  # гр.6
            _opt(x, "ОКАТО", d.get("okato"))                     # гр.8
            _opt(x, "ВХУ", d.get("vhu"))                         # гр.9
            _optn(x, "ДопустимыйОбъём", d.get("limit"))          # гр.10
            _optn(x, "УчтеноПриборами", d.get("measured"))       # гр.12
            # гр.13-18 — отведено по категориям очистки (структура бланка,
            # а не самодельные «в пределах/сверх норматива»)
            if any(d.get(k) not in (None, "") for k in
                   ("polluted_no_treat", "insufficiently_treated",
                    "normatively_clean", "treatment_code",
                    "normatively_treated", "treatment_capacity")):
                cat = el(x, "ОтведеноПоКатегориям")
                _optn(cat, "БезОчистки", d.get("polluted_no_treat"))
                _optn(cat, "НедостаточноОчищенные",
                      d.get("insufficiently_treated"))
                _optn(cat, "НормативноЧистые", d.get("normatively_clean"))
                if d.get("normatively_treated") not in (None, "") or \
                        d.get("treatment_code") not in (None, ""):
                    el(cat, "НормативноОчищенные",
                       float(D(d.get("normatively_treated", 0))),
                       код=str(d.get("treatment_code", "")))
                _optn(cat, "МощностьОчистных", d.get("treatment_capacity"))
            _xml_months(x, d)                                    # гр.19-30
            # сброшенные ЗВ по выпуску (гр.31-78, коды по Прил. 5 к № 445)
            for zv in d.get("pollutants", [])[:S2_ZV_PAIRS_MAX]:
                z = el(x, "ЗВ", код=str(zv.get("code", "")),
                       ед="т" if str(zv.get("code", "")) in ZV_IN_TONNES
                       else "кг")
                el(z, "Наименование", zv.get("name", ""))
                el(z, "Масса", float(D(zv.get("mass", 0))))
        el(root, "ОборотнаяВода", float(D(w.get("recycled", 0))))
        el(root, "ПовторноПоследовательная", float(D(w.get("reused", 0))))
        el(root, "ИспользованоНаСобственныеНужды", float(D(w.get("used_own", 0))))
        write_tree(root, out_path)
        return out_path

    # ---------------- Печатная форма --------------------------------------
    def render_print(self, out_path: Path) -> Path:
        out_path = self._ensure_dir(out_path)
        w = self._water()
        wb = xlsx.new_workbook()
        self._title_page(wb, w)
        self._section1(wb, w)
        self._section2(wb, w)
        self._summary(wb, w)
        return xlsx.save(wb, out_path)

    def _title_page(self, wb, w):
        """Адресная часть бланка № 445: без неё форму нельзя распечатать
        и подписать. Кодовая часть гр.1-6 — по принятому отчёту:
        ОКУД / ОКПО / ИНН / ОКВЭД2 / ОКАТО / ГУИВ (не ОКТМО и не ОГРН)."""
        ws = wb.create_sheet("Титульный лист")
        o = self.ctx.organization
        year = self.ctx.period.year or ""
        xlsx.widths(ws, {"A": 22, "B": 24, "C": 18, "D": 18, "E": 18, "F": 20})
        xlsx.merge(ws, "A1:F1", "ФЕДЕРАЛЬНОЕ СТАТИСТИЧЕСКОЕ НАБЛЮДЕНИЕ",
                   bold=True, border=False)
        xlsx.merge(ws, "A3:F3", "СВЕДЕНИЯ ОБ ИСПОЛЬЗОВАНИИ ВОДЫ",
                   bold=True, border=False)
        xlsx.merge(ws, "A4:F4", f"за {year} г.", bold=True, border=False)
        xlsx.merge(ws, "E5:F5", "Форма № 2-ТП (водхоз), годовая",
                   border=False, italic=True, align="right")
        xlsx.merge(ws, "A5:D6",
                   "Предоставляют: юридические лица и индивидуальные "
                   "предприниматели, осуществляющие пользование водными "
                   "объектами, — территориальному органу Росводресурсов "
                   "в субъекте Российской Федерации  ·  Срок: с первого "
                   "рабочего дня по 22 января после отчётного периода  ·  "
                   "Годовая  ·  Приказ Росстата: об утверждении формы от "
                   "02.10.2024 № 445; о внесении изменений от 06.08.2026 "
                   "№ 473", border=False, align="left")
        xlsx.cell(ws, "A8", "Наименование отчитывающейся организации",
                  border=False, align="left")
        xlsx.merge(ws, "A9:F9", o.name, border=False, align="left")
        xlsx.cell(ws, "A11", "Почтовый адрес", border=False, align="left")
        xlsx.merge(ws, "A12:F12", o.address, border=False, align="left")
        # подписи граф кодовой части — п. 1.11 указаний к форме
        head = ["Код формы по ОКУД",
                "Код отчитывающейся организации (индивидуального "
                "предпринимателя) по ОКПО (для обособленного подразделения "
                "и головного подразделения юридического лица — "
                "идентификационный номер)",
                "ИНН", "Код основного (наиболее водоёмкого) вида "
                "деятельности по ОКВЭД2",
                "Код по ОКАТО (восьмизначный, до третьего уровня)",
                "Код по государственному учёту использования вод (ГУИВ)"]
        okved1 = next((c.strip() for c in (o.okved or "").replace(";", ",").split(",")
                       if c.strip()), "")
        # ОКАТО и ГУИВ в модели организации отсутствуют — берём из
        # extra.water (okato/guiv); нет данных — графа остаётся пустой
        vals = ["0609060", o.okpo, o.inn, okved1,
                w.get("okato", ""), w.get("guiv", "")]
        for i, (h, v) in enumerate(zip(head, vals)):
            col = chr(65 + i)
            xlsx.cell(ws, f"{col}15", h, bold=True, fill=True, size=9)
            xlsx.cell(ws, f"{col}16", i + 1, italic=True, size=9)
            xlsx.cell(ws, f"{col}17", v or "")
        # блок подписи — стр. 4 бланка (должность, ФИО, телефон, e-mail, дата)
        xlsx.merge(ws, "A19:F19",
                   "Должностное лицо, ответственное за предоставление "
                   "первичных статистических данных:", border=False, align="left")
        xlsx.merge(ws, "A21:F21",
                   f"{o.official_title}  ______________  {o.director_name}",
                   border=False, align="left")
        xlsx.merge(ws, "A22:F22", "(должность)            (Ф.И.О.)            "
                   "(подпись)", border=False, italic=True, align="left")
        xlsx.merge(ws, "A23:F23",
                   f"{o.phone}    E-mail {o.email}    "
                   "«___» __________ 20___ год", border=False, align="left")
        xlsx.merge(ws, "A24:F24", "(номер контактного телефона)            "
                   "(дата составления документа)", border=False, italic=True,
                   align="left")
        xlsx.heights(ws, {3: 30, 15: 90})

    def _numbered_sheet(self, wb, name, title, cols, first_row_no, rows):
        """Лист с пронумерованными графами бланка: строка подписей + строка
        номеров (как на официальной форме). Колонка A — графа «А» бланка
        «N строки» (Раздел 1 — 11, 12, …; Раздел 2 — 21, 22, …), справочное
        имя объекта печатается ПОСЛЕ последней графы (в бланке его нет).
        Итоговой строки в бланке нет — итоги на служебном листе «Сводка»."""
        ws = wb.create_sheet(name)
        last = get_column_letter(2 + len(cols))
        xlsx.merge(ws, f"A1:{last}1", title, bold=True, border=False,
                   align="left")
        xlsx.merge(ws, f"A2:{last}2", "Объёмы — тыс м³ с точностью до 0,01; "
                   "расстояние — км до 0,1 (п. 2.1, 2.3.3, 3.1, 3.3 указаний)",
                   border=False, italic=True, align="left")
        xlsx.cell(ws, "A3", "N строки", bold=True, fill=True, size=8)
        xlsx.cell(ws, "A4", "А", italic=True, size=8)
        for i, (num, lbl) in enumerate(cols):
            c = get_column_letter(2 + i)
            xlsx.cell(ws, f"{c}3", lbl, bold=True, fill=True, size=8)
            xlsx.cell(ws, f"{c}4", num, italic=True, size=8)
        xlsx.cell(ws, f"{last}3", "Объект (справочно, не графа бланка)",
                  bold=True, fill=True, size=8)
        r = 5
        for n, (lead, vals) in enumerate(rows):
            xlsx.cell(ws, f"A{r}", first_row_no + n)
            for i, v in enumerate(vals):
                xlsx.cell(ws, f"{get_column_letter(2 + i)}{r}", v)
            xlsx.cell(ws, f"{last}{r}", lead, align="left")
            r += 1
        xlsx.widths(ws, {"A": 9, last: 26,
                         **{get_column_letter(2 + i): 9
                            for i in range(len(cols))}})
        xlsx.heights(ws, {3: 96})
        return ws, r

    def _section1(self, wb, w):
        """Раздел 1 — 49 граф (реквизиты документа, коды-классификаторы,
        лимит, год + помесячно, использование по кодам видов, передача)."""
        rows = []
        for s in w.get("intake", []):
            vals = [
                s.get("doc_type", ""), s.get("doc_no", ""),      # гр.1-2
                s.get("doc_date", ""),                           # гр.3
                str(s.get("source_code", "")),                   # гр.4
                s.get("water_body_code", ""),                    # гр.5
                _numv(s.get("distance_km"), 1),                  # гр.6
                s.get("supplier_guiv", ""),                      # гр.7
                s.get("quality", ""),                            # гр.8
                s.get("okato", ""), s.get("vhu", ""),            # гр.9-10
                _numv(s.get("limit")),                           # гр.11
                float(D(s.get("volume", 0))),                    # гр.12
                *_months_row(s),                                 # гр.13-24
                _numv(s.get("measured")),                        # гр.25
                _numv(s.get("losses")),                          # гр.26
                s.get("use_okato", ""), s.get("use_vhu", ""),    # гр.27-28
                _numv(s.get("recycled")),                        # гр.29
                _numv(s.get("reused")),                          # гр.30
                _numv(s.get("used_total")),                      # гр.31
                *_pairs_row(s.get("uses", []), S1_USE_PAIRS),    # гр.32-41
                *_pairs_row(s.get("transfers", []),
                            S1_TRANSFER_PAIRS),                  # гр.42-47
                *_pairs_row([s["transfer_after_use"]]
                            if s.get("transfer_after_use") else [], 1),
                                                                 # гр.48-49
            ]
            rows.append((s.get("name", "") or s.get("type", ""), vals))
        self._numbered_sheet(
            wb, "Раздел 1",
            "Раздел 1. Забрано из природных источников, получено от "
            "поставщиков, использовано, передано и потеряно воды",
            _S1_COLS, 11, rows)

    def _section2(self, wb, w):
        """Раздел 2 «Водоотведение» — 30 граф + пары «код ЗВ / масса»
        (гр.31-78); отдельная самодельная таблица ЗВ бланку не соответствует
        и заменена штатными графами."""
        discharge = w.get("discharge", [])
        npairs = min(S2_ZV_PAIRS_MAX,
                     max((len(d.get("pollutants", [])) for d in discharge),
                         default=0))
        cols = list(_S2_COLS)
        for i in range(npairs):
            cols += [(31 + 2 * i, "Содержание ЗВ в отведённых водах — код "
                                  "ЗВ (Прил. 5)"),
                     (32 + 2 * i, "Содержание ЗВ в отведённых водах — масса "
                                  "(т/кг по Прил. 5)")]
        rows = []
        for d in discharge:
            zv = [{"code": p.get("code", ""), "volume": p.get("mass", 0)}
                  for p in d.get("pollutants", [])]
            vals = [
                d.get("doc_type", ""), d.get("doc_no", ""),      # гр.1-2
                d.get("doc_date", ""),                           # гр.3
                str(d.get("receiver_code", "")),                 # гр.4
                d.get("water_body_code", ""),                    # гр.5
                _numv(d.get("distance_km"), 1),                  # гр.6
                d.get("quality", ""),                            # гр.7
                d.get("okato", ""), d.get("vhu", ""),            # гр.8-9
                _numv(d.get("limit")),                           # гр.10
                float(D(d.get("volume", 0))),                    # гр.11
                _numv(d.get("measured")),                        # гр.12
                _numv(d.get("polluted_no_treat")),               # гр.13
                _numv(d.get("insufficiently_treated")),          # гр.14
                _numv(d.get("normatively_clean")),               # гр.15
                str(d.get("treatment_code", "")),                # гр.16
                _numv(d.get("normatively_treated")),             # гр.17
                _numv(d.get("treatment_capacity")),              # гр.18
                *_months_row(d),                                 # гр.19-30
                *_pairs_row(zv, npairs, 3),                      # гр.31+
            ]
            rows.append((d.get("receiver", ""), vals))
        ws, r = self._numbered_sheet(
            wb, "Раздел 2", "Раздел 2. Водоотведение", cols, 21, rows)
        # сноска <1> и примечание бланка — дословно
        xlsx.merge(ws, f"A{r + 2}:P{r + 2}",
                   "<1> БПК полн (132), взвешенные вещества (113), "
                   "нефтепродукты (нефть, углеводороды нефти) (80), "
                   "сульфат-ион (40), сухой остаток (минерализация) (83), "
                   "хлорид-ион (52), фосфат-ион (90), азот общий (2), "
                   "аммоний-ион (3) приводятся в тоннах, прочие ЗВ — в "
                   "килограммах.", border=False, italic=True, align="left")
        xlsx.merge(ws, f"A{r + 3}:P{r + 3}",
                   "Примечание: значение показателей граф 32, 34, …, 78 "
                   "округляется до трёх знаков после запятой.",
                   border=False, italic=True, align="left")

    def _summary(self, wb, w):
        """Сводка — служебный лист (не бланк): итоги и реквизиты формы."""
        total_in = sum((D(s.get("volume", 0)) for s in w.get("intake", [])),
                       D(0))
        total_out = sum((D(d.get("volume", 0)) for d in w.get("discharge", [])),
                        D(0))
        ws = wb.create_sheet("Сводка")
        xlsx.header_row(ws, 1, ["Показатель", "Значение, тыс. м³/год"],
                        widths=[40, 20])
        xlsx.data_row(ws, 2, ["Забрано воды, всего", float(total_in)])
        xlsx.data_row(ws, 3, ["Отведено (сброшено), всего", float(total_out)])
        xlsx.data_row(ws, 4, ["Оборотное водоснабжение",
                              float(D(w.get("recycled", 0)))])
        xlsx.data_row(ws, 5, ["Повторно-последовательное водоснабжение",
                              float(D(w.get("reused", 0)))])
        xlsx.data_row(ws, 6, ["Использовано на собственные нужды",
                              float(D(w.get("used_own", 0)))])
        xlsx.data_row(ws, 8, ["Форма / ОКУД", "2-ТП (водхоз) / 0609060"])
        xlsx.data_row(ws, 9, ["Основание", NPA_SHORT])
        xlsx.data_row(ws, 10, ["Адресат", "территориальный орган "
                               "Росводресурсов в субъекте РФ (БВУ; Модуль "
                               "респондента ИАС «2-ТП (водхоз)»)"])
        year = self.ctx.period.year or 0
        xlsx.data_row(ws, 11, [f"Срок (за {year})",
                               f"с 1-го рабочего дня по 22.01.{year + 1}"])
        xlsx.data_row(ws, 12, ["Пояснительная записка",
                               "при изменении объёмов забора/получения, "
                               "расходов оборотного/повторного водоснабжения, "
                               "водоотведения, масс ЗВ более чем на 10 % "
                               "(п. 1.10 в ред. № 473)"])
