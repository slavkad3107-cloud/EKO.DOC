"""Заполнение бланка журнала учёта отходов (Приказ №1028) из папки «Формы».

Пользователь ведёт журнал в готовом файле Excel, где часть ячеек — формулы
(ссылки на титул, суммы). Поэтому: не рисуем лист заново, а вписываем данные
в бланк-образец, оставляя формулы на месте (их пересчитает Excel).

Колонки ищем по подписям граф, а не по фиксированным адресам: бланки у разных
организаций смещены на строку-другую.

Важно: бланк-образец — это, как правило, ЧУЖОЙ уже заполненный журнал
(другой организации), а не пустая форма. Поэтому строки данных образца
предварительно очищаются — иначе отходы и контрагенты чужой организации
уезжали в отчётный документ клиента. Запись ограничена областью таблицы
(до шапки блока-продолжения); не хватило строк — таблица раздвигается,
объединённые ячейки не трогаем (раньше запись в них роняла заполнение
с AttributeError, сбой молча глотался, и бланк не применялся никогда).

Формулы. «Расчётной» считается только формула-агрегат ВНУТРИ своей строки
(=SUM(F8:J8)) — её оставляем Excel. Межлистовые ссылки образца
('Приложение 1'!B6, ='Приложение 2 (год)'!E16, =G8 из другой строки) —
это привязка чужого бланка к конкретным ячейкам: при другом составе/порядке
отходов графы разъезжались (Прил. 3 всегда показывало отход №1, а «для
захоронения» было жёстко равно «передано» из строки 16 Прил. 2). Такие
формулы затираем и пишем значения.

Приложение 2 состоит из двух блоков: основная часть (графы 1–7) и
«продолжение» (графы 8–16: обработано, утилизировано, обезврежено, передано,
размещено, остаток). Второй блок заполняется по тем же строкам, что и
первый — «№ строки» совпадает с «№ п/п».

Редакции: до 31.08.2026 — № 1028 в ред. приказа № 825 от 13.12.2023;
с 01.09.2026 — приказ Минприроды № 227 от 16.04.2026 (см. edition()).
"""
from __future__ import annotations

import re
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl.cell.cell import MergedCell

from ecodoc.core import fkko
from ecodoc.core.models import Issue
from ecodoc.render import template_xlsx as tx

CODE = "waste-movement"

# --- редакции Порядка учёта --------------------------------------------------
# № 1028 (ред. № 825 от 13.12.2023, действует с 01.09.2024) — до 31.08.2026;
# № 227 от 16.04.2026 (Минюст 29.05.2026 N 86726) — с 01.09.2026 по 01.09.2032.
ED_825 = "825"
ED_227 = "227"
ED_227_START = (2026, 9, 1)

REF_825 = ("к Порядку учета в области обращения с отходами, утвержденному "
           "приказом Министерства природных ресурсов и экологии Российской "
           "Федерации от 8 декабря 2020 г. N 1028 (в ред. приказа от 13.12.2023 N 825)")
REF_227 = ("к Порядку учета в области обращения с отходами производства и "
           "потребления, утвержденному приказом Минприроды России "
           "от 16.04.2026 N 227")

# заголовок Таблицы 1: приказом № 825 «Состав…» переименован в «Перечень…»
APP1_TITLE = "Перечень образующихся видов отходов, подлежащих учету"


def edition(ctx) -> str:
    """Какая редакция Порядка действует для отчёта.

    Явно: ctx.extra["order_edition"] = "825" | "227". Иначе — по дате
    составления (extra["report_date"], ДД.ММ.ГГГГ или ГГГГ-ММ-ДД), а без неё
    — по периоду: обобщение за 2026 год и за 3–4 кв. 2026 делается уже после
    01.09.2026, т.е. по № 227; всё более раннее — по № 825."""
    e = ctx.extra if isinstance(ctx.extra, dict) else {}
    forced = str(e.get("order_edition") or "").strip()
    if forced in (ED_825, ED_227):
        return forced
    rd = str(e.get("report_date") or "").strip()
    m = (re.match(r"(\d{2})\.(\d{2})\.(\d{4})", rd)
         or re.match(r"(\d{4})-(\d{2})-(\d{2})", rd))
    if m:
        g = m.groups()
        ymd = (int(g[2]), int(g[1]), int(g[0])) if len(g[0]) == 2 else tuple(map(int, g))
        return ED_227 if ymd >= ED_227_START else ED_825
    year = int(ctx.period.year or 0)
    q = ctx.period.quarter
    if year > 2026 or (year == 2026 and q in (None, 3, 4)):
        return ED_227
    return ED_825


def ref_text(ed: str) -> str:
    return REF_227 if ed == ED_227 else REF_825


def period_text(ctx) -> str:
    """Период отчёта БЕЗ предлога «за»: «1 кв 2025 года» / «2025 года».

    В бланке слово «за» уже стоит в шапке («…отходами за»), а на титул
    идёт формула =Титул!F11 — поэтому «за» в значение не вписываем, иначе
    получалось «за за 2025 год год». Свободный текст периода (как у
    пользователя: «1 кв - 3 кв 2025 года», «полугодие 2026 года» по № 227)
    — через ctx.extra["period_text"]."""
    e = ctx.extra if isinstance(ctx.extra, dict) else {}
    if e.get("period_text"):
        return str(e["period_text"]).strip()
    year = ctx.period.year or ""
    q = ctx.period.quarter
    if q:
        return f"{q} кв {year} года"
    return f"{year} года" if year else ""


def _num(value, ed: str = ED_825) -> float | None:
    """Тоннаж бланка: Порядок требует «в тоннах с точностью до трех знаков
    после запятой» — двоичный хвост float (469.0220892119701) в журнал не
    идёт. По № 227 (п. 12) масса менее 0,001 т учитывается с точностью до
    ЧЕТЫРЁХ знаков (0,0004 т не пропадает). Ноль не пишем: пустая графа
    читается как отсутствие."""
    if value in (None, ""):
        return None
    try:
        d = Decimal(str(value).replace(",", "."))
    except Exception:
        return None
    if not d:
        return None
    step = Decimal("0.001")
    if ed == ED_227 and abs(d) < Decimal("0.001"):
        step = Decimal("0.0001")
    q = d.quantize(step, rounding=ROUND_HALF_UP)
    return float(q) if q else None


# --- разбор шапки --------------------------------------------------------------
# подпись графы (начало) → ключ данных (num — порядковый номер строки).
# Ключи t_*/partner/contract/… наполняются в зависимости от вида листа:
# «переданные» (Таблица 3) — из transferred_* и extra["waste_receivers"],
# «полученные» (Таблица 4) — из extra["waste_suppliers"]; подписи граф в
# обоих листах одинаковые, различить их можно только по заголовку листа.
_COLUMNS = [
    ("№ п/п", "num"),
    ("n п/п", "num"),
    ("№ строки", "num"),                 # блок-продолжение Прил. 2
    ("n строки", "num"),
    ("наименование отход", "name"),
    ("наименование вида отход", "name"),  # № 227
    ("код фкко", "fkko_code"),
    ("код по фкко", "fkko_code"),          # № 227
    ("класс опасности", "hazard_class"),
    ("происхождение", "origin"),
    ("агрегатное состояние", "aggregate_state"),
    ("химический", "composition"),
    ("образовано отходов", "generated"),
    ("получено отходов", "received"),
    ("поступление отходов с собственных", "received_own"),     # № 227, гр. 8
    ("количество полученных отходов", "received"),   # Таблица 4 («Всего»)
    ("количество переданных отходов", "transferred"),  # Таблица 3 («Всего»)
    ("обработано", "processed"),
    ("утилизировано", "used"),
    ("обезврежено", "neutralized"),
    ("передано отходов", "transferred"),
    ("передача отходов", "transferred_own"),  # № 227, гр. 13 («на собственные объекты»)
    ("для обработки", "t_processing"),
    ("для утилизации", "t_util"),
    ("для обезвреживания", "t_neutral"),
    ("для хранения", "t_storage"),
    ("для захоронения", "t_burial"),
    ("сведения о лицах", "partner"),
    ("дата и номер договора", "contract"),
    ("срок действия договора", "contract_term"),
    ("реквизиты лицензии", "license"),
]

# подграфы «хранение/накопление/всего/захоронение» сами по себе ничего не
# значат — ключ зависит от родительской графы над ними (объединённой шапки)
_SUBCOLUMNS = {
    "наличие отходов на начало": {"хранение": "acc_start",
                                  "накопление": "acc_start_nakopl"},
    "наличие отходов на конец": {"хранение": "acc_end",
                                 "накопление": "acc_end_nakopl"},
    "размещено отходов": {"всего": "placed", "хранение": "placed_storage",
                          "захоронение": "placed_burial"},
    "количество переданных отходов": {"всего": "transferred"},
    "количество полученных отходов": {"всего": "received"},
}

_TITLE_FIELDS = [
    ("наименование юридического лица", "org_name"),
    ("индивидуального предпринимателя", "org_name"),
    ("инн", "inn"),
    ("огрн", "ogrn"),
    ("отчетный год", "period"),
    ("отчётный год", "period"),
]

# служебные подписи титула: строка организации ВЫШЕ их, значения — правее
_TITLE_AUX = ("период", "подпись", "фио", "дата", "ответственный")


def _text(v) -> str:
    if not isinstance(v, str):
        return ""
    t = re.sub(r"\s+", " ", v).strip().lower()
    # опечатка в бланке пользователя («Хренение» в графе 13 Прил. 2) — без
    # поправки подграфа «хранение» под «Размещено» не находилась
    return t.replace("хренение", "хранение")


def _merged_range_at(ws, row: int, col: int):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng
    return None


def _parent_text(ws, row: int, col: int) -> str:
    """Текст графы НАД подграфой: верхняя-левая ячейка объединённого диапазона,
    накрывающего ячейку выше, либо сама ячейка выше."""
    if row <= 1:
        return ""
    rng = _merged_range_at(ws, row - 1, col)
    if rng is not None:
        return _text(ws.cell(row=rng.min_row, column=rng.min_col).value)
    return _text(ws.cell(row=row - 1, column=col).value)


def _collect_columns(ws, head: int, depth: int = 3) -> dict[str, int]:
    """Графы блока по подписям строк шапки head..head+depth-1.

    Подграфы (хранение/накопление/всего/захоронение) привязываются через
    родительскую графу — иначе «хранение» на начало и на конец периода
    неразличимы."""
    cols: dict[str, int] = {}
    for r in range(head, head + depth):
        for cell in ws[r]:
            text = _text(cell.value)
            if not text:
                continue
            parent = _parent_text(ws, cell.row, cell.column)
            for plabel, subs in _SUBCOLUMNS.items():
                if parent.startswith(plabel) and text in subs:
                    cols.setdefault(subs[text], cell.column)
            for label, attr in _COLUMNS:
                if attr not in cols and text.startswith(label):
                    cols[attr] = cell.column
    return cols


def header_row(ws, max_row: int = 40) -> tuple[int, dict] | None:
    """Найти строку шапки таблицы и колонки граф по подписям."""
    pos = tx.find_anchor(ws, "код фкко", limit_rows=max_row) or \
        tx.find_anchor(ws, "код по фкко", limit_rows=max_row)
    if not pos:
        return None
    row = pos[0]
    # шапка бывает в 2–3 строки (Таблица 4 — три), «код фкко» — в верхней
    cols = _collect_columns(ws, max(1, row - 2), depth=5)
    return (row, cols) if "fkko_code" in cols else None


def continuation_row(ws, after: int, max_row: int = 400) -> tuple[int, dict] | None:
    """Шапка блока «продолжение» Приложения 2 (строка «№ строки») НИЖЕ after."""
    for r in range(after + 1, min(ws.max_row, max_row) + 1):
        for c in ws[r]:
            if _text(c.value).startswith(("№ строки", "n строки")):
                cols = _collect_columns(ws, r, depth=3)
                return (r, cols) if len(cols) > 1 else None
    return None


def _merged_rows(ws) -> set[int]:
    rows: set[int] = set()
    for rng in ws.merged_cells.ranges:
        rows.update(range(rng.min_row, rng.max_row + 1))
    return rows


def _is_numbering_row(ws, row: int) -> bool:
    """Строка нумерации граф: «А», 1, 2, 3… — подряд идущие номера."""
    nums: list[int] = []
    for c in ws[row]:
        v = c.value
        if v in (None, ""):
            continue
        if isinstance(v, str):
            v = v.strip()
            if len(v) <= 2 and not v.isdigit():
                continue                    # буквенные графы «А», «Б»
            if not v.isdigit():
                return False
            v = int(v)
        if isinstance(v, float):
            if not v.is_integer():
                return False
            v = int(v)
        if isinstance(v, bool) or not isinstance(v, int):
            return False
        nums.append(v)
    return len(nums) >= 3 and nums == list(range(nums[0], nums[0] + len(nums)))


def _is_next_block(ws, row: int) -> bool:
    """Начало следующего блока таблицы («продолжение», «№ строки»)."""
    for c in ws[row]:
        if _text(c.value).startswith(("продолжение", "№ строки", "n строки")):
            return True
    return False


def data_region(ws, header: int) -> tuple[int, int]:
    """Границы области данных под шапкой (включительно).

    Начало — после строк шапки (они задеты объединениями) и строки нумерации
    граф; конец — перед первой строкой следующего блока: она либо входит в
    объединённый диапазон (шапка «продолжения» бланка), либо подписана
    «продолжение»/«№ строки». Раньше данные писались с первой ПУСТОЙ строки —
    т.е. ПОСЛЕ чужих строк образца и прямо по шапке блока-продолжения."""
    merged = _merged_rows(ws)
    row = header + 1
    while row < header + 20 and (row in merged or _is_numbering_row(ws, row)):
        row += 1
    start = row
    end = start
    while (end - start < 500 and (end + 1) not in merged
           and not _is_next_block(ws, end + 1)):
        end += 1
    return start, end


_REF_RE = re.compile(r"\$?[A-Z]{1,3}\$?(\d+)")


def is_row_formula(value, row: int) -> bool:
    """Формула-агрегат внутри СВОЕЙ строки (=SUM(F8:J8) в строке 8) —
    единственный вид формул, который оставляем Excel. Ссылки на другие листы
    ('Приложение 1'!B6) и другие строки (=G8 в строке 16) — привязка чужого
    образца к конкретным ячейкам, их надо заменить значениями."""
    if not (isinstance(value, str) and value.startswith("=")):
        return False
    if "!" in value:
        return False
    refs = [int(m.group(1)) for m in _REF_RE.finditer(value)]
    return bool(refs) and all(r == row for r in refs)


def _clear_rows(ws, start: int, end: int) -> None:
    """Очистить строки данных образца: бланк из «Формы» — чужой ЗАПОЛНЕННЫЙ
    журнал, и без очистки его отходы/контрагенты попадали в документ клиента.
    Внутристрочные формулы-агрегаты оставляем (их пересчитает Excel),
    межлистовые/межстрочные ссылки образца затираем."""
    for r in range(start, end + 1):
        for cell in ws[r]:
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if v is None or is_row_formula(v, r):
                continue
            cell.value = None


def _insert_row(ws, idx: int) -> None:
    """insert_rows + ручной сдвиг объединённых диапазонов: openpyxl «не
    управляет зависимостями» при вставке строк — merge-диапазоны остаются на
    старых координатах и наехали бы на новые строки данных."""
    ws.insert_rows(idx)
    for rng in ws.merged_cells.ranges:
        if rng.min_row >= idx:
            rng.shift(0, 1)


def _sheet_kind(ws) -> str:
    """Вид листа по его заголовку: "transferred" (Таблица 3 — переданные),
    "received" (Таблица 4 — полученные), "" — общие данные (Прил. 1/2)."""
    text = " ".join(_text(c.value) for row in ws.iter_rows(min_row=1, max_row=4)
                    for c in row if isinstance(c.value, str))
    if "переданных другим лицам" in text:
        return "transferred"
    if "полученных от других лиц" in text:
        return "received"
    return ""


def _by_fkko(ctx, key: str) -> dict[str, dict]:
    """Контрагенты по коду ФККО из ctx.extra (waste_receivers/waste_suppliers):
    в модели WasteFlow сведений о лицах и договорах нет."""
    e = ctx.extra if isinstance(ctx.extra, dict) else {}
    out: dict[str, dict] = {}
    for r in e.get(key, []):
        if isinstance(r, dict) and r.get("fkko"):
            out[fkko.norm(str(r["fkko"])) or str(r["fkko"])] = r
    return out


def site_line(ctx) -> str:
    """Строка площадки как в принятых журналах: «Площадка по адресу: "<адрес>"»
    (бланк пользователя: ООО "…" Площадка по адресу: "143570, …")."""
    e = ctx.extra if isinstance(ctx.extra, dict) else {}
    if e.get("site_line"):
        return str(e["site_line"]).strip()
    if getattr(ctx, "objects", None):
        ob = ctx.objects[0]
        if ob.address:
            return f'Площадка по адресу: "{ob.address.strip()}"'
        if ob.name:
            return f"Площадка: {ob.name.strip()}"
    return ""


def requisites_line(ctx) -> str:
    """«ИНН …, ОГРН …» — на титуле бланка отдельных полей нет, а реквизиты
    в журнале нужны; печатаем строкой под организацией."""
    org = ctx.organization
    parts = []
    if org.inn:
        parts.append(f"ИНН {org.inn}")
    if org.ogrn:
        parts.append(f"ОГРН {org.ogrn}")
    return ", ".join(parts)


def org_line(ctx, with_requisites: bool = True) -> str:
    org = ctx.organization
    lines = [x for x in (org.name or org.short_name or "", site_line(ctx),
                         requisites_line(ctx) if with_requisites else "") if x]
    return "\n".join(lines)


def _replace_below_heading(ws, line: str) -> None:
    """Бланки без подписей полей (реальный журнал): строка организации — первая
    текстовая ячейка под заголовком формы; в образце там чужая организация."""
    if not line:
        return
    pos = tx.find_anchor(ws, "данные учета в области обращения с отходами")
    if not pos:
        return
    for r in range(pos[0] + 1, pos[0] + 12):
        for cell in ws[r]:
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if not isinstance(v, str) or not v.strip() or v.startswith("="):
                continue
            if v.strip().lower().startswith(_TITLE_AUX):
                return          # начались служебные подписи — строки орг. нет
            cell.value = line
            return


def _replace_right(ws, label: str, value) -> None:
    """Заменить значение ПРАВЕЕ подписи (не обязательно в соседней ячейке):
    в бланке пользователя между «Ответственный исполнитель» и ФИО — 4 пустых
    колонки, и там остаются данные чужой организации."""
    pos = tx.find_anchor(ws, label)
    if not pos:
        return
    row, col = pos
    for cell in ws[row]:
        if cell.column <= col or isinstance(cell, MergedCell):
            continue
        v = cell.value
        if v in (None, "") or (isinstance(v, str) and v.startswith("=")):
            continue
        if str(v).strip().lower().startswith(_TITLE_AUX):
            return              # дальше идёт следующая подпись — значения нет
        cell.value = value
        return


def _fill_title(ws, ctx, values: dict) -> None:
    """Титул: сначала подписи-поля; в бланках без подписей — замена строки
    чужой организации, периода, исполнителя и даты (образец — чужой отчёт)."""
    pairs = {}
    for label, key in _TITLE_FIELDS:
        if values.get(key) not in (None, ""):
            pairs[label] = values[key]
    missing = tx.fill_by_labels(ws, pairs)
    org_labels = {l for l, k in _TITLE_FIELDS if k == "org_name" and l in pairs}
    # ИНН/ОГРН без своих полей в бланке — дописываем в строку организации
    req_missing = {"инн", "огрн"} & set(pairs) <= set(missing)
    if not org_labels or org_labels <= set(missing):
        _replace_below_heading(ws, org_line(ctx, with_requisites=req_missing))
    if values.get("period"):
        tx.fill_by_labels(ws, {"период": values["period"]})
    e = ctx.extra if isinstance(ctx.extra, dict) else {}
    _replace_right(ws, "ответственный исполнитель",
                   ctx.organization.director_name or "")
    _replace_right(ws, "дата", str(e.get("report_date", "")))


def _drop_year_after_period_formula(ws) -> None:
    """В шапке листов образца: «…за» + формула =Титул!F11 + ячейка «год».
    Период на титуле пишем уже со словом «года» («1 кв 2025 года»), поэтому
    соседнее «год» даёт «2025 года год» — убираем его."""
    for row in ws.iter_rows(min_row=1, max_row=6):
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("=") and "титул" in v.lower():
                nxt = ws.cell(row=c.row, column=c.column + 1)
                if _text(nxt.value) in ("год", "года") and not isinstance(nxt, MergedCell):
                    nxt.value = None


def _actualize_edition(ws, ed: str, issues: list | None) -> None:
    """Бланки пользователя в «Формах» — редакции 2020 г.: реквизит приказа без
    «в ред. № 825», Таблица 1 названа «Состав…» (№ 825: «Перечень…»), в
    Таблице 4 есть графа «для накопления…», исключённая № 825. Реквизит и
    заголовок актуализируем; графу удалить нельзя без переразметки объединений
    бланка — предупреждаем."""
    ref = ref_text(ed)
    for row in ws.iter_rows(min_row=1, max_row=8):
        for c in row:
            v = c.value
            if not isinstance(v, str) or isinstance(c, MergedCell):
                continue
            low = _text(v)
            if "n 1028" in low or "№ 1028" in low or "n 227" in low:
                m = re.match(r"\s*(приложение\s*(?:n|№)\s*\d+)", v, re.I)
                head = m.group(1) if m else v.split("\n")[0]
                c.value = f"{head}\n{ref}"
            elif low.startswith("состав образующихся видов отходов"):
                c.value = APP1_TITLE
            elif low.startswith("для накопления и последующей передачи") and issues is not None:
                issues.append(Issue(
                    "warning", "бланк",
                    f"лист «{ws.title}»: графа «для накопления и последующей "
                    "передачи…» исключена приказом № 825 (с 01.09.2024) — "
                    "бланк устаревшей редакции, графа не заполняется"))


def _own_objects(ctx) -> dict[str, dict]:
    """№ 227, графы 8/13 Табл. 2 (поступление с собственных объектов /
    передача на собственные объекты): в WasteFlow таких полей нет — берём из
    ctx.extra["own_transfers"] = [{"fkko", "received_own", "transferred_own"}]."""
    return _by_fkko(ctx, "own_transfers")


def _row_data(w, n: int, ed: str, own: dict) -> dict:
    """Значения всех граф Прил. 1/2 для одного отхода (n — № п/п из Прил. 1:
    в бланке Прил. 3 ссылается на A6 Прил. 1, т.е. номер отхода сквозной)."""
    placed = _num(Decimal(str(w.placed_norm or 0)) + Decimal(str(w.placed_over or 0)), ed)
    p_st = _num(w.placed_storage, ed)
    p_bu = _num(w.placed_burial, ed)
    if placed is not None and p_st is None and p_bu is None:
        # модель не делит собственное размещение — типовой случай размещения
        # у полигона: относим к захоронению
        p_bu = placed
    o = own.get(fkko.norm(w.fkko_code) or w.fkko_code, {})
    return {
        "num": n,
        "name": w.name,
        "fkko_code": fkko.fmt(w.fkko_code) if w.fkko_code else "",
        "hazard_class": w.hazard_class or None,
        "origin": getattr(w, "origin", ""),
        "aggregate_state": getattr(w, "aggregate_state", ""),
        "composition": getattr(w, "composition", ""),
        "acc_start": _num(w.accumulated_start, ed),
        "acc_start_nakopl": _num(w.accumulated_start_nakopl, ed),
        "generated": _num(w.generated, ed),
        "received": _num(w.received, ed),
        "received_own": _num(o.get("received_own"), ed),
        "processed": _num(w.processed, ed),
        "used": _num(w.used, ed),
        "neutralized": _num(w.neutralized, ed),
        "transferred": _num(w.transferred, ed),
        "transferred_own": _num(o.get("transferred_own"), ed),
        "placed": placed,
        "placed_storage": p_st,
        "placed_burial": p_bu,
        "acc_end": _num(w.accumulated_end, ed),
        "acc_end_nakopl": _num(w.accumulated_end_nakopl, ed),
    }


def _write(ws, row: int, cols: dict, data: dict) -> None:
    for attr, col in cols.items():
        value = data.get(attr)
        if value in (None, ""):
            continue
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue            # объединение — не пишем (и не падаем)
        if is_row_formula(cell.value, row):
            continue            # агрегат по своей строке — посчитает Excel
        cell.value = value


def _fill_block(ws, header: int, cols: dict, rows: list[dict]) -> int:
    """Очистить область под шапкой и записать строки; вернуть конец области."""
    start, end = data_region(ws, header)
    _clear_rows(ws, start, min(end, ws.max_row))
    row = start
    for data in rows:
        if row > end:               # таблица кончилась — раздвигаем её,
            _insert_row(ws, row)    # не наезжая на блок-продолжение
            end += 1
        _write(ws, row, cols, data)
        row += 1
    return end


def fill(ctx, out_path: str | Path, sample: Path | None = None,
         issues: list | None = None) -> Path | None:
    """Заполнить бланк журнала. None — если образца нет (рисуем кодом).
    issues — куда складывать предупреждения о бланке (render_issues формы)."""
    from ecodoc.core import forms_registry
    sample = sample or forms_registry.sample_for(CODE)
    if sample is None or sample.suffix.lower() not in (".xlsx", ".xlsm"):
        return None

    ed = edition(ctx)
    wb = tx.open_template(sample)
    org = ctx.organization
    values = {"org_name": org.name or org.short_name, "inn": org.inn,
              "ogrn": org.ogrn, "period": period_text(ctx)}

    if wb.sheetnames:
        _fill_title(wb[wb.sheetnames[0]], ctx, values)

    receivers = _by_fkko(ctx, "waste_receivers")
    suppliers = _by_fkko(ctx, "waste_suppliers")
    own = _own_objects(ctx)
    if ed == ED_227 and own and issues is not None:
        issues.append(Issue(
            "warning", "бланк",
            "бланк редакции 2020 г.: граф 8/13 Таблицы 2 о собственных объектах "
            "(приказ № 227) в нём нет — эти данные в бланк не попали"))
    filled_sheets = 0
    for name in wb.sheetnames:
        ws = wb[name]
        _actualize_edition(ws, ed, issues)
        _drop_year_after_period_formula(ws)
        head = header_row(ws)
        if not head:
            continue
        header, cols = head
        # Прил. 3 — только переданные, Прил. 4 — только полученные отходы:
        # состав строк в таблицах разный (как и в форме, рисуемой кодом)
        kind = _sheet_kind(ws)
        keep = {"transferred": lambda w: _num(w.transferred, ed) is not None,
                "received": lambda w: _num(w.received, ed) is not None,
                }.get(kind, lambda w: True)
        rows: list[dict] = []
        for n, w in enumerate(ctx.wastes, 1):
            if not (w.fkko_code or w.name) or not keep(w):
                continue
            data = _row_data(w, n, ed, own)
            info: dict = {}
            key = fkko.norm(w.fkko_code) or w.fkko_code
            if kind == "transferred":
                info = receivers.get(key, {})
                data.update({
                    "t_processing": _num(w.transferred_processing, ed),
                    "t_util": _num(w.transferred_util, ed),
                    "t_neutral": _num(w.transferred_neutral, ed),
                    "t_storage": _num(w.transferred_storage, ed),
                    "t_burial": _num(w.transferred_burial, ed),
                    "partner": info.get("receiver", ""),
                    "license": info.get("license", ""),
                })
            elif kind == "received":
                info = suppliers.get(key, {})
                data["partner"] = info.get("supplier", "")
            if kind:
                data["contract"] = info.get("contract", "")
                data["contract_term"] = info.get("contract_term", "")
            rows.append(data)
        end = _fill_block(ws, header, cols, rows)
        # Прил. 2: блок «продолжение» (графы 8–16) — те же строки, тот же номер
        if not kind:
            cont = continuation_row(ws, end)
            if cont:
                _fill_block(ws, cont[0], cont[1], rows)
        filled_sheets += 1
    if not filled_sheets:
        return None
    return tx.save_as(wb, out_path)
