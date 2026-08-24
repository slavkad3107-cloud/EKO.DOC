"""Декларация о плате за НВОС (XML + Excel).

Форма зависит от отчётного года (см. editions.py):

* за 2025 год — Приказ Минприроды России от 10.12.2020 № 1043 в ред. Приказа
  от 29.04.2025 № 241 (действует с 01.09.2025; 9 разделов расчёта, итоговая
  часть со строками 020=021…028, 130–215);
* за 2026 год и далее — Приказ Минприроды России от 01.04.2026 № 182
  (действует с 01.09.2026 по 01.09.2031; добавлен Раздел 10 — искусственные
  грунты, признанные отходами, строка 029, сдвинуты коды 139–149; срок,
  выпавший на выходной, переносится на следующий рабочий день).

Срок представления — до 10 марта года, следующего за отчётным.

Разделы расчёта: Р1 выбросы стационарными; Р2/Р3 сжигание/рассеивание ПНГ
в пределах/сверх; Р4 сбросы; Р5 отходы производства; Р6 ТКО; Р7 побочные
продукты производства; Р8 вскрышные/вмещающие породы; Р9 побочные продукты
животноводства; Р10 искусственные грунты (только № 182).

Печатная книга повторяет состав бланка: титул (поля 1–17 по принятой
декларации + СЗПК/МСП по действующей редакции), лист «Информация о суммах
платы», расчёт «стр.2» с кодами строк по выбранной редакции, лист «Информация
об авансовых платежах», Раздел 1 в 18 граф, Раздел 4 в 20 (ред. 241) или 19
(№ 182) граф с группировкой по выпускам, Разделы 5–10 по графам бланка со
строками «ИТОГО» и «Всего по тем классам опасности…».

Правило: каждая графа и строка взяты из текста Приложения 2; где в модели
данных нет — графа печатается пустой, а validate() предупреждает.
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from ecodoc.core.models import Issue, Medium, ReportContext
from ecodoc.core.money import D, fmt_money
from ecodoc.core.registry import register
from ecodoc.render import xlsx
from ecodoc.render.xmlutil import el, write_tree
from ecodoc.reports.base import Report
from ecodoc.reports.declaration_nvos.calc import SECTIONS, PaymentResult, calculate
from ecodoc.reports.declaration_nvos.editions import (
    KBK, KIND_RU, KIND_SECTIONS, Edition, edition_for_year, rosprirodnadzor_for)

from lxml import etree
from openpyxl.utils import get_column_letter as _col

_BAND_RU = {"norm": "в пределах норматива", "limit": "в пределах лимита",
            "over": "сверх лимита/норматива"}

# Три способа исчисления авансового платежа (п. 4 ст. 16.4 ФЗ-7) — лист
# «Информация об авансовых платежах»; нужное отмечается знаком X
_ADV_METHODS = (
    ("quarter", "одна четвертая часть суммы платы за негативное воздействие "
                "на окружающую среду"),
    ("ndv", "одна четвертая часть суммы платы, определенная по НДВ, ВРВ "
            "(НДС, ВРС, лимитам на размещение отходов)"),
    ("pek", "по данным производственного экологического контроля"),
)
# синонимы значений ctx.extra['declaration']['advance_method']
_ADV_SYNONYM = {"quarter": "quarter", "1/4": "quarter", "q": "quarter",
                "ndv": "ndv", "vrv": "ndv", "limit": "ndv", "ндв": "ndv",
                "pek": "pek", "пэк": "pek"}

# Обратная совместимость: старые константы КБК (без пробелов — см. editions)
_KBK_AIR, _KBK_PNG, _KBK_WATER, _KBK_WASTE, _KBK_TKO = (
    KBK["air"], KBK["png"], KBK["water"], KBK["waste"], KBK["tko"])

# Характеристика объекта размещения отходов — четыре клетки бланка Разделов
# 5/6/8; отмечается знаком V. Ключ — значение extra['declaration']['oro']
# ['status'] (синонимы — для удобства ввода).
_ORO_STATUS = (
    ("groro", "Включен в государственный реестр объектов размещения отходов"),
    ("not_groro", "Не включен в государственный реестр объектов размещения "
                  "отходов"),
    ("no_impact", "Не оказывает негативное воздействие на окружающую среду"),
    ("decision", "Решение территориального органа Федеральной службы по "
                 "надзору в сфере природопользования об исключении негативного "
                 "воздействия на окружающую среду"),
)
_ORO_SYNONYM = {"groro": "groro", "гроро": "groro", "included": "groro",
                "not_groro": "not_groro", "нет": "not_groro",
                "no_impact": "no_impact", "decision": "decision"}


@register
class DeclarationNVOS(Report):
    code = "declaration-nvos"
    title = "Декларация о плате за НВОС"

    def __init__(self, context: ReportContext):
        super().__init__(context)
        self._calc: PaymentResult | None = None

    @property
    def calc(self) -> PaymentResult:
        if self._calc is None:
            self._calc = calculate(self.ctx)
        return self._calc

    @property
    def edition(self) -> Edition:
        """Редакция формы по отчётному году (2025 → № 241, далее → № 182)."""
        return edition_for_year(self.ctx.period.year)

    # ------------------------------------------------------------------ #
    def validate(self) -> list[Issue]:
        from ecodoc.core.validators import inn_valid, ogrn_valid
        issues: list[Issue] = []
        o = self.ctx.organization
        if not o.inn:
            issues.append(Issue("error", "ИНН", "не указан ИНН плательщика"))
        elif not inn_valid(o.inn):
            issues.append(Issue("error", "ИНН",
                                f"ИНН {o.inn} не проходит проверку контрольной "
                                f"суммы — вероятна опечатка"))
        if o.ogrn and not ogrn_valid(o.ogrn):
            issues.append(Issue("warning", "ОГРН",
                                f"ОГРН {o.ogrn} не проходит проверку — сверьте"))
        if not o.name:
            issues.append(Issue("error", "наименование", "не указано наименование организации"))
        if not o.oktmo:
            issues.append(Issue("warning", "ОКТМО", "не указан ОКТМО — обязателен для распределения платы"))
        if not self.ctx.period.year:
            issues.append(Issue("error", "период", "не указан отчётный год"))
        if not self.ctx.objects:
            issues.append(Issue("warning", "объекты", "не указан ни один объект НВОС"))
        else:
            from ecodoc.calendar.engine import category_of
            cats = {category_of(o) for o in self.ctx.objects}
            cats.discard("")
            if cats and cats == {"IV"}:
                issues.append(Issue(
                    "error", "категория",
                    "все объекты — IV категории: такие объекты плату за НВОС "
                    "не вносят и декларацию не подают (п. 1 ст. 16.1 ФЗ-7). "
                    "Проверьте категорию во вкладке ОБЪЕКТ."))
        if not (self.ctx.pollutants or self.ctx.wastes):
            issues.append(Issue("error", "данные", "нет ни выбросов/сбросов, ни отходов — нечего декларировать"))
        for w in self.calc.warnings:
            issues.append(Issue("warning", "ставка", w))
        issues.extend(self._blank_gaps())
        return issues

    def _blank_gaps(self) -> list[Issue]:
        """Графы бланка, для которых в модели нет данных: печатаются пустыми,
        а пользователя предупреждаем — чтобы пустая клетка не ушла в РПН
        незамеченной."""
        issues: list[Issue] = []
        decl = self._decl_extra()
        ed = self.edition
        if not self._rospr_name():
            issues.append(Issue(
                "warning", "титул, поле 2",
                "не определён территориальный орган Росприроднадзора: регион "
                "объекта не в справочнике — впишите наименование в "
                "extra['declaration']['rospr']"))
        bs = self.calc.by_section
        if D(bs.get("Р4", 0)) > 0 and not decl.get("water_permit"):
            issues.append(Issue(
                "warning", "Раздел 4",
                "не указаны реквизиты документа, на основании которого "
                "осуществляются сбросы (extra['declaration']['water_permit'])"))
        if D(bs.get("Р4", 0)) > 0 and not decl.get("water_outlet"):
            issues.append(Issue(
                "warning", "Раздел 4",
                "не указан выпуск (номер, ОКТМО выпуска) — "
                "extra['declaration']['water_outlet']"))
        waste_sections = [s for s in ("Р5", "Р6", "Р7", "Р8", "Р9", "Р10")
                          if D(bs.get(s, 0)) > 0]
        if waste_sections:
            if not decl.get("waste_permit"):
                issues.append(Issue(
                    "warning", "Разделы 5–10",
                    "не указаны реквизиты разрешительного документа на "
                    "размещение отходов (extra['declaration']['waste_permit'])"))
            oro = decl.get("oro") or {}
            if not (oro.get("name") or oro.get("number")):
                issues.append(Issue(
                    "warning", "Разделы 5–10",
                    "не указан объект размещения отходов (наименование, "
                    "№ ГРОРО, адрес, характеристика) — extra['declaration']['oro']"))
            if not decl.get("waste_limits"):
                issues.append(Issue(
                    "warning", "Разделы 5–10",
                    "не указаны установленные лимиты на размещение отходов "
                    "по видам (графа 5) — extra['declaration']['waste_limits']"))
        for kind in ("byprod", "rock", "livestock", "soil"):
            sects = KIND_SECTIONS[kind]
            if any(D(bs.get(s, 0)) > 0 for s in sects) and not self._kbk(kind):
                issues.append(Issue(
                    "warning", "стр.2",
                    f"КБК для «{KIND_RU[kind]}» в бланке не предзаполнен — "
                    f"укажите его в extra['declaration']['kbk']['{kind}']"))
        if D(bs.get("Р10", 0)) > 0 and not ed.has_section10:
            issues.append(Issue(
                "error", "Раздел 10",
                "искусственные грунты (Раздел 10) есть только в форме № 182 "
                "(декларация за 2026 год и позже); в форме ред. № 241 такой "
                "платы нет — проверьте waste_kind='soil' у отходов"))
        return issues

    # ------------------------------------------------------------------ #
    def render_xml(self, out_path: Path) -> Path:
        out_path = self._ensure_dir(out_path)
        o = self.ctx.organization
        per = self.ctx.period
        c = self.calc

        root = etree.Element("ДекларацияНВОС", version="0.1", форма=self.code,
                             редакция=self.edition.key)
        # ── плательщик ──
        plat = el(root, "Плательщик")
        el(plat, "Наименование", o.name)
        el(plat, "ИНН", o.inn)
        el(plat, "КПП", o.kpp)
        el(plat, "ОГРН", o.ogrn)
        el(plat, "ОКТМО", o.oktmo)
        el(plat, "Адрес", o.address)
        el(plat, "Руководитель", o.director_name)
        el(root, "ОтчётныйГод", per.year)

        # ── объекты ──
        objs = el(root, "ОбъектыНВОС")
        for ob in self.ctx.objects:
            x = el(objs, "Объект")
            el(x, "Код", ob.code)
            el(x, "Наименование", ob.name)
            el(x, "Категория", ob.category)
            el(x, "ОКТМО", ob.oktmo or o.oktmo)

        # ── строки расчёта ──
        lines_el = el(root, "СтрокиРасчёта")
        for ln in c.lines:
            x = el(lines_el, "Строка", среда=ln.medium, норматив=ln.band,
                   раздел=ln.section)
            el(x, "Код", ln.code)
            el(x, "Наименование", ln.name)
            el(x, "Масса", ln.mass)
            el(x, "Ставка", ln.rate)
            el(x, "КоэффИндексации", ln.k_ind)
            el(x, "КоэффНорматива", ln.k_band)
            el(x, "КоэффДоп", ln.k_extra)
            el(x, "Плата", ln.amount)

        # ── итоги ──
        tot = el(root, "Итоги")
        for key in SECTIONS:
            if key == "Р10" and not self.edition.has_section10:
                continue
            el(tot, "Раздел", c.by_section.get(key, 0), код=key, наим=SECTIONS[key])
        el(tot, "ПлатаВыбросы", c.total_air)
        el(tot, "ПлатаСбросы", c.total_water)
        el(tot, "ПлатаОтходы", c.total_waste)
        el(tot, "ПлатаВсего", c.total)

        write_tree(root, out_path)
        return out_path

    # ------------------------------------------------------------------ #
    def render_print(self, out_path: Path) -> Path:
        out_path = self._ensure_dir(out_path)
        wb = xlsx.new_workbook()
        bs = self.calc.by_section
        # порядок листов — как в печатной форме бланка: титул, «Информация о
        # суммах платы», расчёт, «Информация об авансовых платежах», разделы
        self._sheet_title(wb)          # стр.1 — титульный лист
        self._sheet_summary(wb)        # суммы платы по объектам НВОС
        self._sheet_calc(wb)           # стр.2 — расчёт суммы платы по разделам
        self._sheet_advances(wb)       # способ исчисления авансовых платежей
        self._sheet_section1(wb)       # Раздел 1 — выбросы, 18 граф по бланку
        if any(ln.section in ("Р2", "Р3") for ln in self.calc.lines):
            self._sheet_lines(wb, "Разделы 2-3 (ПНГ)", Medium.AIR.value,
                              sections=("Р2", "Р3"))
        self._sheet_section4(wb)       # Раздел 4 — сбросы по выпускам
        self._sheet_section5(wb)       # Раздел 5 — отходы производства, 27 граф
        # Разделы 6–10 — только при наличии платы: пустой лист ТКО/пород/грунтов
        # в принятых декларациях не печатают
        if D(bs.get("Р6", 0)) > 0 or any(ln.section == "Р6" for ln in self.calc.lines):
            self._sheet_section6(wb)
        if any(ln.section == "Р7" for ln in self.calc.lines):
            self._sheet_section7(wb)
        if any(ln.section == "Р8" for ln in self.calc.lines):
            self._sheet_section8(wb)
        if any(ln.section == "Р9" for ln in self.calc.lines):
            self._sheet_section9(wb)
        if self.edition.has_section10 and any(ln.section == "Р10"
                                              for ln in self.calc.lines):
            self._sheet_section10(wb)
        xlsx.save(wb, out_path)
        return out_path

    # ---- общие помощники ----------------------------------------------
    def _decl_extra(self) -> dict:
        """Пользовательские реквизиты декларации из ctx.extra['declaration']."""
        e = self.ctx.extra if isinstance(self.ctx.extra, dict) else {}
        return e.get("declaration") or {}

    def _rospr_name(self) -> str:
        """Поле 2 титула: наименование ТО РПН — явно заданное пользователем,
        иначе по региону (коду) первого объекта НВОС или по ОКТМО."""
        e = self.ctx.extra if isinstance(self.ctx.extra, dict) else {}
        explicit = str(self._decl_extra().get("rospr") or e.get("rospr") or "")
        if explicit:
            return explicit
        ob = self.ctx.objects[0] if self.ctx.objects else None
        from ecodoc.core.nvos import subject_code
        # ТО РПН подбирается по субъекту (78), а в базе префикс ОКТМО (40)
        region = (subject_code(getattr(ob, "region_code", "") if ob else "")
                  or subject_code(getattr(ob, "code", "") if ob else ""))
        if not region and ob and getattr(ob, "code", ""):
            region = str(ob.code).split("-")[0]   # «40-0178-…» → 40
        oktmo = (getattr(ob, "oktmo", "") if ob else "") or self.ctx.organization.oktmo
        return rosprirodnadzor_for(region, oktmo)

    def _kbk(self, kind: str) -> str:
        """КБК вида платы: пользовательский (extra['declaration']['kbk'])
        поверх справочника; хранится и печатается сплошными 20 цифрами."""
        user = (self._decl_extra().get("kbk") or {}).get(kind)
        val = str(user or KBK.get(kind, "") or "")
        return val.replace(" ", "")

    def _kind_base(self) -> dict:
        """Начисленная плата по видам платы (строки 040/060/080/100/120/132/…)."""
        bs = self.calc.by_section
        return {k: sum((D(bs.get(s, 0)) for s in KIND_SECTIONS[k]), Decimal("0"))
                for k in self.edition.kinds}

    def _kind_after_deduction(self) -> dict:
        """Строки 171–179: плата по виду минус затраты на мероприятия по
        снижению НВОС (строки 151–159); отрицательной быть не может."""
        decl = self._decl_extra()
        kinds = self.edition.kinds
        ded = {k: D(v) for k, v in (decl.get("deduction") or {}).items()
               if k in kinds}
        base = self._kind_base()
        out = {}
        for k in kinds:
            v = base[k] - ded.get(k, Decimal("0"))
            out[k] = v if v > 0 else Decimal("0")
        return out

    def _object_head(self, ws, r: int, last_col: str, extra_rows=()) -> int:
        """Шапка объекта НВОС на листах разделов: категория, наименование,
        код, ОКТМО, адрес — одинаковый блок в Разделах 4–8 и 10."""
        o = self.ctx.organization
        ob = self.ctx.objects[0] if self.ctx.objects else None
        split = "F"   # подпись A..F, значение G..last
        head = [
            ("Категория объекта, оказывающего негативное воздействие "
             "на окружающую среду", (ob.category if ob else "")),
            ("Наименование объекта", ob.name if ob else ""),
            ("Код объекта", ob.code if ob else ""),
            ("ОКТМО объекта", (ob.oktmo if ob else "") or o.oktmo),
            ("Адрес объекта", (ob.address if ob else "") or o.address),
        ] + list(extra_rows)
        for label, value in head:
            xlsx.merge(ws, f"A{r}:{split}{r}", label, align="left")
            xlsx.merge(ws, f"G{r}:{last_col}{r}",
                       value or "", align="left")
            r += 1
        return r

    def _oro_rows(self) -> list:
        """Блок объекта размещения отходов (Разделы 5/6/8): реквизиты
        разрешительного документа, наименование/№/адрес ОРО, характеристика
        (четыре клетки, нужное отмечено V). Данные — extra['declaration']."""
        decl = self._decl_extra()
        permit = decl.get("waste_permit") or {}
        if not isinstance(permit, dict):
            permit = {"number": str(permit)}
        oro = decl.get("oro") or {}
        status = _ORO_SYNONYM.get(str(oro.get("status") or "").strip().lower(), "")
        rows = [
            ("Реквизиты разрешительного документа, на основании которого "
             "осуществляется размещение отходов: №",
             str(permit.get("number") or "")),
            ("Срок действия", str(permit.get("valid_until") or permit.get("term") or "")),
            ("Наименование объекта размещения отходов", str(oro.get("name") or "")),
            ("Регистрационный номер объекта размещения отходов "
             "(в случае его присвоения)", str(oro.get("number") or "")),
            ("Адрес объекта размещения отходов", str(oro.get("address") or "")),
            ("Характеристика объекта размещения отходов:", ""),
        ]
        for key, label in _ORO_STATUS:
            rows.append(("    " + label, "V" if status == key else ""))
        return rows

    def _sign_block(self, ws, r: int, last_col: str) -> None:
        """Подпись исполнителя под листом раздела — как в бланке."""
        executor = (str(self._decl_extra().get("executor") or "")
                    or self.ctx.organization.director_name)
        xlsx.merge(ws, f"A{r}:{last_col}{r}", "Достоверность и полноту "
                   "сведений, указанных на данной странице, подтверждаю:",
                   border=False, align="left")
        xlsx.merge(ws, f"A{r+1}:{last_col}{r+1}",
                   f"Исполнитель ____________ {executor} (подпись, фамилия, "
                   "имя, отчество (при наличии))   цифрами: день, месяц, год "
                   "«____» __________ 20___ г.", border=False, align="left")

    # стр.1 — титульный лист официального бланка -----------------------
    def _sheet_title(self, wb):
        """Титул: поля 1–17 (нумерация — как в принятой декларации), текст
        полей — дословно по бланку выбранной редакции. Поле 1 — две клетки
        «первичный / уточненный» с отметкой V и номером корректировки; поле 9 —
        два числовых пропуска (страницы декларации и листы приложений)."""
        o = self.ctx.organization
        ed = self.edition
        e = self.ctx.extra if isinstance(self.ctx.extra, dict) else {}
        year = self.ctx.period.year or ""
        ws = wb.create_sheet("стр.1")
        xlsx.widths(ws, {"A": 4, "B": 46, "C": 30, "D": 26})
        xlsx.merge(ws, "B1:D1", ed.header, border=False, italic=True, size=9)
        xlsx.merge(ws, "B2:D2",
                   "ДЕКЛАРАЦИЯ О ПЛАТЕ ЗА НЕГАТИВНОЕ ВОЗДЕЙСТВИЕ НА ОКРУЖАЮЩУЮ СРЕДУ",
                   bold=True, border=False)
        xlsx.merge(ws, "B3:D3", f"за {year} г.", border=False, bold=True)
        # ИП и юрлицо заполняют разные поля бланка: у ИП пусты поля ЮЛ (3, 8),
        # у юрлица — поле ФИО предпринимателя (4). Признак — 12-значный ИНН.
        is_ip = len(str(o.inn or "").strip()) == 12
        decl = self._decl_extra()
        # вид документа: номер корректировки > 0 (или явный флаг) — уточнённая
        corr = str(decl.get("correction") or e.get("correction") or "").strip()
        kind = str(decl.get("doc_kind") or e.get("doc_kind") or "").lower()
        is_corr = bool(corr and corr not in ("0", "—")) or kind.startswith(
            ("уточ", "corr"))
        pages = str(decl.get("pages") or "")
        attach = str(decl.get("attachment_sheets") or "")
        szpk = decl.get("szpk") or {}
        # (номер, подпись бланка, значение C, значение D)
        rows = [
            ("1", "Вид документа (нужное отметить знаком V):",
             f"первичный [{' ' if is_corr else 'V'}]",
             f"уточненный [{'V' if is_corr else ' '}] / "
             f"номер корректировки {corr if is_corr else '—'}"),
            ("2", "Настоящая Декларация представляется в: (наименование "
                  "территориального органа Федеральной службы по надзору "
                  "в сфере природопользования)", self._rospr_name(), None),
            ("3", "Организационно-правовая форма юридического лица и его "
                  "полное наименование", "" if is_ip else o.name, None),
            ("4", "Фамилия, имя, отчество (при наличии) индивидуального "
                  "предпринимателя", o.name if is_ip else "", None),
            ("5", "Адрес юридического лица в пределах места нахождения "
                  "юридического лица - для юридического лица, адрес регистрации "
                  "по месту жительства - для индивидуального предпринимателя:",
             o.address, None),
            ("6", "Код города и номер контактного телефона:", o.phone, None),
            ("7", "Идентификационный номер налогоплательщика", o.inn, None),
            ("8", "Код причины постановки на учет", "" if is_ip else o.kpp, None),
        ]
        if ed.title_has_msp:
            # поле появилось в форме № 182; признак МСП в модели — extra
            msp = decl.get("is_msp")
            if msp is None:
                msp = e.get("is_msp")
            rows.append(("", "Относится к субъектам малого и среднего "
                             "предпринимательства (ст. 4 ФЗ от 24.07.2007 "
                             "№ 209-ФЗ)",
                         "" if msp is None else ("V" if msp else ""), None))
        rows += [
            ("9", "Настоящая Декларация составлена на ___ страницах с "
                  "приложением подтверждающих документов или их копий на "
                  "___ листах",
             f"страниц: {pages}", f"листов приложений: {attach}"),
            # СЗПК — сторона соглашения о защите и поощрении капиталовложений
            # (ФЗ от 01.04.2020 № 69-ФЗ); заполняется только при наличии
            ("", "Признак СЗПК (является стороной СЗПК)",
             "V" if szpk.get("is_party") else "", None),
            ("", "Срок действия соглашения с … по … (цифрами: день, месяц, год)",
             str(szpk.get("from") or ""), str(szpk.get("to") or "")),
            ("", "Номер соглашения", str(szpk.get("number") or ""), None),
            ("10", "Руководитель юридического лица или лицо, уполномоченное на "
                   "осуществление действий на подписание настоящей Декларации "
                   "от имени юридического лица, либо индивидуальный "
                   "предприниматель (фамилия, имя, отчество (при наличии), "
                   "должность)",
             f"{'' if is_ip else o.director_position} "
             f"{o.director_name or (o.name if is_ip else '')}".strip(),
             "(подпись)  «____» __________ 20___ г."),
            # поля 11–17 бланка: подписи и служебные отметки. Данных для них в
            # модели нет — берём из ctx.extra['declaration'] по понятным
            # ключам, иначе печатаем пустую графу (как в принятых отчётах)
            ("11", "Руководитель обособленного подразделения организации "
                   "(по доверенности) (фамилия, имя, отчество (при наличии))",
             str(decl.get("branch_head") or ""), "цифрами: день, месяц, год"),
            ("12", "Исполнитель (фамилия, имя, отчество (при наличии))",
             str(decl.get("executor") or ""),
             "(подпись)  «____» __________ 20___ г."),
            ("13", "Главный бухгалтер (при наличии) (фамилия, имя, отчество "
                   "(при наличии))", str(decl.get("accountant") or ""),
             "(подпись)  «____» __________ 20___ г."),
            ("14", "М.П. (при наличии)", "", None),
            ("", "Сведения территориального органа Федеральной службы по "
                 "надзору в сфере природопользования", "", None),
            ("15", "Настоящая Декларация представлена: (цифрами: день, месяц, год)",
             "", None),
            ("16", "уполномоченным представителем / по почте (нужное отметить "
                   "знаком X) на ___ страницах. Зарегистрирован за №",
             "", None),
            ("17", "Фамилия, имя, отчество (при наличии) и должность "
                   "должностного лица территориального органа Федеральной "
                   "службы по надзору в сфере природопользования, подпись",
             "", None),
        ]
        r = 5
        for num, label, c_val, d_val in rows:
            xlsx.cell(ws, f"A{r}", num)
            xlsx.cell(ws, f"B{r}", label, align="left")
            if d_val is None:
                xlsx.merge(ws, f"C{r}:D{r}", c_val or "", align="left")
            else:
                xlsx.cell(ws, f"C{r}", c_val or "", align="left")
                xlsx.cell(ws, f"D{r}", d_val or "", align="left")
            r += 1

    # «Информация о суммах платы, подлежащих внесению в бюджет» ---------
    def _sheet_summary(self, wb):
        """Стр. 3–4 бланка: категория объекта НВОС → сумма платы к внесению.

        По бланку — одна строка данных на объект (категория может быть
        пустой), строка «ИТОГО» и блок подписи исполнителя. Плата в расчёте
        к объекту не привязана, поэтому при одном объекте вся сумма — его,
        при нескольких графа суммы остаётся пустой (источника разбивки нет),
        итог при этом верен.
        """
        o = self.ctx.organization
        decl = self._decl_extra()
        total = sum(self._kind_after_deduction().values(), Decimal("0"))
        ws = wb.create_sheet("Информация о суммах платы")
        xlsx.widths(ws, {"A": 50, "B": 28})
        xlsx.merge(ws, "A1:B1", "Информация о суммах платы, подлежащих "
                   "внесению в бюджет", bold=True)
        xlsx.cell(ws, "A2", "Категория объекта, оказывающего негативное "
                  "воздействие на окружающую среду", bold=True, fill=True)
        xlsx.cell(ws, "B2", "Сумма платы, подлежащая внесению в бюджет",
                  bold=True, fill=True)
        r = 3
        objs = self.ctx.objects or [None]
        single = len(objs) == 1
        for ob in objs:
            xlsx.cell(ws, f"A{r}", (getattr(ob, "category", "") or "") if ob else "")
            xlsx.cell(ws, f"B{r}", fmt_money(total) if single else "")
            r += 1
        xlsx.cell(ws, f"A{r}", "ИТОГО", bold=True)
        xlsx.cell(ws, f"B{r}", fmt_money(total), bold=True)
        r += 2
        executor = str(decl.get("executor") or "") or o.director_name
        xlsx.cell(ws, f"A{r}", "Достоверность и полноту сведений, указанных "
                  "на данных страницах, подтверждаю:", border=False, align="left")
        xlsx.cell(ws, f"A{r+1}", f"Исполнитель ____________ {executor} "
                  "(подпись, Ф.И.О)", border=False, align="left")
        xlsx.cell(ws, f"A{r+2}", "Дата (цифрами: день, месяц, год): "
                  "«____» __________ 20___ г.", border=False, align="left")

    # «Информация об авансовых платежах, подлежащих внесению в бюджет» --
    def _sheet_advances(self, wb):
        """Лист бланка: способ исчисления авансового платежа по каждому виду
        платы (строки 010–060), нужное отмечается знаком X.

        Способ в модели не хранится — берётся из
        ctx.extra['declaration']['advance_method'] =
        {'air': 'quarter'|'ndv'|'pek', ...}; без него графы пустые.
        Субъекты МСП авансы не вносят — у них лист остаётся незаполненным.
        """
        o = self.ctx.organization
        decl = self._decl_extra()
        methods = {k: _ADV_SYNONYM.get(str(v).strip().lower(), "")
                   for k, v in (decl.get("advance_method") or {}).items()}
        ws = wb.create_sheet("Авансовые платежи")
        xlsx.widths(ws, {"A": 66, "B": 10, "C": 22})
        xlsx.merge(ws, "A1:C1", "Информация об авансовых платежах, "
                   "подлежащих внесению в бюджет", bold=True)
        xlsx.cell(ws, "A2", "Показатели", bold=True, fill=True)
        xlsx.cell(ws, "B2", "Строки", bold=True, fill=True)
        xlsx.cell(ws, "C2", "Значения показателей", bold=True, fill=True)
        oktmo = ((self.ctx.objects[0].oktmo if self.ctx.objects else "")
                 or o.oktmo or "")
        r = 3
        xlsx.cell(ws, f"A{r}", "Код по ОКТМО объекта, оказывающего негативное "
                  "воздействие на окружающую среду (объекта размещения отходов)",
                  align="left")
        xlsx.cell(ws, f"B{r}", "010")
        xlsx.cell(ws, f"C{r}", oktmo)
        r += 1
        xlsx.merge(ws, f"A{r}:C{r}", "Выбранный способ исчисления авансового "
                   "платежа, в том числе:", align="left")
        r += 1
        codes = {"air": "020", "png": "030", "water": "040",
                 "waste": "050", "tko": "060"}
        labels = {
            "air": "авансовый платеж за выбросы, в том числе:",
            "png": "за выбросы ПНГ",
            "water": "авансовый платеж за сбросы",
            "waste": "авансовый платеж за размещение отходов производства",
            "tko": "авансовый платеж за размещение ТКО",
        }
        for k in self.edition.five:
            xlsx.cell(ws, f"A{r}", labels[k], align="left", bold=True)
            xlsx.cell(ws, f"B{r}", codes[k], bold=True)
            xlsx.cell(ws, f"C{r}", "(нужное отметить знаком X)")
            r += 1
            for mk, mlabel in _ADV_METHODS:
                xlsx.cell(ws, f"A{r}", "  " + mlabel, align="left")
                xlsx.cell(ws, f"B{r}", "")
                xlsx.cell(ws, f"C{r}", "X" if methods.get(k) == mk else "")
                r += 1
        r += 1
        executor = str(decl.get("executor") or "") or o.director_name
        xlsx.cell(ws, f"A{r}", f"Исполнитель ____________ {executor} "
                  "(подпись, Ф.И.О)", border=False, align="left")

    # стр.2 — расчёт суммы платы, коды строк по редакции -----------------
    def _sheet_calc(self, wb):
        """Таблица «Расчет суммы платы, подлежащей внесению в бюджет».

        Коды строк — строго по тексту Приложения 2 выбранной редакции
        (editions.py): 020 = 021+…+028 (ред. 241) или +029 (№ 182); блоки
        «КБК / ОКТМО / сумма / подстроки» по каждому виду платы; 150 мероприятия,
        160 с учётом корректировки, 170 к внесению, 180 зачтено (181–185),
        186 номер решения о зачёте, 190 авансы по кварталам (191–195),
        200 итог к внесению, 210 к возврату/зачёту (211–215).
        """
        o = self.ctx.organization
        c = self.calc
        ed = self.edition
        kinds = ed.kinds
        bs = c.by_section
        # суммы по (раздел, корзина) для подстрок 041/042/043 и т.п.
        sb: dict = {}
        for ln in c.lines:
            sb[(ln.section, ln.band)] = sb.get((ln.section, ln.band), Decimal("0")) + D(ln.amount)

        def band(sect, b):
            return fmt_money(sb.get((sect, b), Decimal("0")))

        decl = self._decl_extra()

        def okt(kind: str) -> str:
            """ОКТМО для блока: свой у объекта размещения, иначе плательщика."""
            return str((decl.get("oktmo") or {}).get(kind) or o.oktmo or "")

        base = self._kind_base()
        # вычеты (затраты на мероприятия по снижению НВОС), зачёт переплаты и
        # авансовые платежи в модели не считаются — их вносит пользователь в
        # ctx.extra['declaration']; при отсутствии в бланке остаются нули
        ded = {k: D(v) for k, v in (decl.get("deduction") or {}).items()
               if k in kinds}
        # по бланку вычет за ПНГ/побочные продукты/породы/животноводство/
        # грунты «равен 0» — пользовательское значение там не печатаем
        for k in ("png", "byprod", "rock", "livestock", "soil"):
            ded.pop(k, None)
        ded_total = sum(ded.values(), Decimal("0"))
        pay = self._kind_after_deduction()      # 171–179 = сумма − вычет
        pay_total = sum(pay.values(), Decimal("0"))
        off = {k: D(v) for k, v in (decl.get("offset") or {}).items()
               if k in ed.five}                 # зачтённая переплата, 181–185
        off_total = sum(off.values(), Decimal("0"))
        # авансы: {'air': {'q1':…,'q2':…,'q3':…}} или {'air': сумма}; старый
        # плоский вид {'q1': …} тоже принимается — без привязки к виду платы
        # такие суммы попадают только в итог строки 190
        zero_q = {"q1": Decimal("0"), "q2": Decimal("0"), "q3": Decimal("0")}
        adv_q = {k: dict(zero_q) for k in ed.five}
        adv_extra = {k: Decimal("0") for k in ed.five}  # сумма без кварталов
        adv_flat = Decimal("0")
        for key, val in (decl.get("advance") or {}).items():
            if key in ed.five:
                if isinstance(val, dict):
                    for q in zero_q:
                        adv_q[key][q] += D(val.get(q) or 0)
                else:
                    adv_extra[key] += D(val or 0)
            elif key in zero_q:
                adv_flat += D(val or 0)
        adv_kind = {k: sum(adv_q[k].values(), Decimal("0")) + adv_extra[k]
                    for k in ed.five}
        adv_total = sum(adv_kind.values(), Decimal("0")) + adv_flat
        # итог к внесению (200) и к возврату/зачёту (210) по видам платы —
        # по сноскам бланка: 20N = 17N − 18N − 19N (не меньше 0), 21N —
        # обратная разница; для «новых» видов 206–209 = 176–179 (авансов и
        # зачёта по ним в бланке нет). Незакреплённые за видом авансы
        # (adv_flat) учитываются только в итогах 200/210.
        fin, ret = {}, {}
        for k in kinds:
            if k in ed.five:
                d = pay[k] - off.get(k, Decimal("0")) - adv_kind[k]
                fin[k] = d if d > 0 else Decimal("0")
                ret[k] = -d if d < 0 else Decimal("0")
            else:
                fin[k] = pay[k]
        diff_total = pay_total - off_total - adv_total
        fin_total = diff_total if diff_total > 0 else Decimal("0")
        ret_total = -diff_total if diff_total < 0 else Decimal("0")

        ws = wb.create_sheet("стр.2")
        xlsx.widths(ws, {"A": 10, "B": 64, "C": 26, "D": 18})
        xlsx.merge(ws, "A1:D1", "Расчёт суммы платы, подлежащей внесению в бюджет "
                   f"({ed.header})", bold=True, border=False)
        xlsx.cell(ws, "A3", "Код строки", bold=True, fill=True)
        xlsx.cell(ws, "B3", "Показатели", bold=True, fill=True, align="left")
        xlsx.cell(ws, "C3", "КБК / ОКТМО", bold=True, fill=True)
        xlsx.cell(ws, "D3", "Сумма, руб.", bold=True, fill=True)
        n = len(kinds)
        # ссылки «(040)», «(132)» в подстроках 021–029 — код строки суммы вида
        sum_row = {"air": "040", "png": "060", "water": "080", "waste": "100",
                   "tko": "120"}
        for k, blk in ed.blocks.items():
            sum_row[k] = blk[2]
        plus = "+".join(f"02{i}" for i in range(1, n + 1))
        # (код, показатель, значение-в-C (КБК/ОКТМО), значение-в-D (сумма))
        rows = [
            ("010", "Код по ОКТМО объекта, оказывающего негативное воздействие "
                    "на окружающую среду", o.oktmo or "", None),
            ("020", "Сумма платы, исчисленная без учета корректировки ее "
                    f"размера, всего (020 = {plus})", "", fmt_money(c.total)),
        ]
        for i, k in enumerate(kinds, start=1):
            rows.append((f"02{i}", f"  {KIND_RU[k]} ({sum_row[k]})", "",
                         fmt_money(base[k])))
        rows += [
            ("030", "КБК: плата за выбросы", self._kbk("air"), None),
            ("031", "ОКТМО", okt("air"), None),
            ("040", "Сумма платы за выбросы всего (040 = 041+042+043)", "", fmt_money(base["air"])),
            ("041", "  плата за выбросы в пределах НДВ, ТН", "", band("Р1", "norm")),
            ("042", "  плата за выбросы в пределах ВРВ", "", band("Р1", "limit")),
            ("043", "  плата за выбросы, превышающие установленные НДВ, ТН, ВРВ", "", band("Р1", "over")),
            ("050", "КБК: плата за выбросы ПНГ", self._kbk("png"), None),
            ("051", "ОКТМО", okt("png"), None),
            ("060", "Сумма платы за выбросы ПНГ, всего (060 = 061+062+063)", "", fmt_money(base["png"])),
            ("061", "  плата за выбросы ПНГ в пределах НДВ, ТН", "", band("Р2", "norm")),
            ("062", "  плата за выбросы ПНГ в пределах ВРВ", "", band("Р2", "limit")),
            ("063", "  плата за выбросы ПНГ сверх НДВ, ТН, ВРВ", "", band("Р3", "over")),
            ("070", "КБК: плата за сбросы", self._kbk("water"), None),
            ("071", "ОКТМО", okt("water"), None),
            ("080", "Сумма платы за сбросы, всего (080 = 081+082+083)", "", fmt_money(base["water"])),
            ("081", "  плата за сбросы в пределах НДС, ТН", "", band("Р4", "norm")),
            ("082", "  плата за сбросы в пределах ВРС", "", band("Р4", "limit")),
            ("083", "  плата за сбросы, превышающие установленные НДС, ТН, ВРС", "", band("Р4", "over")),
            ("090", "КБК: плата за размещение отходов производства", self._kbk("waste"), None),
            ("091", "Код по ОКТМО объекта размещения отходов", okt("waste"), None),
            ("100", "Сумма платы за размещение отходов производства, всего "
                    "(100 = 101+102)", "", fmt_money(base["waste"])),
            ("101", "  плата за размещение отходов производства в пределах "
                    "установленного лимита на их размещение", "", band("Р5", "norm")),
            ("102", "  плата за размещение отходов производства сверх "
                    "установленного лимита на их размещение", "", band("Р5", "over")),
            ("110", "КБК: плата за размещение ТКО", self._kbk("tko"), None),
            ("111", "Код по ОКТМО объекта размещения ТКО", okt("tko"), None),
            ("120", "Сумма платы за размещение ТКО, всего (120 = 121+122+123)",
             "", fmt_money(base["tko"])),
            # 121 — плата регионального оператора за принятые ТКО: в расчёте
            # не участвует, источника в модели нет — значение только из
            # ctx.extra['declaration']['tko_accepted'], иначе 0,00
            ("121", "  плата за размещение принятых ТКО", "",
             fmt_money(D(decl.get("tko_accepted") or 0))),
            ("122", "  плата за размещение ТКО в пределах установленного "
                    "лимита на их размещение", "", band("Р6", "norm")),
            ("123", "  плата за размещение ТКО сверх установленного лимита "
                    "на их размещение", "", band("Р6", "over")),
        ]
        # блоки побочных продуктов / пород / животноводства / грунтов —
        # коды по редакции (130–145 в ред. 241, 130–149 в № 182)
        oktmo_label = {
            "byprod": "Код по ОКТМО объекта размещения, в том числе складирования "
                      "побочных продуктов производства, признанных отходами",
            "rock": "Код по ОКТМО объекта размещения вскрышных и вмещающих "
                    "горных пород, признанных отходами",
            "livestock": "Код по ОКТМО места выявления нарушения или объекта "
                         "размещения отходов",
            "soil": "Код по ОКТМО объекта размещения, в том числе складирования "
                    "искусственных грунтов, признанных отходами",
        }
        sum_label = {
            "byprod": "Сумма платы за размещение, в том числе складирование "
                      "побочных продуктов производства, признанных отходами",
            "rock": "Сумма платы за размещение вскрышных и вмещающих горных "
                    "пород, признанных отходами",
            "livestock": "Сумма платы за размещение побочных продуктов "
                         "животноводства, признанных отходами",
            "soil": "Сумма платы за размещение, в том числе складирование "
                    "искусственных грунтов, признанных отходами",
        }
        for k in kinds:
            if k not in ed.blocks:
                continue
            kbk_code, okt_code, sum_code, subs = ed.blocks[k]
            sect = KIND_SECTIONS[k][0]
            formula = (f" ({sum_code} = {'+'.join(s[0] for s in subs)})"
                       if subs else "")
            rows.append((kbk_code, f"КБК: {KIND_RU[k]}", self._kbk(k), None))
            rows.append((okt_code, oktmo_label[k], okt(k), None))
            rows.append((sum_code, sum_label[k] + formula, "", fmt_money(base[k])))
            for code, (b, label) in subs:
                rows.append((code, "  " + label, "", band(sect, b)))
        # ── итоговая часть: 150 / 160 / 170 / 180 / 186 / 190 / 200 / 210 ──
        def codes(prefix):
            return "+".join(f"{prefix}{i}" for i in range(1, n + 1))

        zero_note = {"png": " (равна 0)", "byprod": " (равна 0)",
                     "rock": " (равна 0)", "livestock": " (равна 0)",
                     "soil": " (равна 0)"}
        rows.append(("150", "Сумма средств на выполнение мероприятий по снижению "
                            "негативного воздействия на окружающую среду, всего "
                            f"(150 = {codes('15')})", "", fmt_money(ded_total)))
        for i, k in enumerate(kinds, start=1):
            rows.append((f"15{i}", f"  {KIND_RU[k].replace('плата', 'платы', 1)}"
                         f"{zero_note.get(k, '')}", "", fmt_money(ded.get(k, 0))))
        rows.append(("160", "Сумма платы, исчисленная с учетом корректировки ее "
                            f"размера (160 = {codes('16')})", "",
                     fmt_money(sum(base.values(), Decimal("0")))))
        for i, k in enumerate(kinds, start=1):
            # по бланку 161–165 — графы итогов разделов, 162/166–169 «равно
            # строке 022/026–029»: т.е. начисленная плата без вычета
            rows.append((f"16{i}", f"  {KIND_RU[k].replace('плата', 'платы', 1)}",
                         "", fmt_money(base[k])))
        rows.append(("170", "Сумма платы, подлежащая внесению в бюджет, всего "
                            f"(170 = {codes('17')})", "", fmt_money(pay_total)))
        for i, k in enumerate(kinds, start=1):
            rows.append((f"17{i}", f"  {KIND_RU[k]} (строка {sum_row[k]} − "
                         f"строка 15{i})", "", fmt_money(pay[k])))
        rows.append(("180", "Сумма платы, зачтенная в предыдущем отчетном периоде "
                            "в счет будущего отчетного периода (авансовых "
                            "платежей текущего отчетного периода), всего "
                            "(180 = 181+182+183+184+185)", "", fmt_money(off_total)))
        five_ru = {"air": "плата за выбросы", "png": "плата за выбросы ПНГ",
                   "water": "плата за сбросы",
                   "waste": "плата за размещение отходов",
                   "tko": "плата за размещение ТКО"}
        for i, k in enumerate(ed.five, start=1):
            rows.append((f"18{i}", f"  {five_ru[k]}", "", fmt_money(off.get(k, 0))))
        rows.append(("186", "Номер Решения о зачете сумм излишне уплаченной "
                            "(взысканной) платы за негативное воздействие на "
                            "окружающую среду в счет будущего отчетного периода",
                     str(decl.get("offset_decision") or ""), None))
        rows.append(("190", "Сведения о суммах внесенных авансовых платежей, "
                            "всего (190 = 191+192+193+194+195)", "",
                     fmt_money(adv_total)))
        adv_ru = {"air": "за выбросы", "png": "за выбросы ПНГ",
                  "water": "за сбросы", "waste": "за размещение отходов производства",
                  "tko": "за размещение ТКО"}
        for i, k in enumerate(ed.five, start=1):
            rows.append((f"19{i}", f"  {adv_ru[k]}", "", fmt_money(adv_kind[k])))
            # в бланке подстроки кварталов стоят в графе «Строки» (кодов у них
            # нет — вместо кода печатается «1 квартал» и т.д.)
            for q_label, q in (("1 квартал", "q1"), ("2 квартал", "q2"),
                               ("3 квартал", "q3")):
                rows.append((q_label, "", "", fmt_money(adv_q[k][q])))
        rows.append(("200", "Итоговая сумма платы для внесения за отчетный "
                            f"период, всего (200 = {codes('20')})", "",
                     fmt_money(fin_total)))
        for i, k in enumerate(kinds, start=1):
            rows.append((f"20{i}", f"  {KIND_RU[k].replace('плата', 'платы', 1)}",
                         "", fmt_money(fin[k])))
        rows.append(("210", "Итоговая сумма платы для возврата и/или зачета, "
                            "всего (210 = 211+212+213+214+215)", "",
                     fmt_money(ret_total)))
        for i, k in enumerate(ed.five, start=1):
            rows.append((f"21{i}", f"  {KIND_RU[k].replace('плата', 'платы', 1)}",
                         "", fmt_money(ret.get(k, 0))))
        r = 4
        for code, label, c_val, d_val in rows:
            bold = code in ("010", "020", "170", "200")
            xlsx.cell(ws, f"A{r}", code, bold=bold)
            xlsx.cell(ws, f"B{r}", label, align="left", bold=bold)
            xlsx.cell(ws, f"C{r}", c_val)
            xlsx.cell(ws, f"D{r}", d_val if d_val is not None else "", bold=bold)
            r += 1
        xlsx.cell(ws, f"A{r+1}", f"Коды строк — по тексту Приложения 2 ({ed.npa}). "
                  "Разделы 2–3 (ПНГ) — по флагу is_flare; строка 025/120 — ТКО "
                  "(ФККО «7 3…»).", border=False, italic=True, size=9, align="left")

    # Раздел 1 — выбросы стационарными источниками, 18 граф по бланку ---
    def _sheet_section1(self, wb):
        """Раздел 1 в составе и порядке граф официального бланка.

        Одна строка = одно вещество (Pollutant), массы раскладываются по
        графам 6/7/8 (НДВ,ТН / ВРВ / сверх), суммы платы — по графам 15/16/17.
        Установленные выбросы (гр. 3–4) и стационарный источник в модели не
        хранятся — берутся из ctx.extra['declaration'] (ключи 'air_limits',
        'stationary_source'), иначе графы остаются пустыми, как в принятых
        отчётах.
        """
        from ecodoc.core.refdata import coefficients
        band_k = coefficients()["band"]
        o = self.ctx.organization
        ob = self.ctx.objects[0] if self.ctx.objects else None
        decl = self._decl_extra()
        lines = [ln for ln in self.calc.lines if ln.section == "Р1"]

        ws = wb.create_sheet("Раздел 1 (выбросы)")
        xlsx.merge(ws, "A1:R1", "Раздел 1. Расчёт суммы платы за выбросы "
                   "загрязняющих веществ в атмосферный воздух стационарными "
                   "источниками", bold=True, border=False)
        # шапка объекта — как на листе бланка
        head = [
            ("Категория объекта, оказывающего негативное воздействие "
             "на окружающую среду", (ob.category if ob else "")),
            ("Наименование объекта", ob.name if ob else ""),
            ("Код объекта", ob.code if ob else ""),
            ("Адрес места нахождения объекта",
             (ob.address if ob else "") or o.address),
            ("Реквизиты документа, на основании которого осуществляются "
             "выбросы", str(decl.get("air_permit") or "")),
        ]
        r = 2
        for label, value in head:
            xlsx.merge(ws, f"A{r}:F{r}", label, align="left")
            xlsx.merge(ws, f"G{r}:R{r}", value, align="left")
            r += 1
        r += 1
        headers = [
            "N п/п", "Наименование загрязняющего вещества",
            "Установленные выбросы, т: НДВ, ТН",
            "Установленные выбросы, т: ВРВ",
            "Фактический выброс, всего, т",
            "в т.ч. в пределах НДВ, ТН", "в т.ч. в пределах ВРВ",
            "в т.ч. сверх НДВ, ТН, ВРВ", "Ставка платы, руб./т",
            "Кнд", "Квр", "Кср / Кпр", "Кот", "Кинд",
            "Сумма платы в пределах НДВ, ТН, руб. (гр.6×9×10×13×14)",
            "Сумма платы в пределах ВРВ, руб. (гр.7×9×11×13×14)",
            "Сумма платы сверх ВРВ, НДВ, ТН, руб. (гр.8×9×12×13×14)",
            "Сумма платы, всего, руб. (гр.15+16+17)"]
        # ширины через xlsx.widths: header_row(widths=…) читает ячейку строки
        # 1, а она здесь объединена под заголовок листа
        xlsx.widths(ws, dict(zip(
            "ABCDEFGHIJKLMNOPQR",
            [6, 26, 11, 11, 11, 11, 11, 11, 11, 7, 7, 8, 7, 7, 13, 13, 13, 14])))
        xlsx.header_row(ws, r, headers)
        r += 1
        xlsx.header_row(ws, r, [str(i) for i in range(1, 19)])  # номера граф
        r += 1
        per, order = self._group_by_substance(lines)
        # заголовок стационарного источника: в модели источника выброса нет
        # (Pollutant.source — это происхождение данных), реквизиты берутся из
        # ctx.extra; без них печатаем общий блок с пустыми реквизитами
        src = decl.get("stationary_source") or {}
        src_name = str(src.get("name") or "")
        src_num = str(src.get("number") or "")
        xlsx.merge(ws, f"A{r}:R{r}",
                   f"Стационарный источник {src_name} № {src_num or '____'}",
                   align="left", bold=True)
        r += 1
        xlsx.merge(ws, f"A{r}:R{r}", "ОКТМО стационарного источника "
                   + str(src.get("oktmo") or (ob.oktmo if ob else "")
                         or o.oktmo or ""), align="left")
        r += 1
        limits = decl.get("air_limits") or {}
        tot = {15: Decimal("0"), 16: Decimal("0"), 17: Decimal("0"),
               18: Decimal("0")}
        n = 0
        for key in order:
            n += 1
            d = per[key]
            lim = limits.get(key[0]) or {}
            m = {b: d["mass"].get(b, Decimal("0")) for b in ("norm", "limit", "over")}
            a = {b: d["amt"].get(b, Decimal("0")) for b in ("norm", "limit", "over")}
            total = a["norm"] + a["limit"] + a["over"]
            # коэффициенты полос: где полосы не было — справочное значение
            k10 = d["k"].get("norm", D(band_k["norm"]))
            k11 = d["k"].get("limit", D(band_k["limit"]))
            k12 = d["k"].get("over", D(band_k["over"]))
            xlsx.data_row(ws, r, [
                n, key[1], str(lim.get("ndv") or ""), str(lim.get("vrv") or ""),
                float(m["norm"] + m["limit"] + m["over"]),
                float(m["norm"]), float(m["limit"]), float(m["over"]),
                float(d["rate"]), float(k10), float(k11), float(k12),
                float(d["k_ot"]), float(d["k_ind"]),
                fmt_money(a["norm"]), fmt_money(a["limit"]),
                fmt_money(a["over"]), fmt_money(total)])
            r += 1
            tot[15] += a["norm"]
            tot[16] += a["limit"]
            tot[17] += a["over"]
            tot[18] += total
        for label in ("Итого:", "Итого по стационарным источникам:"):
            xlsx.merge(ws, f"A{r}:N{r}", label, align="left", bold=True)
            for col, key in (("O", 15), ("P", 16), ("Q", 17), ("R", 18)):
                xlsx.cell(ws, f"{col}{r}", fmt_money(tot[key]), bold=True)
            r += 1
        # корректировка размера платы программой не рассчитывается —
        # строки бланка печатаются пустыми
        xlsx.merge(ws, f"A{r}:N{r}", "Всего по всем стационарным источникам "
                   "по тем загрязняющим веществам, по которым осуществляется "
                   "корректировка размера платы", align="left")
        for col in ("O", "P", "Q", "R"):
            xlsx.cell(ws, f"{col}{r}", "")
        r += 1
        xlsx.merge(ws, f"A{r}:R{r}", "в том числе:", align="left")

    @staticmethod
    def _group_by_substance(lines):
        """Строки расчёта (по полосам) обратно в «одно вещество — одна строка»."""
        per: dict = {}
        order: list = []
        for ln in lines:
            key = (ln.code, ln.name)
            if key not in per:
                per[key] = {"mass": {}, "amt": {}, "k": {},
                            "rate": ln.rate, "k_ot": ln.k_extra,
                            "k_ind": ln.k_ind}
                order.append(key)
            per[key]["mass"][ln.band] = ln.mass
            per[key]["amt"][ln.band] = ln.amount
            per[key]["k"][ln.band] = ln.k_band
        return per, order

    # Раздел 4 — сбросы в водные объекты, по выпускам ---------------------
    def _sheet_section4(self, wb):
        """Раздел 4 по графам бланка: 20 граф в ред. 241 (с графой 13 «Кп» —
        коэффициент пересчёта по взвешенным веществам), 19 граф в № 182 (без
        Кп). Строки группируются по выпускам «Выпуск № / ОКТМО выпуска», затем
        «Итого», «Итого по всем выпускам», «Всего … корректировка».

        Установленные сбросы (гр. 3–4) и выпуск в модели не хранятся —
        extra['declaration']['water_limits'] = {код: {'nds':…, 'vrs':…}} и
        ['water_outlet'] = {'name':…, 'number':…, 'oktmo':…}; без них графы
        пустые. Кп (ред. 241), Кво и Кот в расчёте равны 1 — печатаем 1, чтобы
        формула графы сошлась с суммой.
        """
        from ecodoc.core.refdata import coefficients
        band_k = coefficients()["band"]
        ed = self.edition
        o = self.ctx.organization
        ob = self.ctx.objects[0] if self.ctx.objects else None
        decl = self._decl_extra()
        lines = [ln for ln in self.calc.lines if ln.section == "Р4"]
        has_kp = ed.section4_has_kp
        ncol = 20 if has_kp else 19
        last = _col(ncol)

        ws = wb.create_sheet("Раздел 4 (сбросы)")
        xlsx.merge(ws, f"A1:{last}1", "Раздел 4. Расчёт суммы платы за сбросы "
                   "загрязняющих веществ в водные объекты", bold=True, border=False)
        permit = decl.get("water_permit") or {}
        if not isinstance(permit, dict):
            permit = {"number": str(permit)}
        head = [
            ("Категория объекта, оказывающего негативное воздействие "
             "на окружающую среду", (ob.category if ob else "")),
            ("Наименование объекта", ob.name if ob else ""),
            ("Код объекта", ob.code if ob else ""),
            ("Адрес объекта", (ob.address if ob else "") or o.address),
            ("Реквизиты документа, на основании которого осуществляются сбросы "
             "загрязняющих веществ в водные объекты: №",
             str(permit.get("number") or "")),
            ("Срок действия", str(permit.get("valid_until") or permit.get("term") or "")),
        ]
        r = 2
        for label, value in head:
            xlsx.merge(ws, f"A{r}:F{r}", label, align="left")
            xlsx.merge(ws, f"G{r}:{last}{r}", value, align="left")
            r += 1
        r += 1
        if has_kp:
            sums = ("Сумма платы за НДС, ТН, руб. (гр.6×9×10×13×14×15×16)",
                    "Сумма платы за ВРС, руб. (гр.7×9×11×13×14×15×16)",
                    "Сумма платы сверх ВРС, НДС, ТН, руб. (гр.8×9×12×13×14×15×16)",
                    "Сумма платы, всего, руб. (гр.17+18+19)")
            coef_hdr = ["Кп", "Кот", "Кво", "Кинд"]
        else:
            sums = ("Сумма платы за НДС, ТН, руб. (гр.6×9×10×13×14×15)",
                    "Сумма платы в пределах ВРС, руб. (гр.7×9×11×13×14×15)",
                    "Сумма платы сверх ВРС, НДС, ТН, руб. (гр.8×9×12×13×14×15)",
                    "Сумма платы, всего, руб. (гр.16+17+18)")
            coef_hdr = ["Кот", "Кво", "Кинд"]
        headers = [
            "N п/п", "Наименование загрязняющего вещества",
            "Установленные сбросы, т: НДС, ТН", "Установленные сбросы, т: ВРС",
            "Фактический сброс загрязняющего вещества в водные объекты, т",
            "в т.ч. в пределах НДС, ТН", "в т.ч. в пределах ВРС",
            "в т.ч. сверх ВРС, НДС, ТН", "Ставка платы (Нпл), руб./т",
            "Кнд", "Квр", "Кпр"] + coef_hdr + list(sums)
        xlsx.widths(ws, dict(zip(
            [_col(i + 1) for i in range(ncol)],
            [6, 26, 11, 11, 11, 11, 11, 11, 11, 7, 7, 8] + [7] * len(coef_hdr)
            + [13, 13, 13, 14])))
        xlsx.header_row(ws, r, headers)
        r += 1
        xlsx.header_row(ws, r, [str(i) for i in range(1, ncol + 1)])
        r += 1
        per, order = self._group_by_substance(lines)
        outlet = decl.get("water_outlet") or {}
        xlsx.merge(ws, f"A{r}:{last}{r}",
                   f"Выпуск {outlet.get('name') or '__________'} "
                   f"№ {outlet.get('number') or '____'}", align="left", bold=True)
        r += 1
        xlsx.merge(ws, f"A{r}:{last}{r}", "ОКТМО выпуска "
                   + str(outlet.get("oktmo") or (ob.oktmo if ob else "")
                         or o.oktmo or ""), align="left")
        r += 1
        limits = decl.get("water_limits") or {}
        tot = [Decimal("0")] * 4
        n = 0
        sum_cols = [_col(ncol - 3 + i) for i in range(4)]
        for key in order:
            n += 1
            d = per[key]
            lim = limits.get(key[0]) or {}
            m = {b: d["mass"].get(b, Decimal("0")) for b in ("norm", "limit", "over")}
            a = {b: d["amt"].get(b, Decimal("0")) for b in ("norm", "limit", "over")}
            total = a["norm"] + a["limit"] + a["over"]
            k10 = d["k"].get("norm", D(band_k["norm"]))
            k11 = d["k"].get("limit", D(band_k["limit"]))
            k12 = d["k"].get("over", D(band_k["over"]))
            # Кп (ред. 241) и Кво в расчёте не участвуют (=1); Кот — k_extra
            coefs = ([1.0] if has_kp else []) + [float(d["k_ot"]), 1.0,
                                                  float(d["k_ind"])]
            xlsx.data_row(ws, r, [
                n, key[1], str(lim.get("nds") or ""), str(lim.get("vrs") or ""),
                float(m["norm"] + m["limit"] + m["over"]),
                float(m["norm"]), float(m["limit"]), float(m["over"]),
                float(d["rate"]), float(k10), float(k11), float(k12)]
                + coefs + [fmt_money(a["norm"]), fmt_money(a["limit"]),
                           fmt_money(a["over"]), fmt_money(total)])
            r += 1
            for i, v in enumerate((a["norm"], a["limit"], a["over"], total)):
                tot[i] += v
        mid = _col(ncol - 4)
        for label in ("Итого", "Итого по всем выпускам"):
            xlsx.merge(ws, f"A{r}:{mid}{r}", label, align="left", bold=True)
            for col, v in zip(sum_cols, tot):
                xlsx.cell(ws, f"{col}{r}", fmt_money(v), bold=True)
            r += 1
        # корректировка размера платы программой не рассчитывается
        xlsx.merge(ws, f"A{r}:{mid}{r}", "Всего по всем выпускам по тем "
                   "загрязняющим веществам, по которым осуществляется "
                   "корректировка размера платы", align="left")
        for col in sum_cols:
            xlsx.cell(ws, f"{col}{r}", "")
        r += 1
        xlsx.merge(ws, f"A{r}:{last}{r}", "в том числе:", align="left")
        self._sign_block(ws, r + 2, last)

    # ---- отходы: общий каркас листов Разделов 5–10 ----------------------
    def _waste_rows(self, section: str):
        """Строки расчёта раздела, сгруппированные по отходу: один отход —
        одна строка бланка, массы/суммы раскладываются по корзинам."""
        per: dict = {}
        order: list = []
        for ln in self.calc.lines:
            if ln.section != section:
                continue
            key = ln.code
            if key not in per:
                per[key] = {"name": ln.name, "mass": {}, "amt": {}, "k": {},
                            "rate": ln.rate, "k_st": ln.k_extra, "k_ind": ln.k_ind}
                order.append(key)
            per[key]["mass"][ln.band] = ln.mass
            per[key]["amt"][ln.band] = ln.amount
            per[key]["k"][ln.band] = ln.k_band
        return per, order

    def _waste_by_code(self) -> dict:
        """WasteFlow по коду ФККО — для граф движения отходов (гр. 6–12)."""
        out = {}
        for w in self.ctx.wastes:
            out.setdefault(str(w.fkko_code), w)
        return out

    def _waste_sheet(self, wb, name: str, title: str, headers: list,
                     widths: list, head_extra: list, rows: list,
                     sum_cols: list, totals: list, with_classes: bool,
                     head_rows=None):
        """Общая вёрстка листа раздела по отходам: заголовок, шапка объекта,
        блок ОРО, две строки заголовков граф (названия и номера), данные,
        «ИТОГО», при необходимости «Всего по тем классам опасности…»,
        подпись исполнителя."""
        ncol = len(headers)
        last = _col(ncol)
        ws = wb.create_sheet(name)
        xlsx.merge(ws, f"A1:{last}1", title, bold=True, border=False)
        if head_rows is None:
            r = self._object_head(ws, 2, last, head_extra)
        else:
            r = 2
            for label, value in head_rows:
                xlsx.merge(ws, f"A{r}:F{r}", label, align="left")
                xlsx.merge(ws, f"G{r}:{last}{r}", value or "", align="left")
                r += 1
        r += 1
        xlsx.widths(ws, dict(zip([_col(i + 1) for i in range(ncol)], widths)))
        xlsx.header_row(ws, r, headers)
        r += 1
        xlsx.header_row(ws, r, [str(i) for i in range(1, ncol + 1)])
        r += 1
        for row in rows:
            xlsx.data_row(ws, r, row)
            r += 1
        mid = _col(sum_cols[0] - 1)
        xlsx.merge(ws, f"A{r}:{mid}{r}", "ИТОГО", align="left", bold=True)
        for col_idx, v in zip(sum_cols, totals):
            xlsx.cell(ws, f"{_col(col_idx)}{r}", fmt_money(v),
                      bold=True)
        r += 1
        if with_classes:
            # корректировка размера платы программой не рассчитывается —
            # строки бланка печатаются пустыми
            xlsx.merge(ws, f"A{r}:{mid}{r}", "Всего по тем классам опасности "
                       "отходов, по которым осуществляется корректировка размера "
                       "платы", align="left")
            for col_idx in sum_cols:
                xlsx.cell(ws, f"{_col(col_idx)}{r}", "")
            r += 1
            xlsx.merge(ws, f"A{r}:{last}{r}", "в том числе:", align="left")
            r += 1
        self._sign_block(ws, r + 1, last)
        return ws

    def _limit_of(self, code: str) -> str:
        """Графа «Установленный лимит на размещение отходов (тонн)» — в модели
        нет, берётся из extra['declaration']['waste_limits'][код ФККО]."""
        lim = (self._decl_extra().get("waste_limits") or {}).get(code)
        return "" if lim in (None, "") else str(lim)

    # Раздел 5 — отходы производства, 27 граф --------------------------------
    def _sheet_section5(self, wb):
        """Раздел 5 по 27 графам ред. 241/№ 182: движение отходов (гр. 6–12),
        размещено (13 = 14 + 15), ставка и коэффициенты (16–24), суммы
        (25 = 14×16×18×20×21×22×23×24; 26 = 15×16×19×20×22×23×24; 27 = 25+26).

        Откуда графы движения: 6 — образовалось; 7 — передано для обработки;
        8 — утилизировано (в т.ч. передано); 9 — обезврежено (в т.ч. передано);
        10 — складировано с прошлого периода >11 мес (в модели нет — пусто);
        11 — остаток на конец периода ≤11 мес (накопление); 12 — передано
        регоператору ТКО (у отходов производства пусто). Графа 17 (коэффициент
        за отходы, утилизированные в течение 11 мес) в формулы сумм не входит
        — печатается пустой; Код/Кпо(Кло)/Кот в расчёте равны 1.
        """
        from ecodoc.core.refdata import coefficients
        wband = coefficients()["waste_band"]
        ed = self.edition
        per, order = self._waste_rows("Р5")
        flows = self._waste_by_code()
        # разметка по бланку: Р5 — только отходы производства; Р6 и далее —
        # свои листы. Если платы по Р5 нет, лист всё равно печатаем (пустой).
        headers = [
            "N п/п", "Наименование вида отходов",
            "Код отходов в соответствии с ФККО",
            "Класс опасности отходов в соответствии с ФККО",
            "Установленный лимит на размещение отходов (тонн)",
            "Движение отходов, образованных в отчетном периоде (тонн): "
            "образовалось за отчетный период (в том числе после обработки)",
            "передано в отчетном периоде для обработки",
            "утилизировано в отчетном периоде, в том числе передано в целях "
            "утилизации",
            "обезврежено в отчетном периоде, в том числе передано в целях "
            "обезвреживания",
            "фактически складировано отходов предыдущего отчетного периода, "
            "не утилизированных в течение 11 месяцев",
            "фактический остаток отходов на конец отчетного периода, срок "
            "накопления которых не превышает 11 месяцев",
            "передано оператору/региональному оператору по обращению с ТКО",
            "Размещено в отчетном периоде, передано другим организациям в целях "
            "размещения (тонн) (гр.14 + гр.15)",
            "в пределах установленного лимита на размещение отходов",
            "сверх установленного лимита на размещение отходов",
            "Ставка платы (Нпл), руб./тонна",
            "Коэффициент к ставке платы за отходы, накопленные и утилизированные "
            "или переданные для утилизации в течение 11 месяцев",
            "Кл", "Ксл", "Код", ed.section5_kpo_name, "Кст", "Кот", "Кинд",
            "Сумма платы в пределах установленного лимита, руб. "
            "(гр.14×16×18×20×21×22×23×24)",
            "Сумма платы сверх установленного лимита, руб. "
            "(гр.15×16×19×20×22×23×24)",
            "Сумма платы за размещение отходов производства, руб. (гр.25+26)",
        ]
        widths = [5, 24, 13, 7, 9, 9, 9, 9, 9, 9, 9, 9, 10, 9, 9, 9, 8,
                  6, 6, 6, 6, 6, 6, 7, 13, 13, 14]
        rows, tot = [], [Decimal("0")] * 3
        for n, code in enumerate(order, start=1):
            d = per[code]
            w = flows.get(code)
            m_norm = d["mass"].get("norm", Decimal("0"))
            m_over = d["mass"].get("over", Decimal("0"))
            a_norm = d["amt"].get("norm", Decimal("0"))
            a_over = d["amt"].get("over", Decimal("0"))

            def f(attr):
                return float(D(getattr(w, attr, 0) or 0)) if w else ""

            used = (D(w.used) + D(w.transferred_util)) if w else Decimal("0")
            neut = (D(w.neutralized) + D(w.transferred_neutral)) if w else Decimal("0")
            rows.append([
                n, d["name"], code, (w.hazard_class if w else ""),
                self._limit_of(code), f("generated"), f("transferred_processing"),
                float(used) if w else "", float(neut) if w else "", "",
                f("accumulated_end_nakopl"), "",
                float(m_norm + m_over), float(m_norm), float(m_over),
                float(d["rate"]), "",
                float(d["k"].get("norm", D(wband["norm"]))),
                float(d["k"].get("over", D(wband["over"]))),
                1.0, 1.0, float(d["k_st"]), 1.0, float(d["k_ind"]),
                fmt_money(a_norm), fmt_money(a_over), fmt_money(a_norm + a_over)])
            tot[0] += a_norm
            tot[1] += a_over
            tot[2] += a_norm + a_over
        self._waste_sheet(
            wb, "Раздел 5 (отходы)", "Раздел 5. Расчёт суммы платы за "
            "размещение отходов производства", headers, widths,
            self._oro_rows(), rows, [25, 26, 27], tot, with_classes=True)

    # Раздел 6 — ТКО, 21 графа ----------------------------------------------
    def _sheet_section6(self, wb):
        """Раздел 6 по 21 графе бланка: 7 = 8 + 9 + 10 (принято / в пределах
        лимита / сверх), суммы 18 = 8×11×12×15×16×17 (принятые ТКО —
        регоператор, в модели нет), 19 = 9×11×12×14×15×16×17,
        20 = 10×11×13×15×16×17, 21 = 18+19+20. Лицензия на размещение
        отходов I–IV классов — extra['declaration']['waste_license']."""
        from ecodoc.core.refdata import coefficients
        wband = coefficients()["waste_band"]
        per, order = self._waste_rows("Р6")
        flows = self._waste_by_code()
        decl = self._decl_extra()
        headers = [
            "N п/п", "Наименование вида отходов",
            "Код отходов в соответствии с ФККО",
            "Класс опасности отходов в соответствии с ФККО",
            "Установленный лимит на размещение отходов (тонн)",
            "Образовалось отходов за отчетный период (тонн)",
            "Размещено в отчетном периоде (тонн) (гр.8 + гр.9 + гр.10)",
            "принято отходов в целях размещения в отчетном периоде (тонн)",
            "в пределах установленного лимита на размещение отходов (тонн)",
            "сверх установленного лимита на размещение отходов (тонн)",
            "Ставка платы за размещение ТКО (Нткопл), руб./тонна",
            "Кл", "Ксл", "Кпо", "Кст", "Кот", "Кинд",
            "Сумма платы: размещение принятых отходов, руб. "
            "(гр.8×11×12×15×16×17)",
            "в пределах установленного лимита, руб. (гр.9×11×12×14×15×16×17)",
            "сверх установленного лимита, руб. (гр.10×11×13×15×16×17)",
            "Сумма платы за размещение ТКО, руб. (гр.18+19+20)",
        ]
        widths = [5, 24, 13, 7, 9, 9, 10, 9, 9, 9, 9, 6, 6, 6, 6, 6, 7,
                  13, 13, 13, 14]
        rows, tot = [], [Decimal("0")] * 4
        for n, code in enumerate(order, start=1):
            d = per[code]
            w = flows.get(code)
            m_norm = d["mass"].get("norm", Decimal("0"))
            m_over = d["mass"].get("over", Decimal("0"))
            a_norm = d["amt"].get("norm", Decimal("0"))
            a_over = d["amt"].get("over", Decimal("0"))
            rows.append([
                n, d["name"], code, (w.hazard_class if w else ""),
                self._limit_of(code),
                float(D(w.generated)) if w else "",
                float(m_norm + m_over), "", float(m_norm), float(m_over),
                float(d["rate"]),
                float(d["k"].get("norm", D(wband["norm"]))),
                float(d["k"].get("over", D(wband["over"]))),
                1.0, float(d["k_st"]), 1.0, float(d["k_ind"]),
                "", fmt_money(a_norm), fmt_money(a_over),
                fmt_money(a_norm + a_over)])
            tot[1] += a_norm
            tot[2] += a_over
            tot[3] += a_norm + a_over
        extra = self._oro_rows()
        # в Разделе 6 после блока ОРО идёт лицензия на размещение I–IV классов
        extra.insert(5, ("Лицензия на деятельность по размещению отходов "
                         "I - IV классов опасности: №",
                         str(decl.get("waste_license") or "")))
        self._waste_sheet(
            wb, "Раздел 6 (ТКО)", "Раздел 6. Расчёт суммы платы за размещение "
            "твердых коммунальных отходов", headers, widths, extra, rows,
            [18, 19, 20, 21], tot, with_classes=True)

    # Раздел 7 — побочные продукты производства, 13 граф --------------------
    def _sheet_section7(self, wb):
        """Раздел 7: 5 — признано отходами; 11 = 5×6×7×9×10 (≤ 11 мес, через
        Ксл), 12 = 5×6×8×9×10 (> 11 мес, через Кпб), 13 = 11 + 12. В модели
        срок размещения не хранится: placed_norm печатается как «≤ 11 мес»,
        placed_over — как «> 11 мес»; Ксл/Кпб оставлены пустыми, так как
        расчёт применяет только коэффициент корзины."""
        per, order = self._waste_rows("Р7")
        flows = self._waste_by_code()
        oro = self._decl_extra().get("oro") or {}
        store = self._decl_extra().get("byprod_storage") or {}
        headers = [
            "N п/п", "Наименование вида отходов",
            "Код отходов в соответствии с ФККО",
            "Класс опасности отходов в соответствии с ФККО",
            "Признано отходами в отчетном периоде (тонн)",
            "Ставка платы (Нпл), руб./тонна", "Ксл", "Кпб", "Кот", "Кинд",
            "за размещение побочных продуктов производства, признанных "
            "отходами, в срок, не превышающий 11 месяцев с даты образования "
            "таких продуктов, руб. (гр.5×6×7×9×10)",
            "за размещение … в срок, превышающий 11 месяцев с даты образования "
            "таких продуктов, руб. (гр.5×6×8×9×10)",
            "Сумма платы за размещение, в том числе складирование побочных "
            "продуктов производства, признанных отходами, руб. (гр.11+12)",
        ]
        widths = [5, 26, 13, 7, 10, 10, 6, 6, 6, 7, 15, 15, 15]
        rows, tot = [], [Decimal("0")] * 3
        for n, code in enumerate(order, start=1):
            d = per[code]
            w = flows.get(code)
            m = d["mass"].get("norm", Decimal("0")) + d["mass"].get("over", Decimal("0"))
            a1 = d["amt"].get("norm", Decimal("0"))
            a2 = d["amt"].get("over", Decimal("0"))
            rows.append([n, d["name"], code, (w.hazard_class if w else ""),
                         float(m), float(d["rate"]), "", "", 1.0,
                         float(d["k_ind"]), fmt_money(a1), fmt_money(a2),
                         fmt_money(a1 + a2)])
            tot[0] += a1
            tot[1] += a2
            tot[2] += a1 + a2
        extra = [
            ("Наименование объекта размещения отходов", str(oro.get("name") or "")),
            ("Регистрационный номер объекта размещения отходов", str(oro.get("number") or "")),
            ("Адрес объекта размещения отходов", str(oro.get("address") or "")),
            ("Наименование объекта складирования побочных продуктов "
             "производства, признанных отходами", str(store.get("name") or "")),
            ("Адрес объекта складирования побочных продуктов производства, "
             "признанных отходами", str(store.get("address") or "")),
        ]
        self._waste_sheet(
            wb, "Раздел 7 (побочные продукты)", "Раздел 7. Расчёт суммы платы "
            "за размещение, в том числе складирование, побочных продуктов "
            "производства, признанных отходами", headers, widths, extra, rows,
            [11, 12, 13], tot, with_classes=False)

    # Раздел 8 — вскрышные и вмещающие породы, 19 граф ---------------------
    def _sheet_section8(self, wb):
        """Раздел 8: 6 = 7 + 8 (признано отходами: в пределах лимита / сверх),
        17 = 7×9×10×12×13×14×15×16, 18 = 8×9×11×12×14×15×16, 19 = 17 + 18."""
        from ecodoc.core.refdata import coefficients
        wband = coefficients()["waste_band"]
        per, order = self._waste_rows("Р8")
        flows = self._waste_by_code()
        headers = [
            "N п/п", "Наименование вида отходов",
            "Код отходов в соответствии с ФККО",
            "Класс опасности отходов в соответствии с ФККО",
            "Установленный лимит на размещение отходов (тонн)",
            "Признано отходами в отчетном периоде (тонн) (гр.7 + гр.8)",
            "в пределах установленного лимита на размещение отходов",
            "сверх установленного лимита на размещение отходов",
            "Ставка платы (Нпл), руб./тонна",
            "Кл", "Ксл", "Код", "Кпо", "Кст", "Кот", "Кинд",
            "Сумма платы в пределах установленного лимита, руб. "
            "(гр.7×9×10×12×13×14×15×16)",
            "Сумма платы сверх установленного лимита, руб. "
            "(гр.8×9×11×12×14×15×16)",
            "Сумма платы за размещение вскрышных и вмещающих горных пород, "
            "признанных отходами, руб. (гр.17+18)",
        ]
        widths = [5, 24, 13, 7, 9, 10, 9, 9, 9, 6, 6, 6, 6, 6, 6, 7, 13, 13, 14]
        rows, tot = [], [Decimal("0")] * 3
        for n, code in enumerate(order, start=1):
            d = per[code]
            w = flows.get(code)
            m_norm = d["mass"].get("norm", Decimal("0"))
            m_over = d["mass"].get("over", Decimal("0"))
            a_norm = d["amt"].get("norm", Decimal("0"))
            a_over = d["amt"].get("over", Decimal("0"))
            rows.append([
                n, d["name"], code, (w.hazard_class if w else ""),
                self._limit_of(code), float(m_norm + m_over), float(m_norm),
                float(m_over), float(d["rate"]),
                float(d["k"].get("norm", D(wband["norm"]))),
                float(d["k"].get("over", D(wband["over"]))),
                1.0, 1.0, float(d["k_st"]), 1.0, float(d["k_ind"]),
                fmt_money(a_norm), fmt_money(a_over), fmt_money(a_norm + a_over)])
            tot[0] += a_norm
            tot[1] += a_over
            tot[2] += a_norm + a_over
        self._waste_sheet(
            wb, "Раздел 8 (породы)", "Раздел 8. Расчёт суммы платы за размещение "
            "вскрышных и вмещающих горных пород, признанных отходами", headers,
            widths, self._oro_rows(), rows, [17, 18, 19], tot, with_classes=False)

    # Раздел 9 — побочные продукты животноводства, 10 граф -------------------
    def _sheet_section9(self, wb):
        """Раздел 9: шапка — ОКТМО и адрес места выявления нарушения (по акту
        контрольного мероприятия, ч. 6 ст. 5 ФЗ № 248-ФЗ), ОРО; 10 = 5×6×7×8×9.
        Кж (коэффициент за отходы животноводства) в расчёте не применяется —
        графа пустая."""
        per, order = self._waste_rows("Р9")
        flows = self._waste_by_code()
        decl = self._decl_extra()
        oro = decl.get("oro") or {}
        place = decl.get("violation_place") or {}
        headers = [
            "N п/п", "Наименование вида отходов",
            "Код отходов в соответствии с ФККО",
            "Класс опасности отходов в соответствии с ФККО",
            "Признано отходами в отчетном периоде (тонн)",
            "Ставка платы (Нпл), руб./тонна", "Кж", "Кот", "Кинд",
            "Сумма платы за размещение побочных продуктов животноводства, "
            "признанных отходами, руб. (гр.5×6×7×8×9)",
        ]
        widths = [5, 26, 13, 7, 10, 10, 6, 6, 7, 16]
        rows, tot = [], [Decimal("0")]
        for n, code in enumerate(order, start=1):
            d = per[code]
            w = flows.get(code)
            m = sum(d["mass"].values(), Decimal("0"))
            a = sum(d["amt"].values(), Decimal("0"))
            rows.append([n, d["name"], code, (w.hazard_class if w else ""),
                         float(m), float(d["rate"]), "", 1.0, float(d["k_ind"]),
                         fmt_money(a)])
            tot[0] += a
        head_rows = [
            ("ОКТМО места выявления нарушения", str(place.get("oktmo") or "")),
            ("Адрес места выявления нарушения", str(place.get("address") or "")),
            ("Наименование объекта размещения отходов", str(oro.get("name") or "")),
            ("Регистрационный номер объекта размещения отходов (в случае его "
             "присвоения)", str(oro.get("number") or "")),
            ("Адрес объекта размещения отходов", str(oro.get("address") or "")),
        ]
        self._waste_sheet(
            wb, "Раздел 9 (животноводство)", "Раздел 9. Расчёт суммы платы за "
            "размещение побочных продуктов животноводства, признанных отходами",
            headers, widths, [], rows, [10], tot, with_classes=False,
            head_rows=head_rows)

    # Раздел 10 — искусственные грунты (только № 182), 13 граф --------------
    def _sheet_section10(self, wb):
        """Раздел 10 формы № 182: 11 = 5×6×7×9×10 (в пределах установленных
        лимитов, через Кл), 12 = 5×6×8×9×10 (в отсутствие лимитов, через Ксл),
        13 = 11 + 12. placed_norm → графа 11, placed_over → графа 12."""
        from ecodoc.core.refdata import coefficients
        wband = coefficients()["waste_band"]
        per, order = self._waste_rows("Р10")
        flows = self._waste_by_code()
        decl = self._decl_extra()
        oro = decl.get("oro") or {}
        store = decl.get("soil_storage") or {}
        headers = [
            "N п/п", "Наименование вида отходов",
            "Код отходов в соответствии с ФККО",
            "Класс опасности отходов в соответствии с ФККО",
            "Признано отходами в отчетном периоде (тонн)",
            "Ставка платы (Нпл), руб./тонна", "Кл", "Ксл", "Кот", "Кинд",
            "за размещение искусственных грунтов, признанных отходами, в срок, "
            "не превышающий 11 месяцев (или иной срок) с даты образования таких "
            "грунтов, в пределах установленных лимитов, руб. (гр.5×6×7×9×10)",
            "… в отсутствие установленных лимитов, руб. (гр.5×6×8×9×10)",
            "Сумма платы за размещение, в том числе складирование искусственных "
            "грунтов, признанных отходами, руб. (гр.11+12)",
        ]
        widths = [5, 26, 13, 7, 10, 10, 6, 6, 6, 7, 15, 15, 15]
        rows, tot = [], [Decimal("0")] * 3
        for n, code in enumerate(order, start=1):
            d = per[code]
            w = flows.get(code)
            m = d["mass"].get("norm", Decimal("0")) + d["mass"].get("over", Decimal("0"))
            a1 = d["amt"].get("norm", Decimal("0"))
            a2 = d["amt"].get("over", Decimal("0"))
            rows.append([n, d["name"], code, (w.hazard_class if w else ""),
                         float(m), float(d["rate"]),
                         float(d["k"].get("norm", D(wband["norm"]))),
                         float(d["k"].get("over", D(wband["over"]))),
                         1.0, float(d["k_ind"]), fmt_money(a1), fmt_money(a2),
                         fmt_money(a1 + a2)])
            tot[0] += a1
            tot[1] += a2
            tot[2] += a1 + a2
        extra = [
            ("Наименование объекта размещения отходов", str(oro.get("name") or "")),
            ("Регистрационный номер объекта размещения отходов", str(oro.get("number") or "")),
            ("Адрес объекта размещения отходов", str(oro.get("address") or "")),
            ("Наименование объекта складирования искусственных грунтов, "
             "признанных отходами", str(store.get("name") or "")),
            ("Адрес объекта складирования искусственных грунтов, признанных "
             "отходами", str(store.get("address") or "")),
        ]
        self._waste_sheet(
            wb, "Раздел 10 (грунты)", "Раздел 10. Расчёт суммы платы за "
            "размещение, в том числе складирование, искусственных грунтов, "
            "признанных отходами", headers, widths, extra, rows, [11, 12, 13],
            tot, with_classes=False)

    # ----- служебный лист (ПНГ: бланк Разделов 2–3 не воспроизводится) -----
    def _sheet_lines(self, wb, title: str, medium: str,
                     sections: tuple = ()):
        c = self.calc
        rows = [ln for ln in c.lines if ln.medium == medium
                and (not sections or ln.section in sections)]
        ws = wb.create_sheet(title)
        headers = ["Раздел", "Код", "Наименование", "Норматив", "Масса, т",
                   "Ставка, руб.", "Кинд", "Кнорм", "Кдоп", "Плата, руб."]
        xlsx.header_row(ws, 1, headers,
                        widths=[8, 12, 32, 22, 12, 14, 8, 8, 8, 16])
        r = 2
        for ln in rows:
            xlsx.data_row(ws, r, [
                ln.section, ln.code, ln.name, _BAND_RU.get(ln.band, ln.band),
                float(ln.mass), float(ln.rate), float(ln.k_ind),
                float(ln.k_band), float(ln.k_extra), fmt_money(ln.amount)])
            r += 1
        # итог по листу
        total = sum((ln.amount for ln in rows), start=type(rows[0].amount)(0)) if rows else 0
        tcell = ws.cell(row=r, column=9, value="ИТОГО:")
        tcell.font = xlsx.BOLD
        v = ws.cell(row=r, column=10, value=fmt_money(total))
        v.font = xlsx.BOLD
