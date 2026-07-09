"""Тонкие хелперы поверх openpyxl для печатных форм Excel."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
LEFT = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
_ALIGNS = {
    "left": Alignment(wrap_text=True, vertical="center", horizontal="left"),
    "center": Alignment(wrap_text=True, vertical="center", horizontal="center"),
    "right": Alignment(wrap_text=True, vertical="center", horizontal="right"),
    "top-left": Alignment(wrap_text=True, vertical="top", horizontal="left"),
    "top-center": Alignment(wrap_text=True, vertical="top", horizontal="center"),
}


def new_workbook() -> Workbook:
    wb = Workbook()
    # удалить дефолтный лист, чтобы формы сами добавляли именованные
    wb.remove(wb.active)
    return wb


def header_row(ws: Worksheet, row: int, values: list[str], widths: list[int] | None = None) -> None:
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.alignment = WRAP
        c.border = BORDER
    if widths:
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w


def data_row(ws: Worksheet, row: int, values: list, aligns: list[str] | None = None) -> None:
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = BORDER
        align = (aligns[col - 1] if aligns and col - 1 < len(aligns) else "left")
        c.alignment = LEFT if align == "left" else WRAP


def save(wb: Workbook, out_path: Path) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ------------------------------------------------------------------
# Помощники для «официальных» многоячейковых форм (кадастр, журнал №1028
# и т.п.): объединение ячеек, точная простановка стилей по адресу A1.
# ------------------------------------------------------------------

def cell(ws: Worksheet, coord: str, value=None, *, bold: bool = False,
         italic: bool = False, size: int | None = None, align: str = "center",
         border: bool = True, fill: bool = False, wrap: bool = True) -> None:
    """Записать значение в ячейку по адресу A1 с оформлением."""
    c = ws[coord]
    if value is not None:
        c.value = value
    c.font = Font(bold=bold, italic=italic, size=size or 11)
    a = _ALIGNS.get(align, CENTER)
    if not wrap:
        a = Alignment(wrap_text=False, vertical=a.vertical, horizontal=a.horizontal)
    c.alignment = a
    if border:
        c.border = BORDER
    if fill:
        c.fill = HEADER_FILL


def merge(ws: Worksheet, rng: str, value=None, *, bold: bool = False,
          italic: bool = False, size: int | None = None, align: str = "center",
          border: bool = True, fill: bool = False, wrap: bool = True) -> None:
    """Объединить диапазон A1:B2 и оформить верхнюю-левую ячейку.

    Границы ставятся на все ячейки диапазона (иначе рамка объединения рвётся).
    """
    ws.merge_cells(rng)
    top_left = rng.split(":")[0]
    cell(ws, top_left, value, bold=bold, italic=italic, size=size,
         align=align, border=border, fill=fill, wrap=wrap)
    if border:
        for row in ws[rng]:
            for c in row:
                c.border = BORDER


def widths(ws: Worksheet, mapping: dict[str, float]) -> None:
    """Задать ширины столбцов: {'A': 6, 'B': 40, ...}."""
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


def heights(ws: Worksheet, mapping: dict[int, float]) -> None:
    """Задать высоты строк: {1: 30, 5: 60, ...}."""
    for r, h in mapping.items():
        ws.row_dimensions[r].height = h
