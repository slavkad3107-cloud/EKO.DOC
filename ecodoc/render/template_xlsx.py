"""Заполнение готового бланка Excel с сохранением формул.

Требование ТЗ по журналу 1028: «Учти, в имеющейся форме есть формулы в экселе
для формирования формы и данных в ней».

`openpyxl.load_workbook(..., data_only=False)` читает формулы как текст `=…`
и при сохранении оставляет их на месте — значит можно вписать исходные данные
в свои ячейки, а расчётные колонки посчитает сам Excel при открытии.

Принцип работы (гибрид):
  * есть бланк-образец у пользователя → заполняем его;
  * образца нет → форма рисуется кодом, как раньше.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path


@dataclass
class Cell:
    """Куда писать: лист + адрес ячейки."""
    sheet: str
    ref: str


def open_template(path: str | Path):
    """Открыть бланк, СОХРАНИВ формулы (data_only=False — важно)."""
    import openpyxl
    return openpyxl.load_workbook(Path(path), data_only=False)


def find_anchor(ws, text: str, limit_rows: int = 200,
                limit_cols: int = 40) -> tuple[int, int] | None:
    """Найти ячейку по началу текста (заголовок графы, подпись поля).

    Бланки у всех разные, поэтому адреса не зашиваем: ищем по подписи."""
    needle = re.sub(r"\s+", " ", str(text or "")).strip().lower()
    if not needle:
        return None
    for row in ws.iter_rows(min_row=1, max_row=limit_rows, max_col=limit_cols):
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            if re.sub(r"\s+", " ", v).strip().lower().startswith(needle):
                return cell.row, cell.column
    return None


def is_formula(ws, row: int, col: int) -> bool:
    v = ws.cell(row=row, column=col).value
    return isinstance(v, str) and v.startswith("=")


def write_row(ws, row: int, start_col: int, values: list,
              skip_formulas: bool = True) -> int:
    """Записать значения в строку, НЕ затирая ячейки с формулами.

    Возвращает, сколько значений записано."""
    written = 0
    for i, value in enumerate(values):
        col = start_col + i
        if value in (None, ""):
            continue
        if skip_formulas and is_formula(ws, row, col):
            continue        # расчётную колонку не трогаем — Excel посчитает сам
        ws.cell(row=row, column=col, value=value)
        written += 1
    return written


def fill_by_labels(ws, pairs: dict, offset: int = 1) -> list[str]:
    """Заполнить поля «подпись → значение»: значение кладём правее подписи.

    Возвращает список подписей, которые не нашлись в бланке."""
    missing = []
    for label, value in pairs.items():
        pos = find_anchor(ws, label)
        if not pos:
            missing.append(label)
            continue
        row, col = pos
        target = col + offset
        if is_formula(ws, row, target):
            continue
        ws.cell(row=row, column=target, value=value)
    return missing


def save_as(wb, out_path: str | Path) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def fill_template(code: str, out_path: str | Path, *, sheet_fills: dict) -> Path | None:
    """Заполнить бланк-образец документа, если он есть у пользователя.

    sheet_fills: {имя листа или "*": {"подпись поля": значение}}.
    Возвращает путь к заполненному файлу либо None, если образца нет
    (тогда вызывающий код рисует форму сам)."""
    from ecodoc.core import forms_registry
    sample = forms_registry.sample_for(code)
    if sample is None or sample.suffix.lower() not in (".xlsx", ".xlsm"):
        return None
    wb = open_template(sample)
    for sheet_name, pairs in (sheet_fills or {}).items():
        sheets = wb.worksheets if sheet_name == "*" else \
            [wb[sheet_name]] if sheet_name in wb.sheetnames else []
        for ws in sheets:
            fill_by_labels(ws, pairs)
    return save_as(wb, out_path)
