"""Выгрузка исходных данных по источникам для загрузки в УПРЗА.

Две формы:
  • Excel-таблица источников и веществ — под ручной/полуавтоматический ввод
    в «Эколог», «Эколог-ПРО», «Призму», «ЭРА» (структуру колонок легко
    сопоставить с мастером импорта конкретной УПРЗА);
  • JSON — машиночитаемый формат для собственного реимпорта и обмена.

Обратная операция (загрузка данных ИЗ выгрузок «Эколога») уже есть в
development/volume_builder.ingest_excel.
"""
from __future__ import annotations

import json
import math
from pathlib import Path

from ecodoc.development.dispersion_map import MapSource

# колонки таблицы источников (порядок = как в мастере импорта «Эколога»)
_SRC_COLS = [
    "№ источника", "Наименование источника", "Тип (1-точечный)",
    "X1, м", "Y1, м", "X2, м", "Y2, м",
    "Высота H, м", "Диаметр устья D, м",
    "Скорость ГВС w0, м/с", "Расход ГВС V1, м³/с",
    "Температура ГВС Tг, °C", "Температура воздуха Tв, °C",
    "Коэф. рельефа η",
]
_SUB_COLS = [
    "№ источника", "Код вещества", "Наименование вещества",
    "Выброс, г/с", "Выброс, т/год", "F (оседание)", "ПДКм.р., мг/м³",
    "Фон, мг/м³",
]


def _v1(s: MapSource) -> float:
    return math.pi * s.D ** 2 / 4 * s.w0


def to_excel(sources: list[MapSource], out_path: str | Path,
             hours_per_year: float = 8760) -> Path:
    """Excel-книга: листы «Источники», «Выбросы», «Инструкция»."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Источники"
    ws.append(_SRC_COLS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for i, s in enumerate(sources, 1):
        ws.append([i, s.name, 1, round(s.x, 2), round(s.y, 2), "", "",
                   s.H, s.D, s.w0, round(_v1(s), 4), s.Tg, s.Tv, s.eta])

    ws2 = wb.create_sheet("Выбросы")
    ws2.append(_SUB_COLS)
    for c in ws2[1]:
        c.font = Font(bold=True)
    for i, s in enumerate(sources, 1):
        ty = s.M * 3600 * hours_per_year / 1e6      # г/с → т/год (грубо)
        ws2.append([i, s.code, s.substance or s.name, s.M, round(ty, 4),
                    s.F, s.pdk or "", s.bg or ""])

    ws3 = wb.create_sheet("Инструкция")
    for line in (
        "ВЫГРУЗКА ИСХОДНЫХ ДАННЫХ ДЛЯ УПРЗА (ЭКО.DOC)",
        "",
        "Лист «Источники» — параметры источников выбросов (ИЗА).",
        "Лист «Выбросы» — выбросы веществ по каждому источнику.",
        "",
        "Как загрузить в «Эколог»:",
        "  1. Проект → Импорт → из Excel (или мастер ввода источников).",
        "  2. Сопоставьте колонки этого файла с полями мастера "
        "(номер, координаты, H, D, w0/V1, температуры, вещество, г/с).",
        "  3. Проверьте единицы: H,D — м; w0 — м/с; V1 — м³/с; выброс — г/с.",
        "",
        "Т/год рассчитан как г/с × 3600 × часов_работы (по умолчанию 8760).",
        "Скорректируйте под фактический режим работы источника.",
        "",
        "Координаты X,Y — в местной системе площадки (м). Если не заданы —",
        "проставьте по генплану перед расчётом в УПРЗА.",
    ):
        ws3.append([line])
    ws3.column_dimensions["A"].width = 80

    for ws_ in (ws, ws2):
        for col in ws_.columns:
            width = max((len(str(c.value)) for c in col if c.value), default=10)
            ws_.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

    wb.save(out_path)
    return out_path


def to_json(sources: list[MapSource], out_path: str | Path) -> Path:
    """Машиночитаемая выгрузка источников (для реимпорта и обмена)."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    data = {"format": "ecodoc-sources/1", "sources": [
        {"name": s.name, "x": s.x, "y": s.y, "H": s.H, "D": s.D,
         "w0": s.w0, "V1": round(_v1(s), 5), "Tg": s.Tg, "Tv": s.Tv,
         "M_gs": s.M, "F": s.F, "eta": s.eta, "A": s.A,
         "substance": s.substance, "code": s.code, "pdk": s.pdk, "bg": s.bg}
        for s in sources]}
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                        encoding="utf-8")
    return out_path
