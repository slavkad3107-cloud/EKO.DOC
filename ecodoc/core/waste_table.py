"""«Табличка по отходам» — по бланку пользователя «форма таблички по отходам.xls».

Лист «расчет» бланка: наименование | код ФККО | т | м³ | плотность | примечание,
позиции сгруппированы по классам опасности со строками «ИТОГО N класса
опасности» и общим итогом. Такую табличку пользователь прикладывает к ТУ и КП
по строительным отходам.

Бланк — старый .xls (openpyxl его не пишет), поэтому файл собирается заново
в той же раскладке, а не правится поверх. Масса берётся из перечня отходов
объекта, объём считается из плотности: м³ = т / ρ. Плотность ищем в
справках-актах (там она есть у строительных отходов); если её нет — графа
остаётся пустой, ничего не выдумываем.
"""
from __future__ import annotations

from pathlib import Path

from ecodoc.core.models import ReportContext

TITLE = "Табличка по отходам (т / м³, по классам опасности)"

HEADS = ["Наименование отхода", "Код ФККО", "т", "м³",
         "Плотность, т/м³", "Примечание", "№"]


def _fkko_spaced(code: str) -> str:
    """11 цифр → «7 33 100 01 72 4», как пишут в документах."""
    d = "".join(ch for ch in str(code or "") if ch.isdigit())
    if len(d) != 11:
        return str(code or "")
    return f"{d[0]} {d[1:3]} {d[3:6]} {d[6:8]} {d[8:10]} {d[10]}"


def densities(ctx: ReportContext) -> dict[str, float]:
    """Плотность по коду ФККО из справок-актов.

    Берём явную графу «Плотн., т/м³», а где её нет — выводим из пары
    масса/объём того же акта (в справках по строительным отходам обычно
    заполняют именно т и м³): так же поступает и «Сводная по отходам»,
    иначе два документа из одних актов расходились бы."""
    from ecodoc.core.waste_agg import norm_fkko
    explicit: dict[str, float] = {}
    mass_sum: dict[str, float] = {}
    vol_sum: dict[str, float] = {}
    for a in ctx.waste_acts:
        code = norm_fkko(a.fkko_code)
        if not code:
            continue

        def num(v) -> float:
            try:
                return float(v or 0)
            except (TypeError, ValueError):
                return 0.0
        d = num(a.density)
        if d > 0 and code not in explicit:
            explicit[code] = d
        m, v = num(a.mass), num(getattr(a, "volume_m3", 0))
        if m > 0 and v > 0:
            mass_sum[code] = mass_sum.get(code, 0.0) + m
            vol_sum[code] = vol_sum.get(code, 0.0) + v
    out = dict(explicit)
    for code, v in vol_sum.items():
        if code not in out and v > 0:
            out[code] = mass_sum[code] / v
    return out


def rows(ctx: ReportContext) -> list[dict]:
    """Позиции таблички: перечень отходов + т/м³/плотность."""
    from ecodoc.development.waste_inventory import collect
    dens = densities(ctx)
    out = []
    for r in collect(ctx):
        mass = float(r.get("generated") or 0)
        rho = dens.get(r["fkko"], 0.0)
        note = ""
        if r.get("passport"):
            note = "паспорт есть"
        elif 0 < int(r.get("hazard") or 0) <= 4:
            note = "делаем паспорт"
        out.append({"name": r["name"], "fkko": r["fkko"],
                    "hazard": int(r.get("hazard") or 0), "mass": mass,
                    "volume": (mass / rho) if (mass and rho) else 0.0,
                    "density": rho, "note": note})
    # в бланке позиции идут по классам опасности
    out.sort(key=lambda r: (r["hazard"] or 9, r["fkko"]))
    return out


def build_xlsx(ctx: ReportContext, out_path: str | Path,
               source_label: str = "") -> Path:
    """Собрать табличку в раскладке бланка пользователя."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, Side

    data = rows(ctx)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "расчет"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    obj = ctx.objects[0] if ctx.objects else None
    ws["A1"] = (ctx.organization.short_name or ctx.organization.name or "")
    ws["A1"].font = bold
    if obj and (obj.name or obj.address):
        ws["A2"] = f"Объект: {obj.name or ''} {obj.address or ''}".strip()
    ws["A3"] = source_label or "По данным учёта отходов"

    head_row = 4
    for col, text in enumerate(HEADS, start=1):
        c = ws.cell(row=head_row, column=col, value=text)
        c.font = bold
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)

    widths = [46, 18, 10, 10, 12, 26, 5]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    r = head_row + 1
    n = 0
    cls_mass = cls_vol = 0.0
    total_mass = total_vol = 0.0
    prev_cls: int | None = None

    def subtotal(cls: int):
        nonlocal r, cls_mass, cls_vol
        ws.cell(row=r, column=1, value=f"ИТОГО {cls} класс опасности").font = bold
        ws.cell(row=r, column=3, value=round(cls_mass, 4)).font = bold
        if cls_vol:
            ws.cell(row=r, column=4, value=round(cls_vol, 3)).font = bold
        for col in range(1, len(HEADS) + 1):
            ws.cell(row=r, column=col).border = border
        r += 1
        cls_mass = cls_vol = 0.0

    for item in data:
        if prev_cls is not None and item["hazard"] != prev_cls:
            subtotal(prev_cls)
        prev_cls = item["hazard"]
        n += 1
        vals = [item["name"], _fkko_spaced(item["fkko"]),
                round(item["mass"], 4) if item["mass"] else None,
                round(item["volume"], 3) if item["volume"] else None,
                item["density"] or None, item["note"], n]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = border
            if col in (1, 6):
                c.alignment = wrap
        cls_mass += item["mass"]
        cls_vol += item["volume"]
        total_mass += item["mass"]
        total_vol += item["volume"]
        r += 1
    if prev_cls is not None:
        subtotal(prev_cls)

    ws.cell(row=r, column=1, value="ВСЕГО").font = bold
    ws.cell(row=r, column=3, value=round(total_mass, 4)).font = bold
    if total_vol:
        ws.cell(row=r, column=4, value=round(total_vol, 3)).font = bold
    for col in range(1, len(HEADS) + 1):
        ws.cell(row=r, column=col).border = border

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
