"""Тесты формы 2-ТП (водхоз) и коэрции типов из формы.

Бланк — приложение к приказу Росстата от 02.10.2024 № 445 (в ред. приказа
от 06.08.2026 № 473), сверено 22.08.2026 по normativ.kontur.ru
(documentId=507797): Раздел 1 — 49 граф, строки 11, 12, …; гр.32-41 — пять
пар «код/объём» использования, гр.42-47 — три пары передачи без
использования, гр.48-49 — передано после использования; Раздел 2 — строки
21, 22, …, 30 граф + пары ЗВ с гр.31; итоговых строк в бланке нет.
"""
import tempfile
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook

from ecodoc.core import registry, serialize
from ecodoc.core.models import Organization, ReportContext, ReportPeriod

# помесячная разбивка из реально принятого отчёта: сумма = 1.29 (гр.12)
_MONTHS = [0.0, 0.0, 0.01, 0.0, 0.01, 0.02, 0.02, 0.0, 0.0, 0.0, 0.31, 0.92]


def _ctx():
    ctx = ReportContext(
        organization=Organization(name="Т", inn="7801234564", okpo="12345678"),
        period=ReportPeriod(year=2024))
    ctx.extra["water"] = {
        "intake": [{"name": "Скв.1", "type": "подземный", "volume": "12.5"}],
        "discharge": [{"receiver": "р. Нева", "quality": "СД",
                       "volume": "8.0"}],
        "recycled": "40"}
    return ctx


def _ctx_full():
    """Полные данные по структуре бланка № 445 (по образцу принятого отчёта)."""
    ctx = ReportContext(
        organization=Organization(
            name="ООО «Протелюкс»", inn="4707038091", okpo="25853809",
            okved="10.89.4", address="Ивангород, ул. Лесная, зд. 13А/1",
            director_name="Брезгин М.Ю.", phone="89214494705",
            email="test@mail.ru"),
        period=ReportPeriod(year=2024))
    ctx.extra["water"] = {
        "okato": "41221505", "guiv": "412285",
        "intake": [{
            "name": "Водозабор р. Нарва", "type": "поверхностный",
            "doc_type": "Л", "doc_no": "480082", "doc_date": "25.02.2020",
            "source_code": "60", "water_body_code": "БАЛ/НАРВА",
            "distance_km": 16.9, "quality": "ТН",
            "okato": "41221505", "vhu": "01.03.00.004",
            "limit": 145.71, "volume": 1.29, "months": list(_MONTHS),
            "measured": 1.29, "losses": 0.0,
            "used_total": 1.29,
            "uses": [{"code": "102", "volume": 1.29}],
            "transfers": [{"code": "ПК", "volume": 0.1}],
            "transfer_after_use": {"code": "СК", "volume": 0.2},
        }],
        "discharge": [{
            "receiver": "р. Нарва", "quality": "СД",
            "doc_type": "Р", "doc_no": "123", "doc_date": "01.01.2023",
            "receiver_code": "20", "water_body_code": "БАЛ/НАРВА",
            "distance_km": 16.9, "okato": "41221505", "vhu": "01.03.00.004",
            "limit": 145.71, "volume": 1.29, "measured": 1.29,
            "normatively_treated": 1.29, "treatment_code": "5",
            "treatment_capacity": 2.5, "months": list(_MONTHS),
            "pollutants": [{"code": "132", "name": "БПКполн",
                            "mass": 0.0504}],
        }],
        "recycled": "40"}
    return ctx


def _report(ctx):
    registry.load_all()
    return registry.get("2tp-water")(ctx)


def test_tp2_water_implemented():
    registry.load_all()
    cls = registry.get("2tp-water")
    assert getattr(cls, "implemented", True)


def test_tp2_water_generates(tmp_path):
    rep = _report(_ctx())
    assert not [i for i in rep.validate() if i.level == "error"]
    xml = rep.render_xml(tmp_path / "w.xml")
    xlsx = rep.render_print(tmp_path / "w.xlsx")
    assert xml.exists() and xlsx.exists()
    assert "Забор" in xml.read_text(encoding="utf-8") or \
           "ЗаборВоды" in xml.read_text(encoding="utf-8")


def test_title_page_requisites(tmp_path):
    """Расхождение 1: адресная часть бланка — организация, адрес, коды
    гр.1-6 (ОКУД/ОКПО/ИНН/ОКВЭД2/ОКАТО/ГУИВ) и подпись должностного лица."""
    rep = _report(_ctx_full())
    p = rep.render_print(tmp_path / "w.xlsx")
    wb = load_workbook(p)
    assert wb.sheetnames[0] == "Титульный лист"
    ws = wb["Титульный лист"]
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row
                    if c.value not in (None, ""))
    assert "СВЕДЕНИЯ ОБ ИСПОЛЬЗОВАНИИ ВОДЫ" in text
    assert "за 2024 г." in text
    assert "Протелюкс" in text and "Ивангород" in text
    # кодовая часть: ОКУД, ОКПО, ИНН, ОКВЭД2, ОКАТО, ГУИВ (не ОКТМО/ОГРН)
    row17 = [ws.cell(row=17, column=c).value for c in range(1, 7)]
    assert row17 == ["0609060", "25853809", "4707038091", "10.89.4",
                     "41221505", "412285"]
    assert "ГУИВ" in text and "ОКАТО" in text and "ОКТМО" not in text
    # подпись: должность, ФИО, телефон, e-mail
    assert "Брезгин М.Ю." in text and "89214494705" in text
    assert "Должностное лицо" in text
    # на бланке — реквизиты и приказа об утверждении, и приказа об изменениях
    assert "02.10.2024 № 445" in text and "06.08.2026" in text
    assert "№ 473" in text


def test_section1_49_graphs(tmp_path):
    """Раздел 1 по бланку № 445 — 49 пронумерованных граф, включая
    документ, коды, лимит (гр.11), помесячную разбивку (гр.13-24), ПЯТЬ пар
    использования (гр.32-41), три пары передачи без использования (гр.42-47)
    и одну пару после использования (гр.48-49); графа «А» — N строки с 11."""
    rep = _report(_ctx_full())
    wb = load_workbook(rep.render_print(tmp_path / "w.xlsx"))
    ws = wb["Раздел 1"]
    # строка 4 — номера граф 1..49 (графа N в колонке N+1, A — «N строки»)
    nums = [ws.cell(row=4, column=2 + i).value for i in range(49)]
    assert nums == list(range(1, 50))
    assert ws.cell(row=4, column=1).value == "А"
    assert ws.cell(row=3, column=1).value == "N строки"
    row = 5  # первая строка данных
    assert ws.cell(row=row, column=1).value == 11        # N строки по бланку
    assert ws.cell(row=row, column=2).value == "Л"          # гр.1 тип документа
    assert ws.cell(row=row, column=3).value == "480082"     # гр.2 номер
    assert ws.cell(row=row, column=5).value == "60"         # гр.4 код типа ист.
    assert ws.cell(row=row, column=6).value == "БАЛ/НАРВА"  # гр.5 водный объект
    assert ws.cell(row=row, column=7).value == 16.9         # гр.6 расстояние
    assert ws.cell(row=row, column=12).value == 145.71      # гр.11 лимит
    assert ws.cell(row=row, column=13).value == 1.29        # гр.12 всего за год
    # гр.13-24 (колонки 14-25) — помесячно; ноябрь = гр.23 = колонка 24
    assert ws.cell(row=row, column=24).value == 0.31
    assert ws.cell(row=row, column=25).value == 0.92        # декабрь
    assert ws.cell(row=row, column=26).value == 1.29        # гр.25 по приборам
    assert ws.cell(row=row, column=33).value == "102"       # гр.32 код вида исп.
    assert ws.cell(row=row, column=34).value == 1.29        # гр.33 объём
    # гр.42-47 — передано БЕЗ использования (п. 2.16), первая пара = гр.42-43
    assert ws.cell(row=row, column=43).value == "ПК"
    assert ws.cell(row=row, column=44).value == 0.1
    # гр.48-49 — передано ПОСЛЕ использования (п. 2.17)
    assert ws.cell(row=row, column=49).value == "СК"
    assert ws.cell(row=row, column=50).value == 0.2
    # шапка: пары 32-41 — виды использования, 42-47 — без использования
    assert "использования" in ws.cell(row=3, column=42).value   # гр.41
    assert "без использования" in ws.cell(row=3, column=43).value  # гр.42
    assert "после использования" in ws.cell(row=3, column=49).value  # гр.48
    # итоговой строки в бланке нет
    assert ws.cell(row=row + 1, column=1).value in (None, "")


def test_section2_graphs_and_zv(tmp_path):
    """Раздел 2 по бланку № 445 — 30 граф (категории очистки гр.13-18,
    помесячно гр.19-30) + пары «код ЗВ / масса» с гр.31; N строки с 21;
    масса ЗВ округляется до трёх знаков (примечание к бланку)."""
    rep = _report(_ctx_full())
    wb = load_workbook(rep.render_print(tmp_path / "w.xlsx"))
    ws = wb["Раздел 2"]
    nums = [ws.cell(row=4, column=2 + i).value for i in range(32)]
    assert nums == list(range(1, 33))  # 30 граф + 1 пара ЗВ (гр.31-32)
    row = 5
    assert ws.cell(row=row, column=1).value == 21           # N строки
    assert ws.cell(row=row, column=2).value == "Р"          # гр.1
    assert ws.cell(row=row, column=5).value == "20"         # гр.4 тип приёмника
    assert ws.cell(row=row, column=8).value == "СД"         # гр.7 код Прил. 2
    assert ws.cell(row=row, column=11).value == 145.71      # гр.10 лимит
    assert ws.cell(row=row, column=12).value == 1.29        # гр.11 всего за год
    assert ws.cell(row=row, column=17).value == "5"         # гр.16 код Прил. 4
    assert ws.cell(row=row, column=18).value == 1.29        # гр.17 норм.-очищ.
    assert ws.cell(row=row, column=19).value == 2.5         # гр.18 мощность
    assert ws.cell(row=row, column=31).value == 0.92        # гр.30 декабрь
    assert ws.cell(row=row, column=32).value == "132"       # гр.31 код ЗВ
    assert ws.cell(row=row, column=33).value == 0.05        # гр.32: 0.0504→0.050
    text = " ".join(str(c.value) for r_ in ws.iter_rows() for c in r_
                    if c.value not in (None, ""))
    # сноска <1> бланка — дословно, включая взвешенные вещества и азот общий
    assert "взвешенные вещества (113)" in text and "азот общий (2)" in text
    assert "округляется до трёх знаков" in text


def test_xml_months_and_new_fields(tmp_path):
    """Расхождение 4: помесячная разбивка в XML — <Месяцы><М н="1">…</М>
    внутри <Источник> и <Выпуск>; лимит и коды не отбрасываются молча."""
    rep = _report(_ctx_full())
    p = rep.render_xml(tmp_path / "w.xml")
    root = etree.parse(str(p)).getroot()
    src = root.find(".//Источник")
    ms = src.find("Месяцы")
    assert ms is not None and len(ms) == 12
    assert ms[0].tag == "М" and ms[0].get("н") == "1"
    assert ms[10].text == "0.31"                # ноябрь
    assert src.findtext("ДопустимыйОбъём") == "145.71"
    assert src.findtext("КодВодногоОбъекта") == "БАЛ/НАРВА"
    assert src.find("Документ").get("номер") == "480082"
    out = root.find(".//Выпуск")
    assert out.find("Месяцы") is not None and len(out.find("Месяцы")) == 12
    assert out.findtext("ДопустимыйОбъём") == "145.71"
    cat = out.find("ОтведеноПоКатегориям")
    assert cat is not None
    assert cat.find("НормативноОчищенные").get("код") == "5"
    assert cat.findtext("МощностьОчистных") == "2.5"
    assert src.find("ПереданоБезИспользования").get("код") == "ПК"
    assert src.find("ПереданоПослеИспользования").get("код") == "СК"
    assert out.find("ЗВ").get("ед") == "т"        # БПК полн (132) — в тоннах
    assert "№ 473" in root.get("НПА")


def test_minimal_old_data_prints_empty_graphs(tmp_path):
    """Страховка обратной совместимости: старые данные (только
    name/type/volume) не ломают 49-графный лист — объём попадает в гр.12,
    остальные графы ПУСТЫЕ (ничего не выдумываем), итоги — только на
    служебном листе «Сводка» (в бланке итоговой строки нет),
    а в XML необязательные элементы не появляются."""
    rep = _report(_ctx())
    wb = load_workbook(rep.render_print(tmp_path / "w.xlsx"))
    ws = wb["Раздел 1"]
    assert ws.cell(row=5, column=13).value == 12.5   # гр.12 — всего за год
    assert ws.cell(row=5, column=2).value in (None, "")   # гр.1 — пусто
    assert ws.cell(row=5, column=12).value in (None, "")  # гр.11 лимит — пусто
    assert ws.cell(row=6, column=1).value in (None, "")   # нет строки ИТОГО
    assert wb["Сводка"].cell(row=2, column=2).value == 12.5
    root = etree.parse(str(rep.render_xml(tmp_path / "w.xml"))).getroot()
    src = root.find(".//Источник")
    assert src.findtext("Объём") == "12.5"
    assert src.find("Месяцы") is None            # нет данных — нет элемента
    assert src.find("ДопустимыйОбъём") is None
    assert src.find("Документ") is None


def test_months_sum_mismatch_warns():
    """гр.12 обязана равняться сумме гр.13-24 — иначе предупреждение."""
    ctx = _ctx_full()
    ctx.extra["water"]["intake"][0]["months"][0] = 5.0  # ломаем сумму
    issues = _report(ctx).validate()
    assert any(i.field == "месяцы" and "не сходится" in i.message
               for i in issues)
    # а на корректных данных предупреждения нет
    ok = _report(_ctx_full()).validate()
    assert not [i for i in ok if i.field == "месяцы"]


def test_quality_code_not_warned():
    """Коды Прил. 2 («СД», «ТН») — не повод для предупреждения; всё прочее
    (старый «НЧ», текст) — повод: бланк принимает только коды."""
    ctx = _ctx_full()
    issues = _report(ctx).validate()
    assert not [i for i in issues if i.field == "качество"]
    ctx.extra["water"]["discharge"][0]["quality"] = "НЧ"
    issues = _report(ctx).validate()
    assert [i for i in issues if i.field == "качество"]


def test_blank_codes_validated():
    """Код очистного сооружения — только 5/6/7 (Прил. 4); сумма гр.14+17 =
    гр.11 (п. 3.10); сумма видов использования = гр.31 (п. 2.15);
    к Разделу 2 прилагается расчёт НДС (п. 3.1)."""
    ctx = _ctx_full()
    issues = _report(ctx).validate()
    assert not [i for i in issues if i.field == "очистка"]
    assert not [i for i in issues if i.field == "использование"]
    assert any(i.field == "приложения" and "допустимого сброса" in i.message
               for i in issues)
    d = ctx.extra["water"]["discharge"][0]
    d["treatment_code"] = "40"                 # код старой формы
    d["normatively_treated"] = 1.0             # 1.0 ≠ гр.11 = 1.29
    ctx.extra["water"]["intake"][0]["uses"][0]["volume"] = 0.5
    issues = _report(ctx).validate()
    msgs = [i.message for i in issues if i.field == "очистка"]
    assert any("Прил. 4" in m for m in msgs) and any("3.10" in m for m in msgs)
    assert any(i.field == "использование" and "2.15" in i.message
               for i in issues)


def test_pairs_limits_warn_and_do_not_overflow(tmp_path):
    """Шестой вид использования и 25-е вещество в бланк не помещаются —
    предупреждение, а печать не съезжает на чужие графы."""
    ctx = _ctx_full()
    src = ctx.extra["water"]["intake"][0]
    src["uses"] = [{"code": str(100 + i), "volume": 0.1} for i in range(6)]
    src["used_total"] = 0.6
    issues = _report(ctx).validate()
    assert any("5 пар" in i.message for i in issues)
    wb = load_workbook(_report(ctx).render_print(tmp_path / "w.xlsx"))
    ws = wb["Раздел 1"]
    assert ws.cell(row=5, column=42).value == 0.1     # гр.41 — пятая пара
    assert ws.cell(row=5, column=43).value == "ПК"    # гр.42 — уже передача


def test_period_year_coerced_from_string():
    """Год/квартал/класс из формы приходят строками — сериализация чинит тип."""
    ctx = _ctx()
    ctx.period.year = 2024
    serialize.to_json(ctx, Path(tempfile.mkdtemp()) / "c.json")
    # смоделируем строковый год в JSON (как из формы)
    import json
    p = Path(tempfile.mkdtemp()) / "c.json"
    data = {"organization": {"inn": "7801234564"},
            "period": {"year": "2024", "quarter": "2"},
            "wastes": [{"fkko_code": "47110101521", "hazard_class": "1",
                        "generated": "0.05"}]}
    p.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    loaded = serialize.from_json(p)
    assert loaded.period.year == 2024 and isinstance(loaded.period.year, int)
    assert loaded.period.quarter == 2
    assert loaded.wastes[0].hazard_class == 1


def test_blank_layout_sheets_like_scan(tmp_path):
    """Сверка со сканом сданного отчёта ПРОТЕЛЮКС-2023 («Формы/Отчетность/
    2-ТП водхоз»): кроме машинных листов печатаются листы «(бланк)» в
    разбивке на блоки граф 1-6 / 7-19 / 20-31 / 32-49 (Раздел 1) и
    1-6 / 7-18 / 19-30 / пары ЗВ по 8 (Раздел 2) с подписями «Код по ОКЕИ»,
    предпечатанными строками 11-15 / 21-25 и «Бланк № 1 / Всего бланков 1»."""
    rep = _report(_ctx_full())
    wb = load_workbook(rep.render_print(tmp_path / "w.xlsx"))
    assert "Раздел 1 (бланк)" in wb.sheetnames and "Раздел 2 (бланк)" in wb.sheetnames
    ws = wb["Раздел 1 (бланк)"]
    assert ws["A1"].value == "Т1"
    assert ws["A3"].value == "Код по ОКЕИ: километр - 008"
    assert [ws.cell(row=5, column=c).value for c in range(2, 8)] == [1, 2, 3, 4, 5, 6]
    assert [ws[f"A{r}"].value for r in range(6, 11)] == [11, 12, 13, 14, 15]   # предпечатанные
    assert ws["B6"].value == "Л" and ws["F6"].value == "БАЛ/НАРВА"
    assert ws["A12"].value == "Код по ОКЕИ: тысяча кубических метров - 114"
    assert [ws.cell(row=14, column=c).value for c in range(2, 15)] == list(range(7, 20))
    assert ws["G15"].value == 1.29 and ws["N15"].value == 0.02          # гр.12, гр.19 июль
    assert [ws.cell(row=23, column=c).value for c in range(2, 14)] == list(range(20, 32))
    assert ws["E24"].value == 0.31 and ws["F24"].value == 0.92         # ноябрь/декабрь
    assert [ws.cell(row=32, column=c).value for c in range(2, 20)] == list(range(32, 50))
    assert ws["B33"].value == "102" and ws["C33"].value == 1.29        # гр.32-33 как в скане
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "Бланк №  1" in text and "Всего бланков  1" in text
    ws2 = wb["Раздел 2 (бланк)"]
    assert ws2["A1"].value == "Т2"
    assert [ws2[f"A{r}"].value for r in range(6, 11)] == [21, 22, 23, 24, 25]
    assert [ws2.cell(row=14, column=c).value for c in range(2, 14)] == list(range(7, 19))
    assert [ws2.cell(row=23, column=c).value for c in range(2, 14)] == list(range(19, 31))
    assert ws2["B32"].value == 31 and ws2["B33"].value == "132" and ws2["C33"].value == 0.05
    text2 = " ".join(str(c.value) for row in ws2.iter_rows() for c in row if c.value)
    assert "приводятся в тоннах" in text2 and "округляется до трёх знаков" in text2
    # титул: «Бланк № 1 / Всего бланков 1 / Годовая» — как на каждой странице скана
    t = " ".join(str(c.value) for row in wb["Титульный лист"].iter_rows() for c in row if c.value)
    assert "Бланк №  1" in t and "Годовая" in t
