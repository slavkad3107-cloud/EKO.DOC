"""Бланки-образцы: учёт папки «Формы» и заполнение xlsx с формулами."""
import json
from decimal import Decimal
from pathlib import Path

import pytest

from ecodoc.core import forms_registry as fr
from ecodoc.core.models import ReportContext, WasteFlow
from ecodoc.render import template_xlsx as tx


@pytest.fixture()
def forms_dir(tmp_path, monkeypatch):
    d = tmp_path / "Формы"
    (d / "Отчетность" / "Движение отходов 1028").mkdir(parents=True)
    (d / "Разработка" / "ПНООЛР").mkdir(parents=True)
    monkeypatch.setenv("ECODOC_FORMS", str(d))
    return d


def _sample_1028(path: Path) -> Path:
    """Мини-бланк журнала: шапка граф + формула (как в настоящем файле)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Приложение 2 (год)"
    ws["A1"] = "Обобщенные данные учета в области обращения с отходами"
    ws["G3"] = "=Титул!F11"                       # формула-ссылка
    heads = ["№ п/п", "Наименование отходов", "Код ФККО",
             "Класс опасности вида отхода", "Наличие на начало, тонн",
             "накопление", "Образовано отходов в отчетном периоде, тонн",
             "Получено отходов от других лиц, тонн"]
    for i, text in enumerate(heads, start=1):
        ws.cell(row=5, column=i, value=text)
    for i in range(1, len(heads) + 1):
        ws.cell(row=6, column=i, value=i)         # строка нумерации граф
    ws["I7"] = "=G7+H7"                           # расчётная графа
    wb.save(path)
    return path


def test_registry_sees_samples_and_gaps(forms_dir):
    _sample_1028(forms_dir / "Отчетность" / "Движение отходов 1028" / "журнал.xlsx")
    slots = {s.title: s for s in fr.scan()}
    assert slots["Движение отходов 1028"].has_sample
    assert not slots["ПНООЛР"].has_sample
    text = fr.summary()
    assert "есть образцы: 1" in text and "ПНООЛР" in text


def test_registry_without_forms_folder(monkeypatch, tmp_path):
    monkeypatch.setenv("ECODOC_FORMS", str(tmp_path / "нет-такой"))
    monkeypatch.setattr(fr, "roots", lambda: [tmp_path / "нет-такой"])
    assert fr.root() is None
    assert "не найдена" in fr.summary()
    assert all(not s.has_sample for s in fr.scan())


def test_check_new_detects_added_file(forms_dir):
    first = fr.check_new()
    assert first["first_run"] and not first["added"]
    _sample_1028(forms_dir / "Отчетность" / "Движение отходов 1028" / "журнал.xlsx")
    second = fr.check_new()
    assert "Движение отходов 1028" in second["added"]
    third = fr.check_new()
    assert not third["added"]                     # повторно не сигналит


def test_find_missing_searches_disk(forms_dir, tmp_path, monkeypatch):
    stash = tmp_path / "склад"
    stash.mkdir()
    (stash / "ПНООЛР образец 2025.docx").write_text("x", encoding="utf-8")
    monkeypatch.setenv("ECODOC_FORMS_SEARCH", str(stash))
    found = fr.find_missing()
    assert "ПНООЛР" in found and found["ПНООЛР"][0].endswith(".docx")


def test_fill_template_keeps_formulas(forms_dir, tmp_path):
    _sample_1028(forms_dir / "Отчетность" / "Движение отходов 1028" / "журнал.xlsx")
    from ecodoc.reports.waste_movement.template import fill
    ctx = ReportContext()
    ctx.organization.name = "ООО Тест"
    ctx.organization.inn = "7801234564"
    ctx.period.year = 2025
    ctx.wastes = [WasteFlow(fkko_code="47110101521", name="Лампы ртутные",
                            hazard_class=1, generated=Decimal("0.052")),
                  WasteFlow(fkko_code="73310001724", name="Мусор офисный",
                            hazard_class=4, generated=Decimal("1.9"))]
    out = fill(ctx, tmp_path / "журнал_заполнен.xlsx")
    assert out is not None and out.exists()

    wb = tx.open_template(out)
    ws = wb["Приложение 2 (год)"]
    assert ws["G3"].value == "=Титул!F11"          # формула на месте
    assert ws["I7"].value == "=G7+H7"
    texts = [[c.value for c in row] for row in ws.iter_rows(min_row=7, max_row=9)]
    flat = [str(v) for row in texts for v in row if v is not None]
    assert "Лампы ртутные" in flat and "47110101521" in flat
    assert "Мусор офисный" in flat and 1.9 in [v for row in texts for v in row]


def test_fill_returns_none_without_sample(tmp_path, monkeypatch):
    monkeypatch.setenv("ECODOC_FORMS", str(tmp_path / "пусто"))
    from ecodoc.reports.waste_movement.template import fill
    ctx = ReportContext()
    assert fill(ctx, tmp_path / "out.xlsx") is None    # значит рисуем кодом


def test_report_uses_sample_when_available(forms_dir, tmp_path):
    """Форма 1028 берёт бланк пользователя, если он есть."""
    _sample_1028(forms_dir / "Отчетность" / "Движение отходов 1028" / "журнал.xlsx")
    from ecodoc.core import registry
    registry.load_all()
    ctx = ReportContext()
    ctx.organization.name = "ООО Тест"
    ctx.organization.inn = "7801234564"
    ctx.period.year = 2025
    ctx.wastes = [WasteFlow(fkko_code="47110101521", name="Лампы",
                            hazard_class=1, generated=Decimal("0.05"))]
    rep = registry.get("waste-movement")(ctx)
    out = rep.render_print(tmp_path / "1028.xlsx")
    wb = tx.open_template(out)
    # лист из бланка пользователя, а не сгенерированный «Титул/Приложение 1…»
    assert "Приложение 2 (год)" in wb.sheetnames
    assert wb["Приложение 2 (год)"]["G3"].value == "=Титул!F11"


def test_api_forms_registry(forms_dir):
    from ecodoc.gui import server
    _sample_1028(forms_dir / "Отчетность" / "Движение отходов 1028" / "журнал.xlsx")
    out = server.api_forms_registry({}, {"check_new": "1"})
    assert out["have"] >= 1 and out["total"] > 10
    row = next(r for r in out["rows"] if r["code"] == "waste-movement")
    assert row["has"] and row["files"]
