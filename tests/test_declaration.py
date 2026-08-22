"""Тесты расчёта платы и генерации Декларации НВОС."""
import sys
from decimal import Decimal
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from ecodoc.core import serialize
from ecodoc.core.money import D, money
from ecodoc.reports.declaration_nvos.calc import calculate
from ecodoc.reports.declaration_nvos.report import DeclarationNVOS


def _ctx():
    return serialize.from_json(ROOT / "samples" / "example_context.json")


def test_money_roundhalfup():
    assert money("1.005") == Decimal("1.01")
    assert money("663,2") == Decimal("663.20")


def test_calc_matches_manual():
    """2025 год: ставки взяты напрямую из Распоряжения № 1852-р.

    Прежняя схема «ставка 2018 × коэффициент индексации» к ним не применяется;
    единственный множитель — дополнительный коэффициент 1,045 (ПП РФ № 1034)."""
    ctx = _ctx()  # отчётный год 2025
    res = calculate(ctx)
    k2025 = D("1.045")
    # Азота диоксид: 1.2 т × 209.59 (ставка 2025) × 1.045 (norm)
    no2 = next(l for l in res.lines if l.code == "0301")
    assert no2.k_ind == k2025
    assert no2.rate == D("209.59")
    assert no2.amount == money(D("1.2") * D("209.59") * k2025)
    # СО сверх лимита: 0.2 × 2.42 × 1.045 × 100
    co_over = next(l for l in res.lines if l.code == "0337" and l.band == "over")
    assert co_over.amount == money(D("0.2") * co_over.rate * k2025 * D("100"))
    # формула согласована по всем строкам
    for ln in res.lines:
        assert ln.amount == money(ln.mass * ln.rate * ln.k_ind * ln.k_band * ln.k_extra)
    # отход 1 класса (лампы) — передан, не размещён => платы за размещение нет
    assert not any(l.medium == "waste" and l.code == "47110101521" for l in res.lines)
    # итог > 0 и согласован
    assert res.total == money(res.total_air + res.total_water + res.total_waste)
    assert res.total > 0


def test_validate_clean_sample():
    rep = DeclarationNVOS(_ctx())
    rep.ctx.organization.name = rep.ctx.organization.name or "x"
    errors = [i for i in rep.validate() if i.level == "error"]
    assert not errors, errors


def test_render(tmp_path):
    rep = DeclarationNVOS(_ctx())
    xml = rep.render_xml(tmp_path / "d.xml")
    xlsx = rep.render_print(tmp_path / "d.xlsx")
    assert xml.exists() and xml.stat().st_size > 0
    assert xlsx.exists() and xlsx.stat().st_size > 0
    assert "ДекларацияНВОС" in xml.read_text(encoding="utf-8")


def test_print_official_sheets(tmp_path):
    import openpyxl
    rep = DeclarationNVOS(_ctx())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    # состав и порядок листов — как в печатной форме бланка
    assert wb.sheetnames[0] == "стр.1"
    for name in ("Информация о суммах платы", "стр.2", "Авансовые платежи",
                 "Раздел 1 (выбросы)", "Раздел 4 (сбросы)",
                 "Раздел 5 (отходы)"):
        assert name in wb.sheetnames, name
    # лист сумм идёт между титулом и расчётом (стр. 3–4 образца)
    assert (wb.sheetnames.index("Информация о суммах платы")
            < wb.sheetnames.index("стр.2"))
    s2 = wb["стр.2"]
    # официальные коды строк бланка по столбцу A
    codes = {c.value for c in s2["A"]}
    for want in ("010", "020", "021", "023", "024", "025", "100", "120"):
        assert want in codes, want
    # после каждого КБК — своя строка ОКТМО
    for want in ("031", "051", "071", "091", "111"):
        assert want in codes, want
    # детализация ПНГ и ТКО
    for want in ("061", "062", "063", "121", "122", "123"):
        assert want in codes, want
    # КБК ТКО присутствует (строка 110) — сплошными 20 цифрами, как в
    # принятой декларации и ЛКПП
    kbk = {c.value for c in s2["C"]}
    assert "04811201042016000120" in kbk
    assert not any(isinstance(v, str) and v.startswith("048 ") for v in kbk)


def test_calc_codes_edition_241(tmp_path):
    """Итоговая часть стр.2 за 2025 год — по ред. № 241 (Приложение 2 к
    Приказу № 1043): 020 = 021+…+028; 130/131/132 побочные продукты
    (133/134), 136/137/140 породы (141/142), 143/144/145 животноводство;
    150 мероприятия (151–158), 160 корректировка (161–168), 170 к внесению
    (171–178), 180 зачтено (181–185), 186 решение, 190 авансы (191–195),
    200 итог (201–208), 210 возврат (211–215). Старых кодов 2021 г.
    (130 = мероприятия, 150 = к внесению) быть не должно."""
    import openpyxl
    from ecodoc.core.money import fmt_money
    rep = DeclarationNVOS(_ctx())
    assert rep.edition.key == "241"
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    ws = wb["стр.2"]
    rows = {str(r[0].value): (str(r[1].value), r[2].value, r[3].value)
            for r in ws.iter_rows(min_col=1, max_col=4)}
    assert "(020 = 021+022+023+024+025+026+027+028)" in rows["020"][0]
    for want in ("026", "027", "028", "130", "131", "132", "133", "134",
                 "136", "137", "140", "141", "142", "143", "144", "145",
                 "150", "158", "160", "168", "170", "178", "180", "185",
                 "186", "190", "195", "200", "208", "210", "215"):
        assert want in rows, want
    assert "029" not in rows and "147" not in rows   # только в № 182
    assert rows["130"][0].startswith("КБК")
    assert "побочных продуктов производства" in rows["130"][0]
    assert rows["150"][0].startswith("Сумма средств на выполнение мероприятий")
    assert "(150 = 151+152+153+154+155+156+157+158)" in rows["150"][0]
    assert rows["160"][0].startswith("Сумма платы, исчисленная с учетом корректировки")
    assert rows["170"][0].startswith("Сумма платы, подлежащая внесению в бюджет")
    assert rows["180"][0].startswith("Сумма платы, зачтенная в предыдущем")
    assert rows["186"][0].startswith("Номер Решения о зачете")
    assert rows["190"][0].startswith("Сведения о суммах внесенных авансовых")
    assert rows["200"][0].startswith("Итоговая сумма платы для внесения")
    assert rows["210"][0].startswith("Итоговая сумма платы для возврата")
    # суммы в своих строках: 170 = 020 (вычетов нет), 200 = 170 (авансов нет)
    total = fmt_money(rep.calc.total)
    assert rows["020"][2] == total
    assert rows["170"][2] == total
    assert rows["200"][2] == total
    assert rows["171"][2] == fmt_money(rep.calc.by_section["Р1"])
    assert rows["174"][2] == fmt_money(rep.calc.by_section["Р5"])
    assert "(строка 040 − строка 151)" in rows["171"][0]
    # подстроки кварталов — по одной под каждой из 191–195
    codes = [str(c.value) for c in ws["A"]]
    assert codes.count("1 квартал") == 5


def test_edition_switch_by_year(tmp_path):
    """За 2026 год и далее — форма Приказа от 01.04.2026 № 182: 020 = …+029,
    строки 139/140/141 (породы), 142/143/144 (животноводство), 145–149
    (искусственные грунты), 150…159, 170…179, 200…209; реквизиты в
    заголовках — № 182."""
    import openpyxl
    from ecodoc.reports.declaration_nvos.editions import edition_for_year
    assert edition_for_year(2024).key == "241"
    assert edition_for_year(2025).key == "241"
    assert edition_for_year(2026).key == "182"
    assert edition_for_year(2030).key == "182"
    ctx = _ctx()
    ctx.period.year = 2026
    rep = DeclarationNVOS(ctx)
    assert rep.edition.key == "182"
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    ws = wb["стр.2"]
    rows = {str(r[0].value): str(r[1].value)
            for r in ws.iter_rows(min_col=1, max_col=4)}
    assert "(020 = 021+022+023+024+025+026+027+028+029)" in rows["020"]
    for want in ("029", "139", "140", "141", "142", "143", "144", "145",
                 "146", "147", "148", "149", "159", "169", "179", "209"):
        assert want in rows, want
    assert "искусственных грунтов" in rows["029"]
    assert rows["139"].startswith("Сумма платы за размещение вскрышных")
    assert rows["145"].startswith("КБК") and "искусственных грунтов" in rows["145"]
    assert "№ 182" in str(ws["A1"].value)
    assert "№ 182" in str(wb["стр.1"]["B1"].value)
    # в ред. 241 заголовки ссылаются на № 1043 / № 241
    rep25 = DeclarationNVOS(_ctx())
    wb25 = openpyxl.load_workbook(rep25.render_print(tmp_path / "d25.xlsx"))
    assert "№ 241" in str(wb25["стр.1"]["B1"].value)

    # Раздел 4 формы № 182 — 19 граф (без Кп), ред. 241 — 20
    def ncols(w):
        ws4 = w["Раздел 4 (сбросы)"]
        nums = next(r for r in ws4.iter_rows(values_only=True)
                    if r[0] == "1" and r[1] == "2")
        return len([v for v in nums if v])
    assert ncols(wb) == 19 and ncols(wb25) == 20


def test_section10_only_in_182(tmp_path):
    """Искусственные грунты (waste_kind='soil') — Раздел 10, есть только в
    форме № 182; в ред. 241 validate() даёт ошибку."""
    import openpyxl
    from ecodoc.core.models import (Organization, ReportContext, ReportPeriod,
                                    WasteFlow)
    from ecodoc.core.money import fmt_money

    def mk(year):
        return ReportContext(
            organization=Organization(name="Т", inn="7801234564", oktmo="40324000"),
            period=ReportPeriod(year=year),
            wastes=[WasteFlow(fkko_code="81110001495", name="Грунт", hazard_class=4,
                              placed_norm="10", waste_kind="soil")])
    rep = DeclarationNVOS(mk(2026))
    assert rep.calc.by_section["Р10"] > 0
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    assert "Раздел 10 (грунты)" in wb.sheetnames
    rows = {str(r[0].value): r[3].value
            for r in wb["стр.2"].iter_rows(min_col=1, max_col=4)}
    assert rows["147"] == fmt_money(rep.calc.by_section["Р10"])
    assert rows["029"] == rows["147"]
    rep25 = DeclarationNVOS(mk(2025))
    assert any(i.level == "error" and "Раздел 10" in i.field
               for i in rep25.validate())
    assert "Раздел 10 (грунты)" not in openpyxl.load_workbook(
        rep25.render_print(tmp_path / "d25.xlsx")).sheetnames


def test_section5_blank_layout(tmp_path):
    """Раздел 5 — 27 нумерованных граф ред. 241, шапка объекта с ОКТМО, блок
    ОРО (реквизиты документа, № ГРОРО, характеристика), одна строка на отход,
    13 = 14 + 15, 27 = 25 + 26, строки «ИТОГО» и «Всего по тем классам…»."""
    import openpyxl
    from ecodoc.core.money import fmt_money
    ctx = _ctx()
    extra = ctx.extra if isinstance(ctx.extra, dict) else {}
    ctx.extra = {**extra, "declaration": {
        "waste_permit": {"number": "ДВОС-1", "valid_until": "31.12.2030"},
        "oro": {"name": "Полигон «Новосёлки»", "number": "78-00001-З-00592-250914",
                "address": "СПб", "status": "groro"},
        "waste_limits": {"40211001515": "15.0"}}}
    rep = DeclarationNVOS(ctx)
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    ws = wb["Раздел 5 (отходы)"]
    rows = list(ws.iter_rows(values_only=True))
    num_row = next(r for r in rows if r[0] == "1" and r[26] == "27")
    assert list(num_row[:27]) == [str(i) for i in range(1, 28)]
    text = " ".join(str(c) for r in rows for c in r if c)
    for want in ("ОКТМО объекта", "Реквизиты разрешительного документа",
                 "ДВОС-1", "Регистрационный номер объекта размещения отходов",
                 "78-00001-З-00592-250914",
                 "Включен в государственный реестр объектов размещения отходов",
                 "ИТОГО", "Всего по тем классам опасности"):
        assert want in text, want
    groro = next(r for r in rows if r[0] and "Включен в государственный реестр"
                 in str(r[0]))
    assert "V" in [c for c in groro if c]
    r5 = [ln for ln in rep.calc.lines if ln.section == "Р5"]
    data = [r for r in rows if isinstance(r[0], int)]
    assert len(data) == len({ln.code for ln in r5})
    office = next(r for r in data if r[2] == "40211001515")
    assert office[4] == "15.0"                        # лимит из extra
    assert office[12] == office[13] + office[14]      # 13 = 14 + 15
    assert office[26] == fmt_money(sum(ln.amount for ln in r5
                                       if ln.code == "40211001515"))
    itogo = next(r for r in rows if r[0] == "ИТОГО")
    assert itogo[26] == fmt_money(sum(ln.amount for ln in r5))


def test_section4_blank_layout(tmp_path):
    """Раздел 4 ред. 241 — 20 нумерованных граф, выпуск с ОКТМО, «Итого по
    всем выпускам»; Итого = сумма платы Р4."""
    import openpyxl
    from ecodoc.core.models import (Medium, Organization, Pollutant,
                                    ReportContext, ReportPeriod)
    from ecodoc.core.money import fmt_money
    ctx = ReportContext(
        organization=Organization(name="Т", inn="7801234564", oktmo="40324000"),
        period=ReportPeriod(year=2025),
        pollutants=[Pollutant(name="Взвешенные вещества", code="",
                              medium=Medium.WATER, mass_norm="1.5",
                              mass_over="0.2")],
        extra={"declaration": {"water_outlet": {"number": "1",
                                                "oktmo": "40324000"}}})
    rep = DeclarationNVOS(ctx)
    r4 = [ln for ln in rep.calc.lines if ln.section == "Р4"]
    assert r4 and sum(ln.amount for ln in r4) > 0
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    ws = wb["Раздел 4 (сбросы)"]
    rows = list(ws.iter_rows(values_only=True))
    num_row = next(r for r in rows if r[0] == "1" and r[19] == "20")
    assert list(num_row[:20]) == [str(i) for i in range(1, 21)]
    text = " ".join(str(c) for r in rows for c in r if c)
    assert "Выпуск" in text and "ОКТМО выпуска 40324000" in text
    assert "Итого по всем выпускам" in text
    itogo = next(r for r in rows if r[0] == "Итого")
    assert itogo[19] == fmt_money(sum(ln.amount for ln in r4))


def test_title_fields_by_blank(tmp_path):
    """Титул: шапка «Приложение 2 к приказу…», поле 1 — две клетки
    первичный/уточненный с отметкой V, поле 2 — ТО РПН по региону объекта
    (справочник), поле 9 — два числовых пропуска."""
    import openpyxl
    ctx = _ctx()
    extra = ctx.extra if isinstance(ctx.extra, dict) else {}
    ctx.extra = {**extra, "declaration": {"pages": 12, "attachment_sheets": 3}}
    rep = DeclarationNVOS(ctx)
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    s1 = wb["стр.1"]
    assert str(s1["B1"].value).startswith("Приложение 2 к приказу Минприроды России")
    rows = {str(r[0].value): (str(r[1].value), r[2].value, r[3].value)
            for r in s1.iter_rows(min_col=1, max_col=4)}
    assert rows["1"][1] == "первичный [V]"
    assert str(rows["1"][2]).startswith("уточненный [ ]")
    # объект СПб (код ТО РПН 40) → Северо-Западное МУ Росприроднадзора
    assert "Северо-Западное межрегиональное управление" in str(rows["2"][1])
    assert ("Организационно-правовая форма юридического лица и его полное "
            "наименование") in rows["3"][0]
    assert "Идентификационный номер налогоплательщика" in rows["7"][0]
    assert "Код причины постановки на учет" in rows["8"][0]
    assert rows["9"][1] == "страниц: 12"
    assert rows["9"][2] == "листов приложений: 3"
    # уточнённая декларация — V во второй клетке и номер корректировки
    ctx.extra["declaration"]["correction"] = "1"
    wb2 = openpyxl.load_workbook(
        DeclarationNVOS(ctx).render_print(tmp_path / "d2.xlsx"))
    r1 = next(r for r in wb2["стр.1"].iter_rows(values_only=True) if r[0] == "1")
    assert r1[2] == "первичный [ ]"
    assert "уточненный [V]" in r1[3] and r1[3].endswith("1")


def test_rospr_directory():
    from ecodoc.reports.declaration_nvos.editions import rosprirodnadzor_for
    assert "Северо-Западное" in rosprirodnadzor_for("78")
    assert "Северо-Западное" in rosprirodnadzor_for("47")
    assert "Северо-Западное" in rosprirodnadzor_for("40")      # код ТО РПН СПб
    assert "Северо-Западное" in rosprirodnadzor_for("", "40908000")  # по ОКТМО
    assert "Москве и Калужской" in rosprirodnadzor_for("77")
    assert "Центральное" in rosprirodnadzor_for("50")
    assert rosprirodnadzor_for("66") == ""


def test_validate_warns_missing_blank_data():
    """Пустые графы бланка не уходят молча: validate() предупреждает про
    отсутствующие реквизиты ОРО / документа / лимитов (Разделы 5–10)."""
    rep = DeclarationNVOS(_ctx())
    fields = {i.field for i in rep.validate() if i.level == "warning"}
    assert "Разделы 5–10" in fields


def test_title_all_17_fields(tmp_path):
    """Титул содержит поля 1–17 бланка; ОГРН и «сокращённого наименования»
    в бланке нет (проверено по принятой декларации: 0 вхождений)."""
    import openpyxl
    rep = DeclarationNVOS(_ctx())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    s1 = wb["стр.1"]
    nums = {str(c.value) for c in s1["A"]}
    for want in [str(i) for i in range(1, 18)]:
        assert want in nums, f"нет поля {want} титульного листа"
    text = " ".join(str(c.value) for row in s1.iter_rows()
                    for c in row if c.value)
    assert "ОГРН" not in text
    assert "Сокращённое" not in text and "Сокращенное" not in text
    assert "страницах с приложением подтверждающих документов" in text


def test_summary_sheet(tmp_path):
    """Лист «Информация о суммах платы»: строка ИТОГО равна начисленной плате
    (вычетов в образце данных нет), есть блок подписи исполнителя."""
    import openpyxl
    from ecodoc.core.money import fmt_money
    rep = DeclarationNVOS(_ctx())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    ws = wb["Информация о суммах платы"]
    pairs = [(a.value, b.value) for a, b in zip(ws["A"], ws["B"])]
    itogo = [b for a, b in pairs if a == "ИТОГО"]
    assert itogo and itogo[0] == fmt_money(rep.calc.total)
    text = " ".join(str(a) for a, _ in pairs if a)
    assert "Исполнитель" in text


def test_advances_sheet(tmp_path):
    """Лист «Информация об авансовых платежах»: строки 010–060, выбранный
    способ исчисления отмечается знаком X из extra['declaration']."""
    import openpyxl
    ctx = _ctx()
    extra = ctx.extra if isinstance(ctx.extra, dict) else {}
    ctx.extra = {**extra, "declaration": {"advance_method": {"air": "pek"}}}
    rep = DeclarationNVOS(ctx)
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    ws = wb["Авансовые платежи"]
    rows = list(ws.iter_rows(values_only=True))
    codes = {r[1] for r in rows}
    for want in ("010", "020", "030", "040", "050", "060"):
        assert want in codes, want
    x_rows = [r for r in rows if r[2] == "X"]
    assert len(x_rows) == 1, "X должен стоять ровно у одного способа"
    assert "производственного экологического контроля" in str(x_rows[0][0])


def test_section1_blank_layout(tmp_path):
    """Раздел 1: 18 нумерованных граф, одна строка на вещество (полосы
    раскладываются в графы 6/7/8 и 15/16/17, а не плодят строки)."""
    import openpyxl
    from ecodoc.core.money import fmt_money, money
    rep = DeclarationNVOS(_ctx())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    ws = wb["Раздел 1 (выбросы)"]
    rows = list(ws.iter_rows(values_only=True))
    num_row = next(r for r in rows if r[0] == "1" and r[17] == "18")
    assert list(num_row[:18]) == [str(i) for i in range(1, 19)]
    r1 = [ln for ln in rep.calc.lines if ln.section == "Р1"]
    uniq = {(ln.code, ln.name) for ln in r1}
    data_rows = [r for r in rows if isinstance(r[0], int)]
    assert len(data_rows) == len(uniq)
    # СО в образце данных имеет и «норматив», и «сверх»: обе полосы в одной
    # строке, гр.18 = сумма платы по всем полосам вещества
    co_lines = [ln for ln in r1 if ln.code == "0337"]
    assert len(co_lines) > 1, "образец должен содержать СО в двух полосах"
    co_row = next(r for r in data_rows if r[1] == co_lines[0].name)
    assert co_row[17] == fmt_money(sum(ln.amount for ln in co_lines))
    assert co_row[4] == float(sum(ln.mass for ln in co_lines))
    # итог листа = сумма платы Раздела 1
    total = money(sum(ln.amount for ln in r1))
    itogo = [r for r in rows if r[0] == "Итого по стационарным источникам:"]
    assert itogo and itogo[0][17] == fmt_money(total)


def test_calc_tko_rows_match_blank(tmp_path):
    """Строки 121–123 по бланку: 121 — плата за размещение ПРИНЯТЫХ ТКО
    (регоператор; в расчёте не участвует, по умолчанию 0), 122 — в пределах
    лимита, 123 — сверх лимита. Раньше суммы стояли со сдвигом на строку
    (121=лимит, 122=сверх)."""
    import openpyxl
    from ecodoc.core.models import (Organization, ReportContext, ReportPeriod,
                                    WasteFlow)
    from ecodoc.core.money import fmt_money
    ctx = ReportContext(
        organization=Organization(name="Т", inn="7801234564", oktmo="40324000"),
        period=ReportPeriod(year=2025),
        wastes=[WasteFlow(fkko_code="73310001724", name="ТКО", hazard_class=4,
                          placed_norm="3", placed_over="1")])
    rep = DeclarationNVOS(ctx)
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    ws = wb["стр.2"]
    rows = {str(r[0].value): (str(r[1].value), r[3].value)
            for r in ws.iter_rows(min_col=1, max_col=4)}
    norm = sum(l.amount for l in rep.calc.lines
               if l.section == "Р6" and l.band == "norm")
    over = sum(l.amount for l in rep.calc.lines
               if l.section == "Р6" and l.band == "over")
    assert norm > 0 and over > 0
    assert "принятых ТКО" in rows["121"][0]
    assert rows["121"][1] == fmt_money(0)
    assert "в пределах установленного лимита" in rows["122"][0]
    assert rows["122"][1] == fmt_money(norm)
    assert "сверх установленного лимита" in rows["123"][0]
    assert rows["123"][1] == fmt_money(over)
    # формула бланка 120 = 121 + 122 + 123
    assert rows["120"][1] == fmt_money(norm + over)


def test_calc_blank_labels(tmp_path):
    """Заголовки итоговой части — по тексту Приложения 2 ред. 241:
    020 «без учета корректировки», 180 «в предыдущем отчетном периоде»,
    210 «для возврата и/или зачета»; подстроки кварталов 191–195 стоят
    в графе кодов, как в бланке."""
    import openpyxl
    rep = DeclarationNVOS(_ctx())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "d.xlsx"))
    ws = wb["стр.2"]
    label = {str(a.value): str(b.value) for a, b in zip(ws["A"], ws["B"])}
    assert "без учета корректировки" in label["020"]
    assert "в предыдущем отчетном периоде" in label["180"]
    assert "для возврата и/или зачета" in label["210"]
    assert "ПНГ в пределах НДВ, ТН" in label["061"]
    codes = [str(c.value) for c in ws["A"]]
    assert codes.count("1 квартал") == 5  # по одному под каждой из 191–195


def test_k_st_applies_to_waste_placement():
    """Кст (стимулирующий коэффициент) умножает плату за размещение."""
    from ecodoc.core.models import (Organization, ReportContext, ReportPeriod,
                                    WasteFlow)
    ctx = ReportContext(
        organization=Organization(name="Т", inn="7801234564", oktmo="40324000"),
        period=ReportPeriod(year=2025),
        wastes=[
            WasteFlow(fkko_code="34620001000", name="без Кст", hazard_class=4,
                      placed_norm="10"),
            WasteFlow(fkko_code="34620001001", name="Кст 0.3", hazard_class=4,
                      placed_norm="10", k_st="0.3"),
        ])
    res = calculate(ctx)
    plain = next(l for l in res.lines if l.name == "без Кст")
    stim = next(l for l in res.lines if l.name == "Кст 0.3")
    assert stim.k_extra == D("0.3")
    assert stim.amount == money(plain.amount * D("0.3"))


def test_declaration_sections_tko_split(tmp_path):
    """ТКО (ФККО «7 3…») уходит в Р6, отходы производства — в Р5."""
    from ecodoc.core.models import ReportContext, ReportPeriod, Organization, WasteFlow
    from ecodoc.reports.declaration_nvos.calc import calculate
    ctx = ReportContext(
        organization=Organization(name="Т", inn="7801234564", oktmo="40324000"),
        period=ReportPeriod(year=2025),
        wastes=[
            WasteFlow(fkko_code="34620001000", name="отход произв.", hazard_class=4,
                      placed_norm="5"),
            WasteFlow(fkko_code="73310001724", name="ТКО", hazard_class=4,
                      placed_norm="3"),
        ])
    c = calculate(ctx)
    assert c.by_section["Р5"] > 0 and c.by_section["Р6"] > 0
    # согласованность: сумма разделов == итог
    assert money(sum(c.by_section.values())) == c.total


if __name__ == "__main__":
    test_money_roundhalfup()
    test_calc_matches_manual()
    test_validate_clean_sample()
    import tempfile
    test_render(Path(tempfile.mkdtemp()))
    print("OK — все проверки прошли")
