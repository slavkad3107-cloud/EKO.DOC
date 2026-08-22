"""Отчёт по ПЭК — структура по Приказу № 173 (титул + таблицы 1.1–6.2).

Шапки таблиц 2.1/2.2/2.3/3.1/3.2/5.1/6.1 проверяются по ДОСЛОВНОМУ составу
граф принятого отчёта (Otcet-o-PEK_6303183.docx) и текста № 173 ред. № 262."""
import openpyxl

from ecodoc.core import registry
from ecodoc.core.models import (NVOSObject, Organization, Pollutant, Medium,
                                ReportContext, ReportPeriod, WasteFlow)


def _ctx(year=2025, **extra):
    e = {"pek": {"program_number": "1-ПЭК", "lab": "ООО Лаб (RA.RU.21XX)"},
         "ppp": [{"name": "Шлак", "formed": "10", "used": "8"}]}
    e.update(extra)
    return ReportContext(
        organization=Organization(name="ООО Т", inn="7801234564", kpp="780101001"),
        period=ReportPeriod(year=year),
        objects=[NVOSObject(code="40-0178-001234-П", name="Производственная площадка",
                            category="III", oktmo="40908000")],
        pollutants=[Pollutant(name="Азота диоксид", code="0301", medium=Medium.AIR,
                              mass_norm="1.0")],
        wastes=[WasteFlow(fkko_code="73310001724", name="ТКО", hazard_class=4,
                          generated="5", received="2", transferred="5",
                          transferred_util="3", transferred_burial="2",
                          accumulated_start_nakopl="1",
                          accumulated_end_nakopl="3")],
        extra=e)


def _render(tmp_path, ctx=None):
    registry.load_all()
    rep = registry.get("pek")(ctx or _ctx())
    return openpyxl.load_workbook(rep.render_print(tmp_path / "pek.xlsx"))


def _text(ws):
    return " || ".join(str(c.value) for row in ws.iter_rows() for c in row
                       if c.value not in (None, ""))


def _row_after_title(ws, title_start):
    """Номер строки заголовка таблицы, начинающегося с title_start."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.startswith(title_start):
            return r
    raise AssertionError(f"нет таблицы {title_start!r}")


def _header_rows(ws, title_start, ncols):
    """Ярусы шапки таблицы: от строки под заголовком до строки нумерации
    граф «1..ncols»; возвращает (список ярусов, номер строки нумерации)."""
    r = _row_after_title(ws, title_start) + 1
    tiers = []
    while not (ws.cell(row=r, column=1).value == 1
               and ws.cell(row=r, column=ncols).value == ncols):
        tiers.append([ws.cell(row=r, column=c).value for c in range(1, ncols + 1)])
        r += 1
        assert r < ws.max_row, "строка нумерации граф не найдена"
    return tiers, r


def _bottom(tiers, ncols):
    """Нижний ярус шапки: для каждой графы — последний непустой текст."""
    out = []
    for c in range(ncols):
        vals = [t[c] for t in tiers if t[c] not in (None, "")]
        out.append(vals[-1] if vals else None)
    return out


def test_pek_six_sections(tmp_path):
    wb = _render(tmp_path)
    assert wb.sheetnames == ["Титул", "Раздел 1", "Раздел 2 (воздух)",
                             "Раздел 3 (вода)", "Раздел 4 (отходы)",
                             "Раздел 5 (ППП)", "Раздел 6 (искусств. грунты)"]
    ws = wb["Раздел 5 (ППП)"]
    _, num = _header_rows(ws, "Таблица 5.1.", 17)
    # 5.1 без графы «№» — первая графа «Наименование ППП»
    assert ws.cell(row=num + 1, column=1).value == "Шлак"
    assert ws.cell(row=num + 1, column=5).value == "10"    # образовано
    assert ws.cell(row=num + 1, column=6).value == "8"     # использовано


def test_pek_all_20_tables_present(tmp_path):
    """Форма № 173 — 20 пронумерованных таблиц; незаполненные печатаются
    с шапкой и прочерками, а не пропускаются."""
    wb = _render(tmp_path)
    text = " || ".join(_text(ws) for ws in wb.worksheets)
    for num in ("1.1", "1.2", "1.3", "2.1", "2.2", "2.3", "2.4", "2.5", "2.6",
                "3.1", "3.2", "3.3", "3.4", "4.1", "4.2", "4.3",
                "5.1", "5.2", "6.1", "6.2"):
        assert f"Таблица {num}." in text, f"нет таблицы {num}"


def test_pek_section_and_table_titles_verbatim(tmp_path):
    """Названия разделов и таблиц — дословно по № 173 / принятому отчёту
    (по ним ЛК РПН опознаёт таблицы)."""
    wb = _render(tmp_path)
    text = " || ".join(_text(ws) for ws in wb.worksheets)
    for t in (
        "2. Сведения о результатах осуществления производственного "
        "экологического контроля в области охраны атмосферного воздуха",
        "Таблица 1.3. Сведения о собственных и (или) привлекаемых "
        "испытательных лабораториях (центрах), аккредитованных в соответствии "
        "с законодательством Российской Федерации об аккредитации в "
        "национальной системе аккредитации",
        "Таблица 2.1. Перечень загрязняющих веществ, включенных в план-график "
        "контроля стационарных источников выбросов",
        "Таблица 2.2. Результаты контроля стационарных источников выбросов "
        "загрязняющих веществ в атмосферный воздух",
        "Таблица 2.3. Перечень загрязняющих веществ, включенных в план-график "
        "проведения наблюдений за загрязнением атмосферного воздуха",
        "Таблица 3.1. Сведения о результатах учета объема забора (изъятия) "
        "водных ресурсов из водных объектов и объема сброса сточных, в том "
        "числе дренажных, вод, их качества",
        "Таблица 3.2. Сведения о результатах наблюдения за водными объектами "
        "и их водоохранными зонами",
        "Таблица 4.1. Сведения о результатах мониторинга состояния и "
        "загрязнения окружающей среды на территории объекта размещения "
        "отходов и в пределах его воздействия на окружающую среду",
        "Таблица 5.1. Сведения об образовании и обращении ППП",
        "Таблица 6.1. Сведения об образовании и обращении ИГ",
    ):
        assert t in text, f"нет дословного названия: {t[:50]}…"
    assert "Результаты контроля качества сточных вод" not in text
    assert "Точка контроля" not in text          # самодельная таблица убрана


def test_pek_table_21_and_23_two_columns(tmp_path):
    """2.1 и 2.3 — перечни ЗВ: ровно 2 графы «№ п/п | Наименование
    загрязняющего вещества» (эталон TABLE 15/20)."""
    ctx = _ctx()
    ctx.extra["pek"]["air_observed"] = ["Азота диоксид", "Углерод оксид"]
    wb = _render(tmp_path, ctx)
    ws = wb["Раздел 2 (воздух)"]
    for title in ("Таблица 2.1.", "Таблица 2.3."):
        tiers, num = _header_rows(ws, title, 2)
        assert tiers == [["№ п/п", "Наименование загрязняющего вещества"]]
        assert ws.cell(row=num, column=3).value is None      # нет третьей графы
        assert ws.cell(row=num + 1, column=1).value == 1
        assert ws.cell(row=num + 1, column=2).value == "Азота диоксид"
    # 2.3: вторая строка — из перечня наблюдений
    _, num = _header_rows(ws, "Таблица 2.3.", 2)
    assert ws.cell(row=num + 2, column=2).value == "Углерод оксид"


def test_pek_table_22_twelve_columns_verbatim(tmp_path):
    """2.2 — 12 граф, двухъярусная шапка (эталон TABLE 19 R0-R2); гр. 9 —
    превышение в раз = гр. 8 / гр. 7, гр. 11 — число случаев превышения."""
    ctx = _ctx(emission_sources=[{
        "number": "0001", "name": "Труба котельной", "workshop_no": "1",
        "workshop": "Котельная",
        "pollutants": [{"code": "0301", "name": "Азота диоксид",
                        "g_s_norm": "0.5", "g_s": "0.6", "date": "12.05.2025",
                        "exceed_cases": 1, "note": "протокол № 7"}]}])
    wb = _render(tmp_path, ctx)
    ws = wb["Раздел 2 (воздух)"]
    tiers, num = _header_rows(ws, "Таблица 2.2.", 12)
    assert len(tiers) == 2
    top = tiers[0]
    assert top[0] == "№ п/п"
    assert top[1] == "Структурное подразделение (площадка, цех или другое)"
    assert top[3] == "Источник"
    assert _bottom(tiers, 12) == [
        "№ п/п", "Номер", "Наименование", "Номер", "Наименование",
        "Наименование загрязняющего вещества",
        "Предельно допустимый выброс или временно согласованный выброс, г/с",
        "Фактический выброс, г/с",
        "Превышение предельно допустимого выброса или временно согласованного "
        "выброса в раз (гр. 8/гр. 7)",
        "Дата отбора проб",
        "Общее количество случаев превышения предельно допустимого выброса "
        "или временно согласованного выброса",
        "Примечание"]
    row = [ws.cell(row=num + 1, column=c).value for c in range(1, 13)]
    assert row == [1, "1", "Котельная", "0001", "Труба котельной",
                   "Азота диоксид", "0.5", "0.6", 1.2, "12.05.2025", 1,
                   "протокол № 7"]


def test_pek_tables_31_32_letters(tmp_path):
    """3.1 — 3 графы (реквизиты письма / террорган Росводресурсов /
    количество ЗВ в забранной воде), 3.2 — 2 графы по № 173; без графы «№»."""
    ctx = _ctx()
    ctx.extra["pek"]["water_letters"] = [
        {"letter": "№ 12 от 20.01.2026", "authority": "Невско-Ладожское БВУ",
         "intake_pollutants_t": "-"}]
    ctx.extra["pek"]["water_observation"] = [
        {"letter": "№ 13 от 20.01.2026", "authority": "Невско-Ладожское БВУ"}]
    wb = _render(tmp_path, ctx)
    ws = wb["Раздел 3 (вода)"]
    tiers, num = _header_rows(ws, "Таблица 3.1.", 3)
    assert len(tiers) == 1
    h = tiers[0]
    assert h[0].startswith("Реквизиты письма (номер (при наличии) и дата), "
                           "которым направлены сведения о результатах учета "
                           "забора (изъятия) водных ресурсов")
    assert h[1].startswith("Территориальный орган Росводресурсов")
    assert h[2].startswith("Количество загрязняющих веществ, содержащихся в "
                           "забранной (изъятой) воде")
    assert h[2].endswith("тонн/год (заполняется в случае использования одного "
                         "и того же водного объекта для забора воды и для "
                         "сброса сточных вод)")
    assert ws.cell(row=num, column=4).value is None
    assert [ws.cell(row=num + 1, column=c).value for c in (1, 2, 3)] == \
        ["№ 12 от 20.01.2026", "Невско-Ладожское БВУ", "-"]

    tiers, num = _header_rows(ws, "Таблица 3.2.", 2)
    assert tiers == [[
        "Реквизиты письма (номер (при наличии) и дата), которым направлены "
        "сведения о результатах наблюдения за водными объектами и их "
        "водоохранными зонами",
        "Территориальный орган Росводресурсов, в который направлены сведения "
        "о результатах наблюдения за водными объектами и их водоохранными "
        "зонами"]]
    assert ws.cell(row=num + 1, column=1).value == "№ 13 от 20.01.2026"


def test_pek_tables_51_61_columns(tmp_path):
    """5.1 — 17 граф, 6.1 — 18 граф (№ 173 ред. № 262), трёхъярусная шапка
    «Произведено продукции с использованием ППП/ИГ»."""
    wb = _render(tmp_path)
    ws = wb["Раздел 5 (ППП)"]
    tiers, num = _header_rows(ws, "Таблица 5.1.", 17)
    assert len(tiers) == 3
    assert ws.cell(row=num, column=18).value is None
    b = _bottom(tiers, 17)
    assert b[0] == "Наименование ППП"
    assert b[1].startswith("Код ППП по Общероссийскому классификатору")
    assert b[4] == "Образовано ППП, тонн"
    assert b[5] == "Использовано ППП в собственном производстве, тонн"
    assert b[6] == b[9] == "Наименование продукции"
    assert b[8] == b[11] == "Количество, тонн"
    assert b[12] == "Передано ППП другим лицам в качестве сырья, тонн"
    assert b[14] == "Наличие ППП на конец года, тонн"
    assert b[16].startswith("Область применения продукции")
    assert tiers[0][6] == "Произведено продукции с использованием ППП"
    assert tiers[1][6] == "в качестве сырья"
    assert tiers[1][9] == "в качестве продукции для потребления"
    assert "Отнесено к отходам" not in _text(ws)

    ws6 = wb["Раздел 6 (искусств. грунты)"]
    tiers, num = _header_rows(ws6, "Таблица 6.1.", 18)
    assert ws6.cell(row=num, column=19).value is None
    b = _bottom(tiers, 18)
    assert b[0] == "Наименование ИГ"
    assert b[4] == "Масса произведенного ИГ, тонн"
    assert b[5].startswith("Реквизиты документов по стандартизации")
    assert b[6] == "Использовано ИГ в собственном производстве, тонн"
    assert b[15] == "Наличие ИГ на конец года, тонн"


def test_pek_table_42_21_graf(tmp_path):
    """Таблица 4.2 — 21 графа по № 173: разбивка «наличие» на хранение и
    накопление, «получено от других лиц», «передано» по пяти целям; гр. 20/21
    на конец года — хранение / накопление симметрично гр. 5/6."""
    wb = _render(tmp_path)
    ws = wb["Раздел 4 (отходы)"]
    num_row = next(r for r in range(1, ws.max_row + 1)
                   if ws.cell(row=r, column=1).value == 1
                   and ws.cell(row=r, column=21).value == 21)
    row = [ws.cell(row=num_row + 1, column=c).value for c in range(1, 22)]
    assert row[0] == 1                    # N строки
    assert row[1] == "ТКО"                # гр.2 наименование — раньше был ФККО
    assert row[2] == "73310001724"        # гр.3 код по ФККО
    assert row[4] == 0.0                  # гр.5 хранение на начало года
    assert row[5] == 1.0                  # гр.6 накопление на начало года
    assert row[7] == 2.0                  # гр.8 получено от других лиц
    assert row[10] == 5.0                 # гр.11 передано всего
    assert row[12] == 3.0                 # гр.13 передано для утилизации
    assert row[15] == 2.0                 # гр.16 передано для захоронения
    assert row[17] == 0.0 and row[18] == 0.0   # гр.18-19 собственные ОРО
    assert row[19] == 0.0                 # гр.20 хранение на конец года
    assert row[20] == 3.0                 # гр.21 накопление на конец года


def test_pek_empty_tables_have_headers_and_dashes(tmp_path):
    """Таблицы без данных (очистные сооружения 3.4 — 17 граф, мониторинг
    ОРО 4.1) присутствуют: шапка + строка нумерации + строка прочерков."""
    wb = _render(tmp_path)
    ws = wb["Раздел 3 (вода)"]
    num_row = next(r for r in range(1, ws.max_row + 1)
                   if ws.cell(row=r, column=1).value == 1
                   and ws.cell(row=r, column=17).value == 17)
    assert ws.cell(row=num_row + 1, column=1).value == "-"
    ws4 = wb["Раздел 4 (отходы)"]
    assert any("мониторинга" in str(c.value) for row in ws4.iter_rows()
               for c in row if c.value)


def test_pek_section1_has_inn_ogrn(tmp_path):
    """Таблица 1.1 — реквизитник формы: ИНН и код объекта попадают в
    раздел 1 (раньше ИНН печатался только на самодельном титуле)."""
    wb = _render(tmp_path)
    ws = wb["Раздел 1"]
    cells = {str(c.value) for row in ws.iter_rows() for c in row if c.value}
    assert "7801234564" in cells                   # ИНН
    assert "40-0178-001234-П" in cells             # код объекта
    assert any("лаборатор" in v.lower() for v in cells)  # таблица 1.3


def test_pek_table_13_lab_string_split(tmp_path):
    """Старый ключ pek.lab «ООО Лаб (RA.RU.21XX)» разбирается: наименование
    в гр. 2, аттестат в гр. 4, адрес — прочерк (раньше гр. 3-4 были пусты)."""
    wb = _render(tmp_path)
    ws = wb["Раздел 1"]
    _, num = _header_rows(ws, "Таблица 1.3.", 4)
    assert [ws.cell(row=num + 1, column=c).value for c in range(1, 5)] == \
        [1, "ООО Лаб", "-", "RA.RU.21XX"]


def test_pek_title_blocks(tmp_path):
    """Титул бланка: подпись руководителя (без грифа «УТВЕРЖДАЮ» — его в
    форме нет), наименование объекта НВОС после «…контроля на», исполнитель,
    «Экз. №», строка «место нахождения, год»."""
    wb = _render(tmp_path)
    ws = wb["Титул"]
    text = " || ".join(str(c.value) for row in ws.iter_rows() for c in row
                       if c.value not in (None, ""))
    assert "Экз. №" in text
    assert "УТВЕРЖДАЮ" not in text
    assert "Руководитель юридического лица" in text
    assert "производственного экологического контроля на" in text
    assert "Производственная площадка" in text        # наименование объекта
    assert "(полное наименование объекта" in text
    assert "Исполнитель, ответственный за подготовку отчёта" in text
    assert "(должность)" in text
    assert "место нахождения (город, населенный пункт)" in text
    for wrong in ("Место нахождения || ", "Телефон / e-mail"):
        assert wrong not in text, f"на титуле лишние реквизиты: {wrong}"


def test_pek_title_has_no_section1_requisites(tmp_path):
    """Страховка от отката к самодельному титулу-реквизитнику: программа
    ПЭК, лаборатория и срок представления — не титульные поля (это
    раздел 1 формы), на титуле их быть не должно."""
    wb = _render(tmp_path)
    ws = wb["Титул"]
    text = " || ".join(str(c.value) for row in ws.iter_rows() for c in row
                       if c.value)
    for wrong in ("Программа ПЭК", "Срок представления",
                  "аттестат аккредитации"):
        assert wrong not in text, f"на титуле лишнее поле: {wrong}"


def test_pek_validate_warns_missing_pek_data():
    """Данных для таблиц 1.3/2.2/2.3/3.1/3.2 в модели нет — validate()
    подсказывает, какие ключи extra.pek внести."""
    registry.load_all()
    ctx = _ctx()
    ctx.extra["pek"] = {}
    ctx.pollutants.append(Pollutant(name="Взвешенные вещества", code="",
                                    medium=Medium.WATER, mass_norm="0.1"))
    msgs = " || ".join(i.message for i in registry.get("pek")(ctx).validate()
                       if i.level == "warning")
    for key in ("extra.pek.labs", "extra.emission_sources",
                "extra.pek.air_observed", "extra.pek.water_letters",
                "extra.pek.water_observation"):
        assert key in msgs, key


def test_pek_xml_has_new_sections(tmp_path):
    registry.load_all()
    rep = registry.get("pek")(_ctx())
    xml = rep.render_xml(tmp_path / "pek.xml").read_text(encoding="utf-8")
    assert "ПобочныеПродуктыПроизводства" in xml
    assert "ИскусственныеГрунтыТКО" in xml


def test_pek_table_11_ten_rows_like_accepted_report(tmp_path):
    """Сверка с «Формы/Отчетность/ОТчет ПЭК/отчет ПЭК.pdf» (ИП Миних, отчёт за
    2021 год по СТАРОЙ форме № 261 в ред. № 383 — старее действующей № 173
    ред. № 262, поэтому форму не меняем). Что совпадает и в старой, и в новой
    форме: Таблица 1.1 — ровно 10 строк в этом порядке (наименование, адрес,
    руководитель, ответственные за ПЭК, ИНН, ОГРН, наименование объекта,
    адрес объекта, код объекта, категория); Таблица 1.3 — 4 графы
    (№, наименование, адрес, реквизиты аттестата); ссылок на № 261 нет."""
    wb = _render(tmp_path)
    ws = wb["Раздел 1"]
    r = _row_after_title(ws, "Таблица 1.1.")
    names = [ws.cell(row=r + 3 + i, column=2).value for i in range(10)]
    assert [ws.cell(row=r + 3 + i, column=1).value for i in range(10)] == list(range(1, 11))
    assert names[0].startswith("Полное (сокращённое) наименование")
    assert names[1] == "Место нахождения (адрес)"
    assert names[4] == "ИНН" and names[5] == "ОГРН"
    assert names[6].startswith("Наименование объекта")
    assert names[7] == "Адрес места нахождения объекта"
    assert names[8] == "Код объекта" and names[9] == "Категория объекта"
    assert ws.cell(row=r + 3 + 9, column=3).value == "III"
    text = _text(wb["Титул"]) if "Титул" in wb.sheetnames else _text(wb[wb.sheetnames[0]])
    assert "№ 261" not in text and "N 261" not in text
