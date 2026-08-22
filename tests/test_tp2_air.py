"""2-ТП (воздух) — пять разделов бланка, строки 101–109, графы А/1/Б/2-7 (Приказ № 661)."""
import openpyxl
from lxml import etree

from ecodoc.core import registry
from ecodoc.core.models import (Medium, NVOSObject, Organization, Pollutant,
                                ReportContext, ReportPeriod)


def _ctx(extra=None, objects=None):
    return ReportContext(
        organization=Organization(name="ООО Т", inn="7801234564", okpo="12345678",
                                  oktmo="40324000", okved="30.20.9",
                                  address="190000, Санкт-Петербург, ул. Тестовая, 1"),
        objects=objects if objects is not None else [
            NVOSObject(code="40-0178-005113-П", name="Площадка", oktmo="40343000")],
        period=ReportPeriod(year=2025),
        pollutants=[
            Pollutant(name="Сера диоксид", code="0330", medium=Medium.AIR, mass_norm="0.8"),
            Pollutant(name="Азота диоксид", code="0301", medium=Medium.AIR, mass_norm="1.2"),
            Pollutant(name="Углерода оксид", code="0337", medium=Medium.AIR, mass_norm="3.5"),
            Pollutant(name="Взвешенные вещества", code="2902", medium=Medium.AIR, mass_norm="0.5"),
        ],
        extra=extra or {})


def test_tp2_air_sections(tmp_path):
    """Бланк состоит из ПЯТИ разделов — все листы обязаны присутствовать."""
    registry.load_all()
    rep = registry.get("2tp-air")(_ctx())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "air.xlsx"))
    assert wb.sheetnames == ["Титул", "Раздел 1", "Раздел 2 (специфич.)",
                             "Раздел 3 (источники)", "Раздел 4 (мероприятия)",
                             "Раздел 5 (группы источн.)"]
    s1 = wb["Раздел 1"]
    # строки 101-109 в столбце A (данные с 5-й строки: шапка + номера граф)
    assert [s1["A" + str(r)].value for r in range(5, 14)] == \
        [101, 102, 103, 104, 105, 106, 107, 108, 109]
    # графа 7 (I): 101 Всего = 102 твёрдые + 103 газообразные; SO2/CO/NOx разнесены
    assert s1["I5"].value == 6.0            # 101 всего
    assert s1["I6"].value == 0.5            # 102 твёрдые (взвешенные)
    assert s1["I7"].value == 5.5            # 103 = 0.8+3.5+1.2
    assert s1["I8"].value == 0.8            # 104 SO2
    assert s1["I10"].value == 1.2           # 106 NOx


def test_tp2_air_sect1_columns_match_blank(tmp_path):
    """Состав/порядок граф Раздела 1 — строго по бланку № 661: А, 1, Б, 2-7.

    Граф ПДВ/ВСВ (8-9) в бланке № 661 НЕТ — они были в старой форме;
    выдуманной графы «отходит от источников» тоже нет."""
    registry.load_all()
    rep = registry.get("2tp-air")(_ctx())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "air.xlsx"))
    s1 = wb["Раздел 1"]
    # служебная строка номеров граф — как в бланке, и ничего правее
    assert [s1[f"{c}4"].value for c in "ABCDEFGHI"] == \
        ["А", "1", "Б", "2", "3", "4", "5", "6", "7"]
    assert s1["J4"].value is None and s1["K4"].value is None
    assert s1["J5"].value is None and s1["K5"].value is None
    # коды ЗВ графы 1 — фиксированные по бланку (106 — 0012, не 0301)
    assert [s1["B" + str(r)].value for r in range(5, 14)] == \
        ["0001", "0002", "0004", "0330", "0337", "0012", "0401", "0006", "0005"]
    joined = " ".join(str(s1[f"{c}3"].value) for c in "ABCDEFGHI").lower()
    assert "организованных" in joined       # гр.3 бланка
    assert "пдв" not in joined and "всв" not in joined   # граф 8-9 в № 661 нет
    assert "очист" in joined                # графы потока очистки
    assert "отходит" not in joined          # такой графы в бланке нет
    assert "лимит" not in joined and "сверх" not in joined  # не декларация
    # формулировки графы Б — дословно по бланку
    labels = [s1["C" + str(r)].value for r in range(5, 14)]
    assert labels[5] == "оксиды азота (в пересчете на NO2)"
    assert labels[7] == "летучие органические соединения (ЛОС)"
    assert labels[1] == "в том числе: твердые"
    # гр.3 без ввода — прочерк (программа не выдумывает ноль)
    assert s1["E5"].value == "-"            # 101, гр.3
    # предупреждение о незаполненной графе 3 — в validate()
    assert any("графа 3" in i.message for i in rep.validate())


def test_tp2_air_org_sources(tmp_path):
    """Гр.3 берётся из extra и попадает в xlsx и XML; «limits» Раздела 1 больше нет."""
    registry.load_all()
    extra = {"tp2_air": {
        "cleaning": {"104": {"g7": "0.6"}},
        "limits": {"104": {"pdv": 1.5, "vsv": "-"}},   # устаревший ключ — игнорируется
    }}
    rep = registry.get("2tp-air")(_ctx(extra))
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "air.xlsx"))
    s1 = wb["Раздел 1"]
    assert s1["E8"].value == 0.6            # 104, гр.3 из cleaning.g7
    assert s1["E7"].value == 0.6            # 103 = сумма заполненных 104-109
    assert s1["E5"].value == 0.6            # 101 = 102 + 103
    assert s1["J8"].value is None           # ПДВ в Раздел 1 не печатается
    root = etree.parse(str(rep.render_xml(tmp_path / "air.xml"))).getroot()
    row104 = root.xpath("//Раздел1/Строка[@код='104']")[0]
    assert row104.findtext("ОтОрганизованных") == "0.600"
    assert row104.findtext("КодЗВ") == "0330"
    assert root.xpath("//Раздел1/Строка[@код='106']")[0].findtext("КодЗВ") == "0012"
    # гр.3 не может превышать гр.2 — ошибка валидации
    bad = registry.get("2tp-air")(_ctx({"tp2_air": {
        "cleaning": {"104": {"g7": "99"}}}}))
    assert any(i.level == "error" and "графа 3" in i.message
               for i in bad.validate())


def test_tp2_air_sect2_composition(tmp_path):
    """Раздел 2: без SO2/CO/NOx (они в 104-106); вещества Перечня — своими
    строками 201…; остальное — одной строкой 8888 суммарно; графа А есть."""
    registry.load_all()
    ctx = _ctx()
    ctx.pollutants += [
        Pollutant(name="Бенз/а/пирен", code="703", medium=Medium.AIR, mass_norm="0.001"),
        Pollutant(name="Метан", code="0410", medium=Medium.AIR, mass_norm="2.0"),
        Pollutant(name="Керосин", code="2732", medium=Medium.AIR, mass_norm="0.3"),
    ]
    rep = registry.get("2tp-air")(ctx)
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "air.xlsx"))
    s2 = wb["Раздел 2 (специфич.)"]
    assert [s2[f"{c}4"].value for c in "ABCD"] == ["А", "1", "Б", "2"]
    rows = [(s2[f"A{r}"].value, s2[f"B{r}"].value, s2[f"D{r}"].value)
            for r in range(5, s2.max_row + 1)]
    codes = [c for _, c, _ in rows]
    for excluded in ("0330", "0337", "0301", "0012"):
        assert excluded not in codes
    assert rows[0] == (201, "0410", 2.0)            # Перечень — отдельной строкой
    assert rows[1] == (202, "0703", 0.001)          # код нормализован до 4 цифр
    assert rows[2][0] == 203 and rows[2][1] == "8888"
    assert rows[2][2] == 0.8                         # 2902 (0.5) + 2732 (0.3)
    assert s2["C7"].value == "Другие специфические вещества"
    assert len(rows) == 3
    root = etree.parse(str(rep.render_xml(tmp_path / "air.xml"))).getroot()
    xml_codes = [x.get("код") for x in root.xpath("//Раздел2/Вещество")]
    assert xml_codes == ["0410", "0703", "8888"]
    assert root.xpath("//Раздел2/Вещество[@код='8888']")[0].get("строка") == "203"
    assert any("8888" in i.message for i in rep.validate())


def test_tp2_air_onv_in_headers(tmp_path):
    """Код ОНВ — в шапке каждого раздела (в Разделе 1 ещё ОКТМО/ОКВЭД2 ОНВ)
    и в XML; несколько ОНВ — предупреждение, а не молчаливое объединение."""
    registry.load_all()
    rep = registry.get("2tp-air")(_ctx())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "air.xlsx"))
    for name in wb.sheetnames[1:]:
        assert "Код ОНВ: 40-0178-005113-П" in str(wb[name]["A2"].value)
    h1 = wb["Раздел 1"]["A2"].value
    assert "Код ОКТМО ОНВ: 40343000" in h1 and "Код ОКВЭД2 ОНВ: 30.20.9" in h1
    assert "ОКТМО" not in str(wb["Раздел 2 (специфич.)"]["A2"].value)
    t = wb["Титул"]
    vals = {t[f"A{r}"].value: t[f"B{r}"].value for r in range(1, 12)}
    assert vals["Почтовый адрес"] == "190000, Санкт-Петербург, ул. Тестовая, 1"
    assert "от 08.11.2018 N 661" in " ".join(str(k) for k in vals)
    assert vals["Код формы по ОКУД"] == "0609012"
    root = etree.parse(str(rep.render_xml(tmp_path / "air.xml"))).getroot()
    assert root.findtext("ОНВ/КодОНВ") == "40-0178-005113-П"
    assert root.find("Раздел3").get("КодОНВ") == "40-0178-005113-П"
    assert not any(i.field == "ОНВ" for i in rep.validate())
    # без объектов — предупреждение; два объекта — предупреждение
    none = registry.get("2tp-air")(_ctx(objects=[]))
    assert any("код ОНВ" in i.message for i in none.validate())
    two = registry.get("2tp-air")(_ctx(objects=[
        NVOSObject(code="A", oktmo="1"), NVOSObject(code="B", oktmo="2")]))
    assert any("2 объектов" in i.message for i in two.validate())


def test_tp2_air_sect3_rows_302_303(tmp_path):
    """Раздел 3: шапка А/Б/1-4, строки 302/303 из sources['pdv'/'vsv'] — в xlsx и XML."""
    registry.load_all()
    extra = {"tp2_air": {"sources": {
        "total": 7, "organized": 6, "allowed": 7.042,
        "pdv": {"total": 7, "organized": 6},
    }}}
    rep = registry.get("2tp-air")(_ctx(extra))
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "air.xlsx"))
    s3 = wb["Раздел 3 (источники)"]
    assert [s3[f"{c}4"].value for c in "ABCDEF"] == ["А", "Б", "1", "2", "3", "4"]
    assert [s3[f"A{r}"].value for r in (5, 6, 7)] == [301, 302, 303]
    assert s3["B5"].value == "Всего"
    assert "(ПДВ)" in s3["B6"].value and "(ВСВ)" in s3["B7"].value
    assert (s3["C5"].value, s3["D5"].value, s3["E5"].value, s3["F5"].value) == \
        (7.0, 6.0, 7.042, 6.0)                      # факт 301 = сумма гр.7
    assert (s3["C6"].value, s3["D6"].value, s3["E6"].value, s3["F6"].value) == \
        (7.0, 6.0, "-", "-")                        # 302 — что задано, прочее прочерк
    assert [s3[f"{c}7"].value for c in "CDEF"] == ["-"] * 4   # 303 не выдумана
    root = etree.parse(str(rep.render_xml(tmp_path / "air.xml"))).getroot()
    r302 = root.xpath("//Раздел3/Строка[@код='302']")[0]
    assert r302.findtext("ВсегоИсточников") == "7"
    assert r302.findtext("РазрешённыйВыброс") == "-"
    assert root.xpath("//Раздел3/Строка[@код='303']")
    # без pdv/vsv — предупреждение про 302/303
    plain = registry.get("2tp-air")(_ctx({"tp2_air": {"sources": {"total": 7}}}))
    assert any("302/303" in i.message for i in plain.validate())


def test_tp2_air_sections_4_5_content(tmp_path):
    """Разделы 4 и 5 — строки 401-405 и 501-505 по бланку, XML — все разделы."""
    registry.load_all()
    extra = {"tp2_air": {
        "measures": [{"production": "Котельная", "name": "Замена горелок",
                      "group": 3, "done": 1, "spent_year": 100.0,
                      "cut_expected": 0.5, "cut_fact": 0.4}],
        "groups": {"502": {"fuel": 0.7, "tech": 0.1}},
    }}
    rep = registry.get("2tp-air")(_ctx(extra))
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "air.xlsx"))
    s4 = wb["Раздел 4 (мероприятия)"]
    assert [s4["A" + str(r)].value for r in range(5, 10)] == \
        [401, 402, 403, 404, 405]           # пять строк всегда, как в бланке
    assert s4["C5"].value == "Замена горелок"
    assert s4["B6"].value is None           # пустая графа не выдумывается
    s5 = wb["Раздел 5 (группы источн.)"]
    assert [s5["A" + str(r)].value for r in range(5, 10)] == \
        [501, 502, 503, 504, 505]
    assert [s5["B" + str(r)].value for r in range(5, 10)] == \
        ["0002", "0330", "0337", "0012", "0007"]   # коды ЗВ по бланку (504 — 0012)
    assert s5["D6"].value == 0.7            # 502 гр.3 (сжигание топлива)
    assert s5["D5"].value == "-"            # без данных — прочерк
    root = etree.parse(str(rep.render_xml(tmp_path / "air.xml"))).getroot()
    assert [c.tag for c in root if c.tag.startswith("Раздел")] == \
        ["Раздел1", "Раздел2", "Раздел3", "Раздел4", "Раздел5"]
    assert root.xpath("//Раздел4/Мероприятие[@код='401']")
    assert root.xpath("//Раздел5/Строка[@код='502']")[0] \
        .findtext("ОтСжиганияТоплива") == "0.700"
    assert root.xpath("//Раздел5/Строка[@код='504']")[0].findtext("КодЗВ") == "0012"


def test_tp2_air_sum_groups_excluded(tmp_path):
    """Группы суммации (6xxx) не попадают в форму — их массы уже в веществах.

    Страховка от отката фильтра: включение группы «301+330» рядом с самими
    301 и 330 удвоило бы выброс в Разделах 1 и 2."""
    registry.load_all()
    ctx = _ctx()
    ctx.pollutants.append(Pollutant(
        name="Группа суммации: азота диоксид, серы диоксид", code="6204",
        medium=Medium.AIR, mass_norm="2.0"))
    rep = registry.get("2tp-air")(ctx)
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "air.xlsx"))
    s1 = wb["Раздел 1"]
    assert s1["I5"].value == 6.0            # 101 всего — без массы группы
    s2 = wb["Раздел 2 (специфич.)"]
    codes = [s2["B" + str(r)].value for r in range(5, s2.max_row + 1)]
    assert "6204" not in codes
    assert s2["D5"].value == 0.5            # 8888 = только взвешенные, группа не добавлена
    assert any("групп" in i.message for i in rep.validate())


def test_tp2_air_cleaning_control_ratio():
    """Контрольное соотношение Указаний: гр.7 = гр.4 − гр.5 + гр.2."""
    registry.load_all()
    # согласованные данные очистки: 0.8 = 0.3 − 0.2 + 0.7 — предупреждения нет
    ok = registry.get("2tp-air")(_ctx({"tp2_air": {"cleaning": {
        "104": {"g2": "0.7", "g3": "0.3", "g4": "0.2", "g6": "0.8"}}}}))
    assert not any("контрольное соотношение" in i.message for i in ok.validate())
    # рассогласованные (уловлено больше, чем поступило в форму) — предупреждение
    bad = registry.get("2tp-air")(_ctx({"tp2_air": {"cleaning": {
        "104": {"g3": "0.3", "g4": "0.2"}}}}))
    assert any("контрольное соотношение" in i.message and "104" in i.message
               for i in bad.validate())


def test_tp2_air_row_classification():
    """Коды без ведущего нуля и «именные» SO2/CO/NOx/керосин не уезжают в 109."""
    from ecodoc.reports.tp2_air.report import _air_row

    def mk(code, name):
        return Pollutant(name=name, code=code, medium=Medium.AIR, mass_norm="1")

    assert _air_row(mk("301", "Азота диоксид")) == 106      # без ведущего нуля
    assert _air_row(mk("330", "Сера диоксид")) == 104
    assert _air_row(mk("337", "Углерод оксид")) == 105
    assert _air_row(mk("3003", "Диоксид серы")) == 104      # старые коды ПДВ
    assert _air_row(mk("3332", "Оксид углерода")) == 105
    assert _air_row(mk("3004", "Диоксид азота")) == 106
    assert _air_row(mk("2732", "Керосин")) == 107           # код в справочнике
    assert _air_row(mk("9999", "Керосин ГОСТ")) == 107      # и по названию
    assert _air_row(mk("9998", "Алканы C12-C19")) == 107
    assert _air_row(mk("9997", "Уайт-спирит")) == 107


def test_tp2_air_specific_list_is_official():
    """Перечень приложения к № 661 — 97 позиций (нумерация в приказе пропускает № 29), ключевые коды на месте."""
    from ecodoc.reports.tp2_air.report import SECT2_SPECIFIC
    assert len(SECT2_SPECIFIC) == 97
    for c in ("0703", "0322", "0410", "0303", "0328", "1325", "2908", "2926"):
        assert c in SECT2_SPECIFIC
    for c in ("0330", "0337", "0301", "0304", "0012", "8888"):
        assert c not in SECT2_SPECIFIC
