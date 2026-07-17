"""Сводная по отходам («Справки-2025») + паспорта отходов + новые скоупы."""
from decimal import Decimal

from ecodoc.core.models import ReportContext, WasteAct
from ecodoc.core.waste_summary import _fmt_fkko, build_rows, build_xlsx


def _ctx():
    ctx = ReportContext()
    ctx.period.year = 2025
    ctx.waste_acts = [
        WasteAct(name="Лампы ртутные", fkko_code="47110101521", hazard_class=1,
                 mass=Decimal("0.05"), density=Decimal("0.55"),
                 operation="обезвреживание", carrier="Меркурий",
                 receiver="Меркурий", date="15.02.2025"),
        WasteAct(name="Лампы ртутные", fkko_code="4 71 101 01 52 1",
                 hazard_class=1, mass=Decimal("0.03"),
                 operation="обезвреживание", date="2025-07-10"),
        WasteAct(name="Смет", fkko_code="73339001714", hazard_class=4,
                 mass=Decimal("1.9"), volume_m3=Decimal("2.4"),
                 operation="размещение", receiver="Полигон", date=""),
        WasteAct(name="Прошлогодний", fkko_code="73339001714", hazard_class=4,
                 mass=Decimal("9"), date="01.03.2024"),
    ]
    return ctx


def test_build_rows_months_quarters_year():
    rows = build_rows(_ctx(), 2025)
    assert len(rows) == 2
    lamp = next(r for r in rows if r["fkko"] == "47110101521")
    # два акта слились: февраль + июль (ISO-дата), год = сумма
    assert lamp["mass_m"] == {2: Decimal("0.05"), 7: Decimal("0.03")}
    assert lamp["mass_q"][1] == Decimal("0.05") and lamp["mass_q"][3] == Decimal("0.03")
    assert lamp["mass_y"] == Decimal("0.08")
    # объём из плотности: 0.05/0.55
    assert abs(lamp["vol_m"][2] - Decimal("0.05") / Decimal("0.55")) < Decimal("1e-9")
    smet = next(r for r in rows if r["fkko"] == "73339001714")
    # акт без даты → «без даты» и в итог года; акт 2024 отфильтрован
    assert smet["mass_nodate"] == Decimal("1.9")
    assert smet["mass_y"] == Decimal("1.9")


def test_fkko_formatted_like_catalog():
    assert _fmt_fkko("47110101521") == "4 71 101 01 52 1"


def test_xlsx_three_sheets(tmp_path):
    p = build_xlsx(_ctx(), tmp_path / "s.xlsx")
    import openpyxl
    wb = openpyxl.load_workbook(p)
    assert wb.sheetnames == ["По месяцам", "По кварталам", "За год"]
    ws = wb["По месяцам"]
    head = [c.value for c in ws[3]]
    assert head[:4] == ["Наименование отхода", "Код отхода (по ФККО)",
                        "Класс опасности", "Плотность, т/м³"]
    assert "Вид обращения" in head and "Приёмщик/полигон" in head
    assert ws.cell(row=4, column=2).value == "4 71 101 01 52 1"


def test_passports_scope_only_touches_passports():
    """scope='passports': паспорт идёт в extra.waste_passports, реквизиты и
    акты из того же документа НЕ принимаются."""
    from ecodoc.ai.analyzer import ExtractionReport, _merge_passports
    ctx = ReportContext()
    rep = ExtractionReport()
    data = {"waste_passports": [
        {"fkko": "4 71 101 01 52 1", "name": "Лампы ртутные", "hazard_class": 1,
         "components": [{"name": "стекло", "percent": "92"},
                        {"name": "ртуть", "percent": "0.02"}]},
        {"fkko": "47110101521", "name": "Лампы (дубль)"},   # дедуп по ФККО
    ]}
    _merge_passports(ctx, data, "паспорт.pdf", rep)
    store = ctx.extra["waste_passports"]
    assert len(store) == 1
    assert store[0]["fkko"] == "47110101521"
    assert len(store[0]["components"]) == 2
    assert any("паспорт отхода" in a.field for a in rep.accepted)


def test_act_carrier_license_roundtrip(tmp_path):
    from ecodoc.core.serialize import from_json, to_json
    ctx = ReportContext()
    ctx.waste_acts = [WasteAct(name="Смет", fkko_code="73339001714",
                               mass=Decimal("1"), carrier="ООО Т",
                               carrier_license="Л-123", license="Л-456")]
    p = tmp_path / "c.json"
    to_json(ctx, p)
    a = from_json(p).waste_acts[0]
    assert a.carrier_license == "Л-123" and a.license == "Л-456"
