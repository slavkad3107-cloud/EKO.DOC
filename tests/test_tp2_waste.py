"""Форма № 2-ТП (отходы) по приказу Росстата от 06.11.2025 № 614 — печать и
проверки. Эталон формулировок — принятый отчёт ИП Миних за 2025
(печать ЛКПП); порядок строк и баланс — Указания к № 614 (п. 7, 8, 13, 14).
XML-тест (конверт Модуля) — в test_waste_forms.py."""
import openpyxl

from ecodoc.core import registry
from ecodoc.core.models import WasteAct, WasteFlow
from tests.test_waste_forms import _ctx


def _print(ctx, tmp_path):
    registry.load_all()
    rep = registry.get("2tp-waste")(ctx)
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "2tp.xlsx"))
    return rep, wb


def test_tp2_waste_print_pages(tmp_path):
    rep, wb = _print(_ctx(), tmp_path)
    assert wb.sheetnames == ["стр.1", "стр.2", "стр.3"]
    assert wb["стр.1"]["A19"].value == "0609013"       # код формы по ОКУД
    # графы А,Б,В,Г (строка номеров 5) + 29 граф (Приказ № 614)
    assert wb["стр.2"]["A5"].value == "А"
    from openpyxl.utils import get_column_letter
    assert wb["стр.2"][f"{get_column_letter(4 + 29)}5"].value == 29  # графа 29


def test_tp2_waste_section1_no_totals_sorted_by_class(tmp_path):
    """Раздел I — только виды отходов (строк «ВСЕГО»/«Всего по классу» в
    бланке нет — эталон МИНИХ: строка 1 = первый отход), в последовательности
    с I по V класс (п. 7 Указаний к № 614), внутри класса — по коду ФККО,
    нумерация с 1 (п. 13)."""
    ctx = _ctx()
    ctx.wastes = [
        WasteFlow(fkko_code="73310001724", name="Мусор офисный", hazard_class=4,
                  generated="42.6", transferred="42.6"),
        WasteFlow(fkko_code="47110101521", name="Лампы ртутные", hazard_class=1,
                  generated="0.115", transferred="0.115", transferred_neutral="0.115"),
        WasteFlow(fkko_code="92130201523", name="Аккумуляторы", hazard_class=2,
                  generated="0.2", transferred="0.2", transferred_util="0.2"),
        WasteFlow(fkko_code="73310002725", name="Смёт", hazard_class=5,
                  generated="1.0", transferred="1.0", transferred_burial="1.0"),
        WasteFlow(fkko_code="40612001313", name="Масло отработанное", hazard_class=3,
                  generated="0.5", transferred="0.5", transferred_util="0.5"),
    ]
    rep, wb = _print(ctx, tmp_path)
    s2 = wb["стр.2"]
    rows = [(s2[f"A{r}"].value, s2[f"B{r}"].value, s2[f"D{r}"].value)
            for r in range(6, 11)]
    assert [r[0] for r in rows] == [1, 2, 3, 4, 5]
    assert [r[2] for r in rows] == [1, 2, 3, 4, 5]
    assert rows[0][1] == "Лампы ртутные" and rows[4][1] == "Смёт"
    texts = [s2[f"B{r}"].value for r in range(1, 20)]
    assert not any(t and str(t).startswith("ВСЕГО") for t in texts)
    assert not any(t and "классу опасности" in str(t) for t in texts)
    assert s2["B11"].value is None                      # после 5 отходов пусто
    # тот же порядок в XML
    xml = rep.render_xml(tmp_path / "2tp.xml").read_text(encoding="utf-8")
    assert xml.index("47110101521") < xml.index("92130201523") < \
        xml.index("40612001313") < xml.index("73310001724") < xml.index("73310002725")


def test_tp2_waste_section1_header_verbatim(tmp_path):
    """Шапка Раздела I — дословно по бланку № 614 (эталон МИНИХ, табл. 8/10):
    многоуровневая, служебные графы полные, группы граф объединены."""
    rep, wb = _print(_ctx(), tmp_path)
    s2 = wb["стр.2"]
    assert s2["A2"].value == "N строки"
    assert s2["C2"].value == ("Код отхода по федеральному классификационному "
                              "каталогу отходов")
    assert s2["D2"].value == "Класс опасности отхода"
    assert s2["E2"].value == "Наличие отходов на начало отчетного года"
    # гр.3–5: «Поступление … из других хозяйствующих субъектов» → всего /
    # из графы 3 → из других субъектов РФ / по импорту из других государств
    assert s2["G2"].value == "Поступление отходов из других хозяйствующих субъектов"
    assert s2["G3"].value == "всего" and s2["H3"].value == "из графы 3"
    assert s2["H4"].value == "из других субъектов РФ"
    assert s2["I4"].value == "по импорту из других государств"
    merged = {str(m) for m in s2.merged_cells.ranges}
    assert "G2:I2" in merged and "G3:G4" in merged and "H3:I3" in merged
    # гр.15–24: «Передача отходов (за исключением ТКО) другим хозяйствующим
    # субъектам» → для обработки → всего передано для обработки / из них …
    assert s2["S2"].value == ("Передача отходов (за исключением ТКО) другим "
                              "хозяйствующим субъектам")
    assert "S2:AB2" in merged
    assert s2["S3"].value == "для обработки" and "S3:T3" in merged
    assert s2["S4"].value == "всего передано для обработки"
    assert s2["T4"].value == "из них в другие субъекты РФ"
    # гр.27/28: «Размещение отходов на эксплуатируемых объектах за отчетный
    # год» → хранение / захоронение
    assert s2["AE2"].value == ("Размещение отходов на эксплуатируемых объектах "
                               "за отчетный год")
    assert s2["AE3"].value == "хранение" and s2["AF3"].value == "захоронение"
    assert s2["AG2"].value == "Наличие отходов на конец отчетного года"
    assert [s2.cell(row=5, column=5 + i).value for i in range(29)] == \
        list(range(1, 30))


def test_tp2_waste_placement_storage_vs_burial(tmp_path):
    """Гр.27 (хранение) и гр.28 (захоронение) — раздельно из placed_storage /
    placed_burial (п. 13 Указаний); XML TP2_RAZM_STOR — из данных. Старые
    базы (только placed_norm/over) — как раньше в захоронение + предупреждение."""
    ctx = _ctx()
    ctx.wastes = [
        WasteFlow(fkko_code="73310001724", name="Мусор", hazard_class=4,
                  generated="10", placed_storage="3", placed_burial="7"),
        WasteFlow(fkko_code="81290101205", name="Грунт", hazard_class=5,
                  generated="20", placed_norm="20"),        # старая база
    ]
    rep, wb = _print(ctx, tmp_path)
    s2 = wb["стр.2"]
    assert (s2["AE6"].value, s2["AF6"].value) == (3.0, 7.0)     # гр.27 / гр.28
    assert (s2["AE7"].value, s2["AF7"].value) == (0.0, 20.0)    # fallback
    xml = rep.render_xml(tmp_path / "2tp.xml").read_text(encoding="utf-8")
    assert "<TP2_RAZM_STOR>3.0</TP2_RAZM_STOR>" in xml
    assert "<TP2_RAZM>7.0</TP2_RAZM>" in xml
    issues = rep.validate()
    fb = [i for i in issues if i.field == "размещение/81290101205"]
    assert fb and fb[0].level == "warning" and "захоронение" in fb[0].message
    assert not [i for i in issues if i.field == "размещение/73310001724"]
    # баланс сходится: 10 − 3 − 7 = 0, 20 − 20 = 0 → без замечаний по балансу
    assert not [i for i in issues if i.field.startswith("баланс/")]


def test_tp2_waste_validate_transfer_purpose_and_balance():
    """validate(): «передано всего» должно раскладываться по гр.14/15–23 —
    иначе error «не указано назначение передачи по акту»; баланс гр.29
    вычитает обработанные (w.processed)."""
    ctx = _ctx()
    ctx.wastes = [
        # передано 0.5, назначение не указано → error
        WasteFlow(fkko_code="40612001313", name="Масло", hazard_class=3,
                  generated="0.5", transferred="0.5"),
        # ТКО без получателей — всё регоператору (гр.14), ошибки нет
        WasteFlow(fkko_code="73310001724", name="Мусор", hazard_class=4,
                  generated="42.6", transferred="42.6"),
        # обработано 2 из 10 — баланс с учётом processed сходится при остатке 8
        WasteFlow(fkko_code="45711101725", name="Картон", hazard_class=5,
                  generated="10", processed="2", accumulated_end="8"),
    ]
    registry.load_all()
    rep = registry.get("2tp-waste")(ctx)
    issues = rep.validate()
    tr = [i for i in issues if i.field == "передача/40612001313"]
    assert tr and tr[0].level == "error"
    assert "не указано назначение передачи по акту" in tr[0].message
    assert not [i for i in issues if i.field == "передача/73310001724"]
    assert not [i for i in issues if i.field == "баланс/45711101725"]


def test_tp2_waste_g14_only_for_regional_operator(tmp_path):
    """Графа 14 — только ТКО, переданные региональному оператору (по
    получателю из акта); ТКО лицензиату на утилизацию — гр.17 (п. 13)."""
    from ecodoc.core.waste_agg import apply_acts
    ctx = _ctx()
    ctx.wastes = []
    ctx.waste_acts = [
        WasteAct(name="Мусор", fkko_code="73310001724", hazard_class=4, mass="30",
                 operation="передача ТКО", receiver="АО «Невский экологический "
                 "оператор» (региональный оператор)"),
        WasteAct(name="Мусор", fkko_code="73310001724", hazard_class=4, mass="12",
                 operation="утилизация", receiver="ООО «Втор-Ресурс»"),
    ]
    apply_acts(ctx)
    rep, wb = _print(ctx, tmp_path)
    s2 = wb["стр.2"]
    assert s2["R6"].value == 30.0        # гр.14 — региональному оператору
    assert s2["U6"].value == 12.0        # гр.17 — для утилизации
    assert not [i for i in rep.validate() if i.field.startswith("передача/")]
    # пояснение под Разделом II — по п. 14 Указаний (не только регоператоры)
    s3 = wb["стр.3"]
    assert "п. 14 Указаний" in s3["A19"].value
    assert "операторы, не передающие" in s3["A19"].value


def test_tp2_waste_unknown_operation_not_lost(tmp_path):
    """Акт без распознанного назначения передачи: масса остаётся в
    transferred (не теряется), агрегация даёт предупреждение, а validate()
    2-ТП — error «не указано назначение передачи по акту»."""
    from ecodoc.core.waste_agg import apply_acts
    ctx = _ctx()
    ctx.wastes = []
    ctx.waste_acts = [
        WasteAct(name="Масло", fkko_code="40612001313", hazard_class=3, mass="0.5",
                 operation="передача", receiver="ООО «Нефтесервис»",
                 date="10.04.2024"),
    ]
    apply_acts(ctx)
    w = ctx.wastes[0]
    assert float(w.transferred) == 0.5 and float(w.transferred_util) == 0
    warns = ctx.extra["waste_agg_warnings"]
    assert len(warns) == 1 and "не указано назначение передачи" in warns[0]
    rep, wb = _print(ctx, tmp_path)
    issues = rep.validate()
    assert [i for i in issues if i.field == "акты" and i.level == "warning"]
    err = [i for i in issues if i.field == "передача/40612001313"]
    assert err and err[0].level == "error"
    # в печати — гр.15–23 нули, гр.14 (не ТКО) тоже
    s2 = wb["стр.2"]
    assert all(s2.cell(row=6, column=c).value == 0.0 for c in range(18, 28))


def test_tp2_waste_fkko_spaced_in_print(tmp_path):
    """Графа В печатается в каноническом виде ФККО с пробелами (как в печати
    ЛКПП: «7 33 100 01 72 4»), а в XML код остаётся без пробелов."""
    rep, wb = _print(_ctx(), tmp_path)
    s2 = wb["стр.2"]
    # строки с 6 — позиции отходов (без итоговых строк), I класс первым
    assert s2["C6"].value == "4 71 101 01 52 1"
    assert s2["C7"].value == "7 33 100 01 72 4"
    xml = rep.render_xml(tmp_path / "2tp.xml").read_text(encoding="utf-8")
    assert "<WST_CODE>73310001724</WST_CODE>" in xml   # XML — без пробелов


def test_tp2_waste_section2_structure(tmp_path):
    """Раздел II — по бланку № 614: официальный заголовок, многоуровневая
    шапка А|Б|В|Г и 29 граф тремя подтаблицами (1–9 / 10–17 / 18–29); шапка
    печатается и у нерегоператора (с пояснением), Раздел III — фикс. строки
    11–31, подпись — по бланку."""
    rep, wb = _print(_ctx(), tmp_path)      # не регоператор
    s3 = wb["стр.3"]
    assert s3["A1"].value.startswith(
        "Раздел II. Сведения об образовании, обработке, утилизации, "
        "обезвреживании, размещении отходов производства и потребления, "
        "представляемые региональными операторами")
    assert s3["A1"].value.rstrip().endswith("тонна")
    # подтаблица 1: шапка 3–5 (3 уровня), буквенная строка 6: А|Б|В|Г + гр.1–9
    assert [s3[f"{c}6"].value for c in "ABCD"] == ["А", "Б", "В", "Г"]
    assert [s3.cell(row=6, column=5 + i).value for i in range(9)] == \
        list(range(1, 10))
    assert s3["G3"].value.startswith("Поступление ТКО к региональному оператору")
    assert s3["G4"].value == "всего ТКО" and s3["H4"].value == "из графы 3"
    assert s3["H5"].value == "ТКО, образованных в жилых помещениях в субъекте РФ"
    # подтаблица 2 (гр.10–17) и 3 (гр.18–29, 5 уровней) — после «продолжение»
    assert s3["A7"].value == "продолжение раздела II"
    assert [s3.cell(row=11, column=5 + i).value for i in range(8)] == \
        list(range(10, 18))
    assert s3["A12"].value == "продолжение раздела II"
    assert s3["G16"].value == "на энергетическую утилизацию"
    assert s3["G17"].value == "всего ТКО"
    assert [s3.cell(row=18, column=5 + i).value for i in range(12)] == \
        list(range(18, 30))
    # структура не подменяется текстом — пояснение идёт ПОД шапкой
    assert s3["A19"].value.startswith("— раздел не заполняется")
    # Раздел III: фиксированные строки 11–31 с прочерками
    assert s3["A21"].value.startswith("Раздел III")
    assert (s3["A22"].value, s3["B22"].value, s3["C22"].value) == \
        ("N строки", "Наименование показателя", "Фактически")
    assert s3["A23"].value == 11 and s3["C23"].value == "-"
    assert s3["A43"].value == 31 and s3["C43"].value == "-"
    # подпись — по бланку № 614 (таблица 16 эталона)
    assert s3["A46"].value.startswith(
        "Должностное лицо, ответственное за предоставление первичных "
        "статистических данных (лицо, уполномоченное")
    assert (s3["A49"].value, s3["B49"].value, s3["C49"].value) == \
        ("(должность)", "(Ф.И.О.)", "(подпись)")
    assert s3["C51"].value == "(дата составления документа)"


def test_tp2_waste_wording_matches_lkpp_sample(tmp_path):
    """Формулировки — дословно по печатной форме ЛКПП (приказ № 614):
    титул с тремя стандартными надписями, блоком «Предоставляют / Сроки /
    Форма / Приказ / Годовая» и полным перечнем респондентов; Раздел I —
    полный официальный заголовок; Раздел III — «из них ТКО»."""
    rep, wb = _print(_ctx(), tmp_path)
    s1 = wb["стр.1"]
    assert s1["A1"].value == "ФЕДЕРАЛЬНОЕ СТАТИСТИЧЕСКОЕ НАБЛЮДЕНИЕ"
    assert s1["A2"].value == "КОНФИДЕНЦИАЛЬНОСТЬ ГАРАНТИРУЕТСЯ ПОЛУЧАТЕЛЕМ ИНФОРМАЦИИ"
    assert s1["A3"].value.startswith("Нарушение порядка предоставления первичных")
    assert "152-ФЗ" in s1["A4"].value
    assert "ТРАНСПОРТИРОВАНИИ" not in s1["A6"].value
    assert s1["A6"].value.endswith("за 2024 г.")
    assert s1["A9"].value == "Предоставляют:"
    assert s1["D9"].value == "Сроки предоставления"
    assert s1["E9"].value == "Форма N 2-ТП (отходы)"
    assert s1["A10"].value.startswith("юридические лица, физические лица")
    assert "операторы по обращению с твердыми коммунальными отходами:" in s1["A10"].value
    assert s1["D11"].value == "1 февраля после отчетного периода"
    assert s1["D12"].value == "15 марта после отчетного периода"
    assert "от 06.11.2025 N 614" in s1["E10"].value
    assert s1["E12"].value == "Годовая"
    assert s1["A17"].value == "Код Формы по ОКУД"
    assert "идентификационный номер" in s1["B17"].value
    assert wb["стр.2"]["A1"].value == (
        "Раздел I. Сведения об образовании, обработке, утилизации, "
        "обезвреживании, размещении отходов производства и потребления; "
        "сведения об образовании и передаче твердых коммунальных отходов "
        "региональному оператору, тонна")
    s3 = wb["стр.3"]
    assert s3["B24"].value == "из них ТКО, ед"     # строка 12 Раздела III
    assert s3["B42"].value == "из них ТКО, га"     # строка 30 Раздела III


def test_tp2_waste_section2_operator_data(tmp_path):
    """Данные регоператора ложатся в свои графы: ключи g1..g29 плюс
    обратная совместимость (received→гр.3, processed→гр.10, placed→гр.27);
    графы без источника — прочерк."""
    ctx = _ctx()
    ctx.extra["tko_operators"] = [
        {"name": "Отходы коммунальные", "fkko": "73111001724",
         "hazard_class": 4, "g1": 10.5, "received": 100.0,
         "processed": 50.0, "placed": 20.0}]
    rep, wb = _print(ctx, tmp_path)
    s3 = wb["стр.3"]
    # подтаблица 1: шапка 3–5, номера 6, данные в строке 7
    assert s3["B7"].value == "Отходы коммунальные"
    assert s3["C7"].value == "7 31 110 01 72 4"     # графа В — с пробелами
    assert s3["E7"].value == 10.5                    # гр.1 ← g1
    assert s3["G7"].value == 100.0                   # гр.3 ← received (legacy)
    assert s3["F7"].value == "-"                     # гр.2 — нет источника
    # подтаблица 2: «продолжение» 8, шапка 9–11, номера 12, данные 13
    assert s3["E13"].value == 50.0                   # гр.10 ← processed
    # подтаблица 3: «продолжение» 14, шапка 15–19, номера 20, данные 21;
    # гр.27 ← placed (колонка N), гр.29 — прочерк
    assert s3["N21"].value == 20.0
    assert s3["P21"].value == "-"
