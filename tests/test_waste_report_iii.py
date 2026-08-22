"""Справка о движении отходов за год (данные для раздела 4 отчёта ПЭК).

Отдельной формы «III кат./МСП» в законодательстве нет (п. 7 ст. 18 ФЗ-89,
приказ № 30 отменён с 01.01.2021) — документ должен честно называться
справкой, а листы повторять шапки табл. 4.2 (21 графа) и 4.3 (12 граф)
отчёта ПЭК по принятому эталону Otcet-o-PEK_6303183.docx.
"""
import openpyxl

from ecodoc.core import registry
from ecodoc.core.models import (NVOSObject, Organization, ReportContext,
                                ReportPeriod, WasteFlow)


def _ctx():
    return ReportContext(
        organization=Organization(name="ООО Т", inn="7801234564", ogrn="1157847008219",
                                  okpo="12345678", oktmo="40324000"),
        period=ReportPeriod(year=2025),
        objects=[NVOSObject(code="40-0178-001234-П", category="III", address="СПб",
                            oktmo="40908000")],
        wastes=[WasteFlow(fkko_code="73310001724", name="ТКО", hazard_class=4,
                          generated="5", transferred="5", transferred_burial="5")],
        extra={"waste_receivers": [
            {"fkko": "73310001724", "receiver": "ООО Полигон", "inn": "4703037467",
             "address": "ЛО, Всеволожский р-н", "license": "78 №00105 от 20.11.2015",
             "operation": "передача на захоронение", "mass": "5"}]})


def test_title_and_general_sheet(tmp_path):
    registry.load_all()
    cls = registry.get("waste-report-iii")
    # находка [5]: не «отчётность III кат./МСП», а справка для раздела 4 ПЭК
    assert "Справка о движении отходов" in cls.title and "раздела 4" in cls.title
    assert "МСП" not in cls.title
    rep = cls(_ctx())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "wr3.xlsx"))
    assert wb.sheetnames == ["Общие сведения", "Движение отходов", "Получатели"]
    general = [str(c.value) for c in wb["Общие сведения"]["B"] if c.value]
    assert any("40-0178-001234-П" in v for v in general)
    left = " ".join(str(c.value) for c in wb["Общие сведения"]["A"] if c.value)
    # документ честно говорит: отдельной подачи нет, сведения идут в отчёт ПЭК
    assert "Отдельной подачи у этого документа НЕТ" in left
    assert "№ 30 отменён" in left and "№ 173" in left and "ст. 18 ФЗ-89" in left
    # находка [2]: устаревшей ветки «МСП — региональный порядок» нет
    assert "региональному порядку" not in left and "МСП регионального надзора" not in left


def test_movement_sheet_is_table_4_2(tmp_path):
    # находка [3]: точная двухъярусная шапка табл. 4.2 отчёта ПЭК — 21 графа
    registry.load_all()
    rep = registry.get("waste-report-iii")(_ctx())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "wr3.xlsx"))
    ws = wb["Движение отходов"]
    assert ws["A1"].value.startswith("Таблица 4.2. Сведения об образовании")
    assert ws["A2"].value == "N строки"
    assert ws["B2"].value == "Наименование видов отходов"
    assert ws["C2"].value == ("Код по федеральному классификационному каталогу "
                              "отходов, далее - ФККО")
    assert ws["E2"].value == "Наличие отходов на начало года, тонн"
    assert (ws["E3"].value, ws["F3"].value) == ("Хранение", "Накопление")
    assert ws["K2"].value.startswith("Передано отходов другим")
    assert [ws[f"{c}3"].value for c in "KLMNOP"] == [
        "Всего", "для обработки", "для утилизации", "для обезвреживания",
        "для хранения", "для захоронения"]
    assert ws["Q2"].value == "Размещено отходов на эксплуатируемых объектах, тонн"
    assert ws["R3"].value.startswith("Хранение на собственных объектах")
    assert ws["S3"].value == "Захоронение на собственных ОРО"
    assert ws["T2"].value == "Наличие отходов на конец года, тонн"
    assert (ws["T3"].value, ws["U3"].value) == ("Хранение", "Накопление")
    # строка нумерации граф 1..21
    assert [ws.cell(row=4, column=i).value for i in range(1, 22)] == list(range(1, 22))
    # данные: N строки, наименование, ФККО, класс, … передано всего/захоронение
    assert (ws["A5"].value, ws["B5"].value, ws["C5"].value, ws["D5"].value) == \
        (1, "ТКО", "73310001724", 4)
    assert ws["G5"].value == 5 and ws["K5"].value == 5 and ws["P5"].value == 5
    assert ws["L5"].value == 0


def test_receivers_sheet_is_table_4_3(tmp_path):
    # находка [4]: лист «Получатели» = табл. 4.3 (12 граф, массы по цели)
    registry.load_all()
    rep = registry.get("waste-report-iii")(_ctx())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "wr3.xlsx"))
    ws = wb["Получатели"]
    assert ws["A1"].value.startswith("Таблица 4.3. Сведения о юридических лицах")
    assert ws["A2"].value == "Номер строки"
    assert ws["C2"].value == "Код отхода по ФККО"
    assert ws["D2"].value.startswith("Наименование, ИНН, адрес")
    assert ws["E2"].value == "Получено отходов, т"
    assert ws["F2"].value.startswith("Цель приема отходов")
    assert ws["H2"].value.startswith("Количество отходов, переданных")
    assert (ws["H3"].value, ws["I3"].value, ws["J3"].value) == (
        "Для обработки", "Для утилизации", "Для обезвреживания")
    assert ws["K3"].value == "Для размещения"
    assert (ws["K4"].value, ws["L4"].value) == ("хранение", "захоронение")
    assert ws["M2"].value is None  # графы «Лицензия»/«Всего» в форме нет
    assert [ws.cell(row=5, column=i).value for i in range(1, 13)] == list(range(1, 13))
    # данные: контрагент одной строкой, масса — в графе цели «захоронение»
    assert ws["B6"].value == "ТКО" and ws["C6"].value == "73310001724"
    assert ws["G6"].value == "ООО Полигон, 4703037467, ЛО, Всеволожский р-н"
    assert ws["L6"].value == 5 and ws["H6"].value == 0
    # лицензия в форме № 173 не печатается
    row6 = [str(c.value) for c in ws[6]]
    assert not any("00105" in v for v in row6)


def test_empty_requisites_print_dash_and_warn(tmp_path):
    # находка [6]: «Руководитель» без ФИО и пустые реквизиты — «—», не « / »
    registry.load_all()
    ctx = _ctx()
    ctx.organization.director_position = "Генеральный директор"
    ctx.organization.director_name = ""
    rep = registry.get("waste-report-iii")(ctx)
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "wr3.xlsx"))
    ws = wb["Общие сведения"]
    rows = {ws.cell(row=i, column=1).value: ws.cell(row=i, column=2).value
            for i in range(1, ws.max_row + 1)}
    assert rows["Руководитель"] == "—"
    assert rows["Телефон / e-mail"] == "— / —"
    assert rows["ОКВЭД / ОКТМО"] == "— / 40324000"
    issues = [i for i in rep.validate() if i.field == "руководитель"]
    assert len(issues) == 1 and issues[0].level == "warning"
    # с ФИО — печатается целиком, предупреждения нет
    ctx.organization.director_name = "Иванов И.И."
    rep = registry.get("waste-report-iii")(ctx)
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "wr3b.xlsx"))
    ws = wb["Общие сведения"]
    vals = {ws.cell(row=i, column=1).value: ws.cell(row=i, column=2).value
            for i in range(1, ws.max_row + 1)}
    assert vals["Руководитель"] == "Генеральный директор Иванов И.И."
    assert not [i for i in rep.validate() if i.field == "руководитель"]


def test_xml_is_internal_with_breakdowns(tmp_path):
    registry.load_all()
    rep = registry.get("waste-report-iii")(_ctx())
    xml = rep.render_xml(tmp_path / "wr3.xml").read_text(encoding="utf-8")
    assert "<СправкаДвижениеОтходов" in xml
    assert "<НаличиеНачалоХранение>" in xml
    assert "<НаличиеНачалоНакопление>" in xml
    assert "<НаличиеКонец>" in xml and "<НаличиеКонецНакопление>" in xml
    assert "<Захоронение>5</Захоронение>" in xml
    assert "<ОбъектНВОС" in xml


def _ctx_nakopl():
    """Позиция с остатками на начало года в ОБОИХ режимах: хранение 2 т +
    накопление 3 т + образовано 10 т = передано 10 т + на конец 5 т —
    данные внутренне согласованы, предупреждения баланса быть не должно."""
    ctx = _ctx()
    ctx.extra = {}
    ctx.wastes = [WasteFlow(fkko_code="73310001724", name="ТКО", hazard_class=4,
                            accumulated_start="2", accumulated_start_nakopl="3",
                            generated="10", transferred="10",
                            accumulated_end="5")]
    return ctx


def test_balance_counts_nakopl_start(tmp_path):
    # без учёта накопления на начало года validate() давал ложное
    # предупреждение на согласованных данных
    registry.load_all()
    rep = registry.get("waste-report-iii")(_ctx_nakopl())
    issues = [i for i in rep.validate() if i.field.startswith("баланс")]
    assert issues == []


def test_balance_counts_nakopl_end(tmp_path):
    # остаток на конец года в режиме «накопление» (гр. 21) — тоже часть баланса
    registry.load_all()
    ctx = _ctx_nakopl()
    ctx.wastes[0].accumulated_end = "1"
    ctx.wastes[0].accumulated_end_nakopl = "4"
    rep = registry.get("waste-report-iii")(ctx)
    assert [i for i in rep.validate() if i.field.startswith("баланс")] == []


def test_balance_still_warns_on_mismatch(tmp_path):
    # реальное расхождение баланса по-прежнему ловится (фикс не «заглушил» проверку)
    registry.load_all()
    ctx = _ctx_nakopl()
    ctx.wastes[0].accumulated_end = "7"  # баланс даёт 5, а заявлено 7
    rep = registry.get("waste-report-iii")(ctx)
    issues = [i for i in rep.validate() if i.field.startswith("баланс")]
    assert len(issues) == 1


def test_xlsx_and_xml_split_start_columns(tmp_path):
    # хранение и накопление на начало года печатаются раздельно и с верными
    # значениями — и в xlsx (графы 5/6 = E/F), и в xml
    registry.load_all()
    rep = registry.get("waste-report-iii")(_ctx_nakopl())
    wb = openpyxl.load_workbook(rep.render_print(tmp_path / "n.xlsx"))
    ws = wb["Движение отходов"]
    assert ws["E5"].value == 2 and ws["F5"].value == 3
    xml = rep.render_xml(tmp_path / "n.xml").read_text(encoding="utf-8")
    assert "<НаличиеНачалоХранение>2</НаличиеНачалоХранение>" in xml
    assert "<НаличиеНачалоНакопление>3</НаличиеНачалоНакопление>" in xml
