"""Тесты официальных бланков: журнал учёта отходов (Приказ №1028) и
региональный кадастр отходов СПб (Формы 1–5)."""
from decimal import Decimal

import openpyxl

from ecodoc.core import registry
from ecodoc.core.models import (NVOSObject, Organization, ReportContext,
                                ReportPeriod, WasteFlow)


def _ctx():
    return ReportContext(
        organization=Organization(
            name="ООО «Ромашка»", short_name="ООО «Ромашка»", inn="7801234564",
            ogrn="1157847008219", okpo="13884779", okved="68.20.2",
            address="197348, СПб, Богатырский пр., д.2", oktmo="40324000",
            director_name="Иванов И.И.", phone="8(812)000", email="a@b.ru"),
        period=ReportPeriod(year=2024),
        objects=[NVOSObject(code="40-0278-004029-П", name="БЦ Эталон",
                            address="СПб, Богатырский пр., 2", oktmo="40324000",
                            region_code="78")],
        wastes=[
            WasteFlow(fkko_code="47110101521", name="Лампы ртутные", hazard_class=1,
                      generated="0.115", transferred="0.115",
                      origin="Использование по назначению",
                      aggregate_state="Изделия", composition="Стекло 92%"),
            WasteFlow(fkko_code="73310001724", name="Мусор офисный", hazard_class=4,
                      generated="42.6", transferred="42.6"),
        ],
        extra={
            "waste_receivers": [
                {"fkko": "47110101521", "receiver": 'ФГУП "ФЭО"',
                 "contract": "№52274", "contract_term": "по 31.12.2024",
                 "license": "ЛО 020"}],
            "accumulation_sites": [
                {"description": "Контейнерная площадка", "capacity_t": 0.35,
                 "capacity_m3": 2.52, "waste_name": "Мусор", "fkko": "73310001724",
                 "hazard_class": 4}],
        },
    )


def test_waste_movement_1028_sheets(tmp_path):
    registry.load_all()
    rep = registry.get("waste-movement")(_ctx())
    assert rep.has_xml is False
    p = rep.render_print(tmp_path / "j.xlsx")
    wb = openpyxl.load_workbook(p)
    assert wb.sheetnames == ["Титул", "Приложение 1", "Приложение 2 (год)",
                             "Приложение 3 (год)", "Приложение 4 (год)"]
    # Приложение 1 — состав образующихся отходов
    a1 = wb["Приложение 1"]
    assert a1["C4"].value == "Код ФККО"
    assert a1["B6"].value == "Лампы ртутные"
    assert a1["E6"].value == "Использование по назначению"
    # Приложение 2 — движение, образовано в графе 6
    a2 = wb["Приложение 2 (год)"]
    assert a2["G8"].value == 0.115
    # Приложение 3 — получатель переданных отходов
    a3 = wb["Приложение 3 (год)"]
    assert a3["K8"].value == 'ФГУП "ФЭО"'


def test_waste_movement_no_xml(tmp_path):
    registry.load_all()
    rep = registry.get("waste-movement")(_ctx())
    try:
        rep.render_xml(tmp_path / "j.xml")
        assert False, "должно бросить NotImplementedError"
    except NotImplementedError:
        pass


def test_waste_movement_app4_current_edition_13_graf(tmp_path):
    """Таблица 4 (лист «Приложение 4») — по действующей редакции (приказ №825,
    с 01.09.2024) ровно 13 граф: графа «для накопления и последующей
    передачи…» исключена; поставщик — из ctx.extra["waste_suppliers"]."""
    registry.load_all()
    ctx = _ctx()
    ctx.wastes.append(WasteFlow(fkko_code="40140000000", name="Полученный отход",
                                hazard_class=4, received="1.2345678"))
    ctx.extra["waste_suppliers"] = [
        {"fkko": "40140000000", "supplier": "ООО «Источник»",
         "contract": "№7 от 01.02.2024", "contract_term": "по 31.12.2024"}]
    p = registry.get("waste-movement")(ctx).render_print(tmp_path / "j.xlsx")
    a4 = openpyxl.load_workbook(p)["Приложение 4 (год)"]
    assert a4["F7"].value == "для обработки"        # «для накопления…» — нет
    assert a4["J7"].value == "для захоронения"
    assert a4["M8"].value == 13 and a4["N8"].value is None   # граф ровно 13
    assert a4["E9"].value == 1.235                  # тонны — 3 знака (№1028)
    assert a4["K9"].value == "ООО «Источник»"
    assert a4["M9"].value == "по 31.12.2024"


def test_waste_movement_tonnage_three_decimals(tmp_path):
    """Порядок №1028: количество отходов — «в тоннах с точностью до трех
    знаков после запятой»; сырой float с двоичным хвостом в журнал не идёт."""
    registry.load_all()
    ctx = _ctx()
    ctx.wastes[1].generated = Decimal("469.02208921197007477")
    ctx.wastes[1].transferred = Decimal("469.02208921197007477")
    p = registry.get("waste-movement")(ctx).render_print(tmp_path / "j.xlsx")
    wb = openpyxl.load_workbook(p)
    assert wb["Приложение 2 (год)"]["G9"].value == 469.022
    assert wb["Приложение 3 (год)"]["E9"].value == 469.022


def _user_blank_1028(path):
    """Мини-копия реального бланка Б2: чужие строки данных, формулы и
    объединённый блок-продолжение — то, на чём fill() раньше падал."""
    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Титул"
    t["B5"] = "ДАННЫЕ УЧЕТА В ОБЛАСТИ ОБРАЩЕНИЯ С ОТХОДАМИ"
    t.merge_cells("B5:L5")
    t["B9"] = 'ООО "ЧУЖАЯ ФИРМА"'
    t.merge_cells("B9:L9")
    t["E11"] = "период"
    t["F11"] = "1 кв - 4 кв 2025 года"
    t["B15"] = "Ответственный исполнитель"
    t["G15"] = "Чужой И.И."
    ws = wb.create_sheet("Приложение 2 (год)")
    ws["A2"] = "Обобщенные данные учета в области обращения с отходами за"
    ws["G3"] = "=Титул!F11"
    heads = {"A5": "№ п/п", "B5": "Наименование отходов", "C5": "Код ФККО",
             "D5": "Класс опасности вида отхода",
             "E5": "Наличие отходов на начало отчетного периода, тонн",
             "G5": "Образовано отходов в отчетном периоде, тонн",
             "H5": "Получено отходов от других лиц в отчетном периоде, тонн"}
    for ref, txt in heads.items():
        ws[ref] = txt
    for c in "ABCDGH":
        ws.merge_cells(f"{c}5:{c}6")
    ws.merge_cells("E5:F5")
    ws["E6"] = "хранение"
    ws["F6"] = "накопление"
    for i, lbl in enumerate(["А", 1, 2, 3, 4, 5, 6, 7]):
        ws.cell(row=7, column=1 + i, value=lbl)
    # чужие строки данных образца (другая организация)
    for r, (name, code, gen) in {8: ("Чужой отход", "11111111111", 0.115),
                                 9: ("Ещё чужой", "22222222222", 5.8)}.items():
        ws.cell(row=r, column=1, value=r - 7)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=code)
        ws.cell(row=r, column=4, value=4)
        ws.cell(row=r, column=7, value=gen)
    # блок-продолжение: объединённая шапка (раньше запись в неё падала)
    ws["J13"] = "продолжение"
    ws["A14"] = "№ строки"
    ws["F14"] = "Размещено отходов на эксплуатируемых объектах, тонн"
    ws.merge_cells("A14:A15")
    ws.merge_cells("F14:H14")
    ws["F15"] = "Всего"
    ws["G15"] = "Хранение"
    ws["H15"] = "Захоронение"
    for i, lbl in enumerate(["А", 8, 9, 10, 11, 12, 13, 14]):
        ws.cell(row=16, column=1 + i, value=lbl)
    # лист переданных отходов (Таблица 3): фильтр по transferred + контрагент
    a3 = wb.create_sheet("Приложение 3 (год)")
    a3["A2"] = "Данные учета переданных другим лицам отходов за"
    for ref, txt in {"A5": "№ п/п", "B5": "Наименование отходов",
                     "C5": "Код ФККО", "D5": "Класс опасности вида отхода",
                     "E5": "Количество переданных отходов за отчетный период, тонн",
                     "K5": "Сведения о лицах, которым переданы отходы",
                     "L5": "Дата и номер договора на передачу отходов",
                     "M5": "Срок действия договора",
                     "N5": "Реквизиты лицензии на осуществление деятельности"}.items():
        a3[ref] = txt
    for c in "ABCDKLMN":
        a3.merge_cells(f"{c}5:{c}6")
    a3.merge_cells("E5:J5")
    for col, lbl in zip("EFGHIJ", ["Всего", "Для обработки", "Для утилизации",
                                   "Для обезвреживания", "Для хранения",
                                   "Для захоронения"]):
        a3[f"{col}6"] = lbl
    for i in range(14):
        a3.cell(row=7, column=1 + i, value=i + 1)
    a3["K8"] = 'ФГУП "ЧУЖОЙ ПРИЁМЩИК"'      # чужая строка образца
    wb.save(path)
    return path


def test_waste_movement_template_clears_sample_and_expands(tmp_path):
    """Сценарий реального бланка Б2: чужие строки образца очищаются (утечки
    данных другой организации нет), >4 отходов не падают на MergedCell —
    таблица раздвигается перед блоком-продолжением, формулы на месте."""
    from ecodoc.reports.waste_movement.template import fill
    sample = _user_blank_1028(tmp_path / "бланк.xlsx")
    ctx = _ctx()
    ctx.wastes = [WasteFlow(fkko_code=f"4711010152{i}", name=f"Отход {i}",
                            hazard_class=4, generated=str(i))
                  for i in range(1, 8)]          # 7 отходов > 5 свободных строк
    ctx.wastes[2].transferred = Decimal("3.0004")     # только он — в Прил. 3
    ctx.wastes[2].transferred_util = Decimal("3.0004")
    ctx.extra["waste_receivers"] = [
        {"fkko": "47110101523", "receiver": "ООО «Приёмщик»",
         "contract": "№9 от 01.03.2024", "license": "ЛО 020"}]
    out = fill(ctx, tmp_path / "из_бланка.xlsx", sample=sample)
    assert out is not None
    wb = openpyxl.load_workbook(out)
    ws = wb["Приложение 2 (год)"]
    flat = [c.value for row in ws.iter_rows() for c in row]
    assert "Чужой отход" not in flat and 0.115 not in flat    # утечка закрыта
    assert ws["B8"].value == "Отход 1" and ws["B14"].value == "Отход 7"
    assert ws["G3"].value == "=Титул!F11"                     # формула на месте
    # блок-продолжение сдвинут вместе с объединениями, а не затёрт
    assert ws["A16"].value == "№ строки"
    assert "F16:H16" in {str(r) for r in ws.merged_cells.ranges}
    # титул: чужая организация/период/исполнитель заменены на наши
    assert wb["Титул"]["B9"].value.startswith("ООО «Ромашка»")
    # период без «за»: в шапках листов оно уже есть («…отходами за» + =Титул!F11)
    assert wb["Титул"]["F11"].value == "2024 года"
    assert wb["Титул"]["G15"].value == "Иванов И.И."
    # Прил. 3: только переданные отходы, чужой приёмщик заменён на нашего
    a3 = wb["Приложение 3 (год)"]
    assert a3["B8"].value == "Отход 3" and a3["B9"].value is None
    assert a3["E8"].value == 3.0 and a3["G8"].value == 3.0    # 3.0004 → 3 знака
    assert a3["K8"].value == "ООО «Приёмщик»"
    assert a3["L8"].value == "№9 от 01.03.2024"
    assert a3["N8"].value == "ЛО 020"


def test_waste_movement_blank_failure_not_silent(tmp_path, monkeypatch):
    """Сбой заполнения бланка больше не глотается молча: форма рисуется
    кодом, а в render_issues появляется предупреждение про бланк."""
    registry.load_all()
    import ecodoc.reports.waste_movement.template as tmpl

    def boom(*a, **k):
        raise RuntimeError("'MergedCell' object attribute 'value' is read-only")

    monkeypatch.setattr(tmpl, "fill", boom)
    rep = registry.get("waste-movement")(_ctx())
    p = rep.render_print(tmp_path / "j.xlsx")
    wb = openpyxl.load_workbook(p)
    assert "Приложение 2 (год)" in wb.sheetnames    # форма нарисована кодом
    warns = [i for i in rep.render_issues if i.level == "warning"]
    assert warns and warns[0].field == "бланк"
    assert "не применён" in warns[0].message


def test_cadastre_spb_forms(tmp_path):
    registry.load_all()
    rep = registry.get("cadastre-spb")(_ctx())
    assert rep.has_xml is False
    assert not [i for i in rep.validate() if i.level == "error"]
    p = rep.render_print(tmp_path / "k.xlsx")
    wb = openpyxl.load_workbook(p)
    assert wb.sheetnames == ["Форма 1", "Форма 2", "Форма 3", "Форма 4", "Форма 5"]
    assert wb["Форма 1"]["D6"].value == "7801234564"     # ИНН
    assert wb["Форма 2"]["C8"].value == "Контейнерная площадка"
    f3 = wb["Форма 3"]
    assert f3["B13"].value == "Лампы ртутные"             # наименование в графе 2
    assert f3["C13"].value == "4 71 101 01 52 1"   # код в графе 3 — формат ФККО
    assert f3["D13"].value == "I"                  # класс римской, как в принятом отчёте
    # мусор офисный 7 33 100 01 72 4 без записи о приёмщике — резервный
    # признак ТКО по коду, масса в гр.17 (графа Q)
    assert f3["Q14"].value == 42.6
    # гр.20/21 (T/U) при передаче не пустые: неизвестное — «-», как в эталоне
    assert f3["T13"].value == "-" and f3["U13"].value == "-"
    assert f3["S13"].value == 0.0
    assert wb["Форма 5"]["C27"].value == "ООО «Ромашка»"


def test_cadastre_spb_form3_split_by_receiver(tmp_path):
    """Гр.17/18 Формы 3 делятся по ФАКТИЧЕСКОМУ приёмщику (как в принятом
    отчёте), а не по «весь блок 7 3 = ТКО»; гр.19-21 — из waste_receivers."""
    registry.load_all()
    ctx = _ctx()
    ctx.wastes = [
        # производственный мусор 7 33 2... передан НЕ регоператору → гр.18,
        # приёмщик в ЛО → гр.19 масса + гр.20 субъект, цель из operation
        WasteFlow(fkko_code="73321001724", name="Мусор производственных помещений",
                  hazard_class=4, generated="133.95", transferred="133.95"),
        # офисный мусор передан регоператору (НЭО) → гр.17, гр.18=0
        WasteFlow(fkko_code="73310001724", name="Мусор офисный", hazard_class=4,
                  generated="2.672", transferred="2.672"),
    ]
    ctx.extra["waste_receivers"] = [
        {"fkko": "73321001724",
         "receiver": "АО «УК по обращению с отходами в Ленинградской области»",
         "license": "Л020-00113-47/00038311", "operation": "обработка"},
        {"fkko": "73310001724", "receiver": "АО «Невский экологический оператор»",
         "operation": "обработка"},
    ]
    p = registry.get("cadastre-spb")(ctx).render_print(tmp_path / "k3.xlsx")
    f3 = openpyxl.load_workbook(p)["Форма 3"]
    # строка 13: 7 33 2... — прочим (гр.18), в другой субъект РФ
    assert f3["Q13"].value == 0.0 and f3["R13"].value == 133.95
    assert f3["S13"].value == 133.95
    assert f3["T13"].value == "Ленинградская область"
    assert f3["U13"].value == "обработка"
    # строка 14: регоператор ТКО → гр.17, гр.18/19 = 0
    assert f3["Q14"].value == 2.672 and f3["R14"].value == 0.0
    assert f3["S14"].value == 0.0


def test_cadastre_spb_form2_no_sites_placeholder(tmp_path):
    """Без данных о местах накопления Форма 2 печатает прочерки (не «…»)
    и validate() предупреждает, что форма уйдёт пустой."""
    registry.load_all()
    ctx = _ctx()
    ctx.extra.pop("accumulation_sites")
    rep = registry.get("cadastre-spb")(ctx)
    warns = [i for i in rep.validate() if i.level == "warning"]
    assert any("накоплен" in i.message for i in warns)
    p = rep.render_print(tmp_path / "k2.xlsx")
    f2 = openpyxl.load_workbook(p)["Форма 2"]
    assert f2["C8"].value == "-"
    assert all(f2[f"{c}8"].value == "-" for c in "BCDEFGHI")


def test_cadastre_spb_form2_site_with_waste_list(tmp_path):
    """Место накопления со списком wastes раскладывается как в принятом
    отчёте: описание/вместимость один раз, по строке на каждый отход."""
    registry.load_all()
    ctx = _ctx()
    ctx.extra["accumulation_sites"] = [
        {"description": "Металлический контейнер", "capacity_t": 15.0,
         "capacity_m3": 27.0, "wastes": [
             {"waste_name": "Мусор офисный", "fkko": "73310001724",
              "hazard_class": 4, "method": "в смеси с другими отходами"},
             {"name": "Смет с территории", "fkko": "73339002715",
              "hazard_class": 5}]},
        {"description": "Контейнер пластмассовый", "capacity_t": 0.352,
         "capacity_m3": 2.2, "waste_name": "Лампы ртутные",
         "fkko": "47110101521", "hazard_class": 1},
    ]
    p = registry.get("cadastre-spb")(ctx).render_print(tmp_path / "k2w.xlsx")
    f2 = openpyxl.load_workbook(p)["Форма 2"]
    # место 1: две строки отходов, описание и вместимость только в первой
    assert (f2["B8"].value, f2["C8"].value, f2["D8"].value) == \
        (1, "Металлический контейнер", 15.0)
    assert f2["G8"].value == "7 33 100 01 72 4"
    assert f2["H8"].value == "IV"
    assert f2["I8"].value == "в смеси с другими отходами"
    # пустая строка при чтении openpyxl видна как None — важно, что не число/текст
    assert (f2["B9"].value, f2["C9"].value, f2["D9"].value) == (None, None, None)
    assert f2["F9"].value == "Смет с территории"     # поле name — тоже понимаем
    assert f2["I9"].value == "отдельно от других отходов"
    # место 2 (плоская запись) — следующая строка со своим номером
    assert (f2["B10"].value, f2["C10"].value) == (2, "Контейнер пластмассовый")
    assert f2["G10"].value == "4 71 101 01 52 1"


def test_cadastre_spb_form4_two_tier_header(tmp_path):
    """Форма 4 — двухъярусная шапка с групповыми заголовками (merge) и
    полными подписями граф, как в принятом Комитетом отчёте."""
    registry.load_all()
    p = registry.get("cadastre-spb")(_ctx()).render_print(tmp_path / "k4.xlsx")
    f4 = openpyxl.load_workbook(p)["Форма 4"]
    merged = {str(r) for r in f4.merged_cells.ranges}
    assert {"N5:P5", "Q5:R5", "X5:AA5", "A5:A8", "N6:N8"} <= merged
    assert f4["N5"].value == "Используемые установки (необходимое оборудование)"
    assert f4["Q5"].value == "Получение вторичной продукции (энергии)"
    assert f4["X5"].value == "Образование вторичных отходов"
    # подписи граф без склейки «Группа: подпись» и без многоточий-заглушек
    assert f4["N6"].value == "Наименование"
    assert "…" not in (f4["S5"].value or "")
    assert "телефон, факс, интернет-сайт разработчика" in f4["S5"].value
    assert "транспортировке" in f4["G5"].value
    assert f4["G5"].value.endswith("и наименование органа, выдавшего ее")
    assert "в случае отсутствия кода по ОКПД" in f4["Q6"].value
    assert "с указанием единицы измерения" in f4["R6"].value
    # номера граф в строке 9, данные (прочерки без treatment_objects) — в 10,
    # как в типовой форме Комитета
    assert f4["A9"].value == 1 and f4["AA9"].value == 27
    assert f4["A10"].value == "-"


def test_tp2_waste_datapacket_xml(tmp_path):
    registry.load_all()
    rep = registry.get("2tp-waste")(_ctx())
    p = rep.render_xml(tmp_path / "2tp.xml")
    xml = p.read_text(encoding="utf-8")
    # реальный конверт Модуля природопользователя
    assert "<DATA_PACKET_NI" in xml and 'DocType="3"' in xml
    assert "<ORG_INFO>" in xml and "<EMISS_OBJECT>" in xml
    assert "<RPT_2TP_WASTE>" in xml and "<RPT_2TP_WASTE_FACT>" in xml
    assert "<WST_CODE>73310001724</WST_CODE>" in xml
    # масса IV класса — с точностью 1 знак (Указания к форме)
    assert "<TP2_FORMING>42.6</TP2_FORMING>" in xml
    assert "<CHECKSUM>" in xml


# --- журнал 1028: блок «продолжение», межлистовые ссылки, период, ред. № 227 ---

def _user_blank_1028_full(path):
    """Мини-копия бланка МО 2025: Прил. 1 + Прил. 2 с ДВУМЯ блоками (шапка
    продолжения объединена, строка 16 — чужой образец с формулой =G8) +
    Прил. 3, где первая строка — межлистовые ссылки на строку 6 Прил. 1."""
    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Титул"
    t["B5"] = "ДАННЫЕ УЧЕТА В ОБЛАСТИ ОБРАЩЕНИЯ С ОТХОДАМИ"
    t["B9"] = 'ООО "ЧУЖАЯ ФИРМА"\nПлощадка по адресу: "чужой адрес"'
    t["E11"] = "период"
    t["F11"] = "1 кв - 3 кв 2025 года"
    a1 = wb.create_sheet("Приложение 1")
    a1["G1"] = ("Приложение N 1\nк Порядку учета в области обращения с отходами, "
                "утвержденному приказом Минприроды России от 8 декабря 2020 года N 1028")
    a1["A2"] = "Состав образующихся видов отходов, подлежащих учету"
    for i, h in enumerate(["№ п/п", "Наименование отходов", "Код ФККО",
                           "Класс опасности вида отхода", "Происхождение",
                           "Агрегатное состояние", "Химический состав"], 1):
        a1.cell(row=4, column=i, value=h)
        a1.cell(row=5, column=i, value=i)
    a1["A6"], a1["B6"], a1["C6"], a1["D6"] = 3, "Чужой отход", "7 33 100 01 72 4", 4
    ws = wb.create_sheet("Приложение 2 (год)")
    ws["A2"] = "Обобщенные данные учета в области обращения с отходами за"
    ws["G3"], ws["H3"] = "=Титул!F11", "год"
    for ref, txt in {"A5": "№ п/п", "B5": "Наименование отходов", "C5": "Код ФККО",
                     "D5": "Класс опасности вида отхода",
                     "E5": "Наличие отходов на начало отчетного периода, тонн",
                     "G5": "Образовано отходов в отчетном периоде, тонн",
                     "H5": "Получено отходов от других лиц в отчетном периоде, тонн"}.items():
        ws[ref] = txt
    for c in "ABCDGH":
        ws.merge_cells(f"{c}5:{c}6")
    ws.merge_cells("E5:F5")
    ws["E6"], ws["F6"] = "хранение", "накопление"
    for i, lbl in enumerate(["А", 1, 2, 3, 4, 5, 6, 7]):
        ws.cell(row=7, column=1 + i, value=lbl)
    ws["A8"], ws["B8"] = 1, "='Приложение 1'!B6"
    ws["C8"], ws["D8"] = "='Приложение 1'!C6", "='Приложение 1'!D6"
    ws["G8"] = 0
    ws["J12"] = "продолжение"
    ws["A13"] = "№ строки"
    ws["B13"] = "Обработано отходов в отчетном периоде, тонн"
    ws["C13"] = "Утилизировано отходов в отчетном периоде, тонн"
    ws["D13"] = "Обезврежено отходов в отчетном периоде, тонн"
    ws["E13"] = "Передано отходов за отчетный период, тонн"
    ws["F13"] = "Размещено отходов на эксплуатируемых объектах в отчетном периоде, тонн"
    ws["I13"] = "Наличие отходов на конец отчетного периода, тонн"
    for c in "ABCDE":
        ws.merge_cells(f"{c}13:{c}14")
    ws.merge_cells("F13:H13")
    ws.merge_cells("I13:J13")
    ws["F14"], ws["G14"], ws["H14"] = "Всего", "Хренение", "Захоронение"  # опечатка как в бланке
    ws["I14"], ws["J14"] = "Хранение", "Накопление"
    for i, lbl in enumerate(["А", 8, 9, 10, 11, 12, 13, 14, 15, 16]):
        ws.cell(row=15, column=1 + i, value=lbl)
    ws["A16"], ws["B16"], ws["E16"], ws["F16"] = 1, 0, "=G8", 0   # чужая строка образца
    a3 = wb.create_sheet("Приложение 3 (год)")
    a3["A2"] = "Данные учета переданных другим лицам отходов за"
    a3["L3"], a3["M3"] = "=Титул!F11", "год"
    for ref, txt in {"A5": "№ п/п", "B5": "Наименование отходов", "C5": "Код ФККО",
                     "D5": "Класс опасности вида отхода",
                     "E5": "Количество переданных отходов за отчетный период, тонн",
                     "K5": "Сведения о лицах, которым переданы отходы",
                     "L5": "Дата и номер договора", "M5": "Срок действия договора",
                     "N5": "Реквизиты лицензии"}.items():
        a3[ref] = txt
    for c in "ABCDKLMN":
        a3.merge_cells(f"{c}5:{c}6")
    a3.merge_cells("E5:J5")
    for col, lbl in zip("EFGHIJ", ["Всего", "Для обработки", "Для утилизации",
                                   "Для обезвреживания", "Для хранения", "Для захоронения"]):
        a3[f"{col}6"] = lbl
    for i in range(14):
        a3.cell(row=7, column=1 + i, value=i + 1)
    a3["A8"], a3["B8"] = "='Приложение 1'!A6", "='Приложение 1'!B6"
    a3["C8"], a3["D8"] = "='Приложение 1'!C6", "='Приложение 1'!D6"
    a3["E8"], a3["J8"] = "=SUM(F8:J8)", "='Приложение 2 (год)'!E16"
    a3["K8"] = 'ООО "ЧУЖОЙ ПРИЁМЩИК"'
    wb.save(path)
    return path


def _ctx3():
    """Три отхода: передаётся только третий (проверка нумерации Прил. 3)."""
    ctx = _ctx()
    ctx.wastes = [
        WasteFlow(fkko_code="47110101521", name="Лампы ртутные", hazard_class=1,
                  generated="0.115", accumulated_end="0.115"),
        WasteFlow(fkko_code="73310001724", name="Мусор офисный", hazard_class=4,
                  generated="42.6", used="2.6", placed_norm="40"),
        WasteFlow(fkko_code="92130201523", name="Обтирочный материал", hazard_class=4,
                  generated="0.111", processed="0.011", transferred="0.1",
                  transferred_neutral="0.1"),
    ]
    ctx.extra["waste_receivers"] = [
        {"fkko": "92130201523", "receiver": "ООО «Приёмщик»", "contract": "№9"}]
    ctx.period = ReportPeriod(year=2025, quarter=1)
    return ctx


def test_waste_movement_blank_continuation_block_filled(tmp_path):
    """[critical] Блок «продолжение» Прил. 2 (графы 8–16) заполняется по тем
    же строкам, что и первая часть; чужая строка образца (=G8) затёрта;
    подграфы «хранение/накопление» различаются по родительской графе."""
    from ecodoc.reports.waste_movement.template import fill
    sample = _user_blank_1028_full(tmp_path / "бланк.xlsx")
    out = fill(_ctx3(), tmp_path / "из_бланка.xlsx", sample=sample)
    ws = openpyxl.load_workbook(out)["Приложение 2 (год)"]
    assert [ws[f"A{r}"].value for r in (16, 17, 18)] == [1, 2, 3]   # № строки
    assert ws["E16"].value is None and ws["I16"].value == 0.115    # =G8 затёрт; остаток
    assert ws["C17"].value == 2.6                                  # утилизировано
    assert ws["F17"].value == 40.0 and ws["H17"].value == 40.0     # размещено всего/захоронение
    assert ws["B18"].value == 0.011 and ws["E18"].value == 0.1     # обработано / передано
    flat = [c.value for row in ws.iter_rows() for c in row]
    assert not any(isinstance(v, str) and v.startswith("='Приложение") for v in flat)


def test_waste_movement_blank_cross_sheet_refs_replaced(tmp_path):
    """[critical] Межлистовые ссылки первой строки Прил. 3 (на отход №1 из
    Прил. 1, «для захоронения» = E16 Прил. 2) заменены значениями; SUM по
    своей строке оставлен Excel; номер отхода — как в Прил. 1 (3, не 1)."""
    from ecodoc.reports.waste_movement.template import fill
    sample = _user_blank_1028_full(tmp_path / "бланк.xlsx")
    out = fill(_ctx3(), tmp_path / "из_бланка.xlsx", sample=sample)
    wb = openpyxl.load_workbook(out)
    a3 = wb["Приложение 3 (год)"]
    assert a3["A8"].value == 3 and a3["B8"].value == "Обтирочный материал"
    assert a3["C8"].value == "9 21 302 01 52 3"
    assert a3["E8"].value == "=SUM(F8:J8)"                 # агрегат по строке
    assert a3["H8"].value == 0.1 and a3["J8"].value is None   # обезвреживание, не захоронение
    assert a3["K8"].value == "ООО «Приёмщик»" and a3["B9"].value is None
    a2 = wb["Приложение 2 (год)"]
    assert a2["B8"].value == "Лампы ртутные" and a2["C8"].value == "4 71 101 01 52 1"


def test_waste_movement_blank_period_and_edition_825(tmp_path):
    """[major] Период из ctx.period без «за» («1 кв 2025 года»), слово «год»
    рядом с =Титул!F11 убрано; реквизит приказа дополнен ред. № 825,
    Таблица 1 — «Перечень…»; титул — «Площадка по адресу» + ИНН/ОГРН."""
    from ecodoc.reports.waste_movement.template import fill
    sample = _user_blank_1028_full(tmp_path / "бланк.xlsx")
    out = fill(_ctx3(), tmp_path / "из_бланка.xlsx", sample=sample)
    wb = openpyxl.load_workbook(out)
    assert wb["Титул"]["F11"].value == "1 кв 2025 года"
    a2 = wb["Приложение 2 (год)"]
    assert a2["G3"].value == "=Титул!F11" and a2["H3"].value is None
    assert wb["Приложение 3 (год)"]["M3"].value is None
    a1 = wb["Приложение 1"]
    assert a1["A2"].value == "Перечень образующихся видов отходов, подлежащих учету"
    assert "N 825" in a1["G1"].value and a1["G1"].value.startswith("Приложение N 1\n")
    title = wb["Титул"]["B9"].value
    assert 'Площадка по адресу: "СПб, Богатырский пр., 2"' in title
    assert "ИНН 7801234564" in title and "ОГРН 1157847008219" in title


def test_waste_movement_edition_227(tmp_path):
    """[major] С 01.09.2026 — приказ № 227: реквизит, графы 8/13 Табл. 2 о
    собственных объектах, 4 знака при массе < 0,001 т, лист полученных —
    самостоятельное Приложение N 4."""
    from ecodoc.reports.waste_movement import template as tmpl
    registry.load_all()
    ctx = _ctx3()
    ctx.period = ReportPeriod(year=2026, quarter=3)
    assert tmpl.edition(ctx) == "227"
    ctx.period = ReportPeriod(year=2026, quarter=2)
    assert tmpl.edition(ctx) == "825"          # 2 кв. 2026 обобщается в июле
    ctx.extra["report_date"] = "05.10.2026"    # а с датой составления — по ней
    assert tmpl.edition(ctx) == "227"
    ctx.period = ReportPeriod(year=2026, quarter=3)
    ctx.wastes[0].generated = Decimal("0.0004")
    ctx.extra["own_transfers"] = [{"fkko": "73310001724", "transferred_own": "1.5"}]
    ctx.extra["period_text"] = "9 месяцев 2026 года"
    p = registry.get("waste-movement")(ctx).render_print(tmp_path / "j.xlsx")
    wb = openpyxl.load_workbook(p)
    a2 = wb["Приложение 2 (год)"]
    assert "N 227" in a2["I1"].value and "производства и потребления" in a2["I1"].value
    assert a2["I5"].value.startswith("Поступление отходов с собственных объектов")
    assert a2["I7"].value == 8
    assert a2["G8"].value == 0.0004                   # 4 знака при массе < 0,001 т
    assert a2["G3"].value == "9 месяцев 2026 года"
    # 3 отхода (строки 8–10) + 3 строки отступа → шапка продолжения в 14–15,
    # нумерация граф 9–18 в строке 16, данные с 17
    assert a2["F14"].value.startswith("Передача отходов (за исключением ТКО)")
    assert [a2[f"{c}16"].value for c in "ABCDEFGHIJK"] == ["А"] + list(range(9, 19))
    assert a2["F18"].value == 1.5                      # передача на собственные объекты
    assert wb["Приложение 4 (год)"]["M1"].value.startswith("Приложение N 4\n")
    # а в ред. № 825 лист полученных — Таблица 4 Приложения N 3
    ctx.period, ctx.extra["report_date"] = ReportPeriod(year=2025), ""
    ctx.extra.pop("period_text")
    p = registry.get("waste-movement")(ctx).render_print(tmp_path / "j2.xlsx")
    wb = openpyxl.load_workbook(p)
    assert wb["Приложение 4 (год)"]["M1"].value.startswith("Приложение N 3 (Таблица 4)\n")
    assert wb["Приложение 2 (год)"]["G8"].value == 0.0    # 3 знака: 0,0004 → 0
    assert wb["Приложение 3 (год)"]["A8"].value == 3      # № отхода из Прил. 1
    assert wb["Приложение 1"]["A2"].value == "Перечень образующихся видов отходов, подлежащих учету"
    assert wb["Титул"]["F11"].value == "за 2025 года"
    assert wb["Титул"]["B10"].value == "ИНН 7801234564, ОГРН 1157847008219"


def test_waste_accounting_calendar_monthly_last_day():
    """[major] Календарь: журнал — ежемесячно, срок — последний день
    следующего месяца (№ 1028 ред. № 825 п. 11; № 227 п. 13), а не 10-е."""
    from ecodoc.calendar.obligations import OBLIGATIONS
    o = next(x for x in OBLIGATIONS if x.code == "waste-accounting")
    assert o.periodicity == "месяц"
    assert len(o.due) == 12 and (1, 31) in o.due and (2, 28) in o.due and (4, 30) in o.due
    assert "полугодие" in o.coverage and "последнего дня" in o.coverage
